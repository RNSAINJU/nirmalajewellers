from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import models
from django.db.models import IntegerField, Q, Sum, F, Case, When, DecimalField
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, DeleteView, ListView, UpdateView
from django.utils.decorators import method_decorator

import openpyxl
from openpyxl.utils import get_column_letter

from common.nepali_utils import ndt
from .forms import CustomerPurchaseForm, MetalStockForm
from .models import (
    CustomerPurchase,
    GoldSilverPurchase,
    MetalStock,
    MetalStockMovement,
    MetalStockType,
    Party,
)
from ornament.models import Kaligar, MainCategory, Ornament, SubCategory
from order.models import Order, OrderMetalStock, OrderOrnament, OrderPayment
from sales.models import Sale, SalesMetalStock

# Decimal alias used in a few imports
D = Decimal

@login_required(login_url='/accounts/login/')
def export_metalstock_xlsx(request):
    wb = openpyxl.Workbook()
    ws_stock = wb.active
    ws_stock.title = 'MetalStock'
    ws_movement = wb.create_sheet(title='MetalStockMovement')

    # MetalStock sheet
    stock_headers = [
        'StockID', 'MetalType', 'StockType', 'Purity', 'Quantity', 'UnitCost', 'RateUnit', 'TotalCost', 'Location', 'Remarks', 'CreatedAt', 'LastUpdated'
    ]
    ws_stock.append(stock_headers)
    # OPTIMIZED: Add select_related to avoid N+1 queries
    for stock in MetalStock.objects.select_related('stock_type').all():
        ws_stock.append([
            stock.id,
            stock.metal_type,
            stock.stock_type.name if stock.stock_type else '',
            stock.purity,
            float(stock.quantity) if stock.quantity is not None else '',
            float(stock.unit_cost) if stock.unit_cost is not None else '',
            stock.rate_unit,
            float(stock.total_cost) if stock.total_cost is not None else '',
            stock.location or '',
            stock.remarks or '',
            str(stock.created_at) if stock.created_at else '',
            str(stock.last_updated) if stock.last_updated else ''
        ])

    # MetalStockMovement sheet
    movement_headers = [
        'MovementID', 'MetalStockID', 'MovementType', 'Quantity', 'Rate', 'ReferenceType', 'ReferenceID', 'Notes', 'MovementDate', 'CreatedAt'
    ]
    ws_movement.append(movement_headers)
    # OPTIMIZED: Add select_related to avoid N+1 queries
    for m in MetalStockMovement.objects.select_related('metal_stock').all():
        ws_movement.append([
            m.id,
            m.metal_stock.id if m.metal_stock else '',
            m.movement_type,
            float(m.quantity) if m.quantity is not None else '',
            float(m.rate) if m.rate is not None else '',
            m.reference_type or '',
            m.reference_id or '',
            m.notes or '',
            str(m.movement_date) if m.movement_date else '',
            str(m.created_at) if m.created_at else ''
        ])

    # Auto-size columns for both sheets
    for ws in [ws_stock, ws_movement]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    from io import BytesIO
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=metalstock_export.xlsx'
    wb.save(response)
    return response

@method_decorator(csrf_exempt, name='dispatch')
class ImportMetalStockXLSXView(LoginRequiredMixin, View):
    def post(self, request):
        from datetime import date, datetime
        file = request.FILES.get('file')
        if not file:
            return HttpResponse('No file uploaded.', status=400)
        wb = openpyxl.load_workbook(file)
        from decimal import Decimal
        from django.utils.dateparse import parse_date
        # --- Import MetalStock sheet ---
        ws_stock = wb['MetalStock'] if 'MetalStock' in wb.sheetnames else wb.active
        stock_rows = list(ws_stock.iter_rows(min_row=2, values_only=True))
        stock_objs = {}
        errors = []
        for idx, row in enumerate(stock_rows, start=2):
            try:
                stock_id, metal_type, stock_type_name, purity, quantity, unit_cost, rate_unit, total_cost, location, remarks, created_at, last_updated = row
                stock_type = MetalStockType.objects.filter(name=stock_type_name).first() if stock_type_name else None
                stock, _ = MetalStock.objects.update_or_create(
                    id=stock_id,
                    defaults={
                        'metal_type': metal_type or '',
                        'stock_type': stock_type,
                        'purity': purity or '',
                        'quantity': Decimal(quantity or 0),
                        'unit_cost': Decimal(unit_cost or 0),
                        'rate_unit': rate_unit or '',
                        'total_cost': Decimal(total_cost or 0),
                        'location': location or '',
                        'remarks': remarks or '',
                    }
                )
                stock_objs[stock_id] = stock
            except Exception as e:
                errors.append(f"Stock Row {idx}: {str(e)}")
        # --- Import MetalStockMovement sheet ---
        if 'MetalStockMovement' in wb.sheetnames:
            ws_movement = wb['MetalStockMovement']
            movement_rows = list(ws_movement.iter_rows(min_row=2, values_only=True))
            for idx, row in enumerate(movement_rows, start=2):
                try:
                    movement_id, metal_stock_id, movement_type, quantity, rate, reference_type, reference_id, notes, movement_date, created_at = row
                    stock = stock_objs.get(metal_stock_id) or MetalStock.objects.filter(id=metal_stock_id).first()
                    if not stock:
                        errors.append(f"Movement Row {idx}: MetalStockID {metal_stock_id} not found.")
                        continue
                    MetalStockMovement.objects.update_or_create(
                        id=movement_id,
                        defaults={
                            'metal_stock': stock,
                            'movement_type': movement_type or '',
                            'quantity': Decimal(quantity or 0),
                            'rate': Decimal(rate or 0),
                            'reference_type': reference_type or '',
                            'reference_id': reference_id or '',
                            'notes': notes or '',
                            'movement_date': parse_date(movement_date) if isinstance(movement_date, str) and movement_date else movement_date if isinstance(movement_date, (date, datetime)) else None,
                        }
                    )
                except Exception as e:
                    errors.append(f"Movement Row {idx}: {str(e)}")
        if errors:
            return HttpResponse('Import completed with errors:\n' + '\n'.join(errors), status=400)
        return HttpResponse('Import completed.')


class PurchaseListView(LoginRequiredMixin, ListView):
    model = GoldSilverPurchase
    template_name = 'goldsilverpurchase/purchase_list.html'
    context_object_name = 'purchases'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('party')
        date_str = self.request.GET.get('date')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        party_id = self.request.GET.get('party')
        metal_type = self.request.GET.get('metal_type')

        if date_str:
            try:
                y, m, d = map(int, str(date_str).split('-'))
                if ndt:
                    queryset = queryset.filter(bill_date=ndt.date(y, m, d))
            except Exception:
                pass
        elif start_date_str and end_date_str:
            try:
                y1, m1, d1 = map(int, str(start_date_str).split('-'))
                y2, m2, d2 = map(int, str(end_date_str).split('-'))
                if ndt:
                    queryset = queryset.filter(
                        bill_date__range=(ndt.date(y1, m1, d1), ndt.date(y2, m2, d2))
                    )
            except Exception:
                pass

        if party_id:
            queryset = queryset.filter(party_id=party_id)

        if metal_type:
            queryset = queryset.filter(metal_type=metal_type)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        
        # OPTIMIZED: Combine multiple aggregates into a single query
        aggregates = qs.aggregate(
            total_quantity=Sum('quantity'),
            total_amount=Sum('amount')
        )
        context['total_quantity'] = aggregates['total_quantity'] or Decimal('0')
        context['total_amount'] = aggregates['total_amount'] or Decimal('0')

        # OPTIMIZED: Combine gold and silver purchases into a single query with conditional aggregation
        from django.db.models import Case, When, DecimalField
        metal_aggregates = GoldSilverPurchase.objects.aggregate(
            gold_total=Sum(
                Case(
                    When(metal_type__icontains='gold', then=F('quantity')),
                    default=0,
                    output_field=DecimalField()
                )
            ),
            silver_total=Sum(
                Case(
                    When(metal_type__icontains='silver', then=F('quantity')),
                    default=0,
                    output_field=DecimalField()
                )
            )
        )
        context['gold_purchase'] = metal_aggregates['gold_total'] or 0
        context['silver_purchase'] = metal_aggregates['silver_total'] or 0

        context['date'] = self.request.GET.get('date', '')
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['parties'] = Party.objects.all()
        context['selected_party'] = self.request.GET.get('party', '')
        context['metal_type'] = self.request.GET.get('metal_type', '')
        return context
class PurchaseCreateView(LoginRequiredMixin, CreateView):
    model = GoldSilverPurchase
    fields = [
        'bill_no', 'bill_date', 'party',
        'particular', 'metal_type', 'purity', 'quantity',
        'rate', 'rate_unit', 'wages', 'discount','amount', 'payment_mode'
    ]
    template_name = 'goldsilverpurchase/purchase_form.html'
    success_url = reverse_lazy('gsp:purchaselist')


class PurchaseUpdateView(LoginRequiredMixin, UpdateView):
    model = GoldSilverPurchase
    fields = [
        'bill_no', 'bill_date', 'party',
        'particular', 'metal_type', 'purity', 'quantity',
        'rate', 'rate_unit', 'wages','discount', 'amount', 'payment_mode',
        'is_paid', 'remarks'
    ]
    template_name = 'goldsilverpurchase/purchase_form.html'
    success_url = reverse_lazy('gsp:purchaselist')


class PurchaseDeleteView(LoginRequiredMixin, DeleteView):
    model = GoldSilverPurchase
    template_name = 'goldsilverpurchase/purchase_confirm_delete.html'
    success_url = reverse_lazy('gsp:purchaselist')


# ---------------------- Party Management ----------------------
class PartyCreateView(LoginRequiredMixin, CreateView):
    model = Party
    fields = ['party_name', 'panno']
    template_name = 'goldsilverpurchase/party_form.html'
    success_url = reverse_lazy('gsp:purchaselist')


class PartyUpdateView(LoginRequiredMixin, UpdateView):
    model = Party
    fields = ['party_name', 'panno']
    template_name = 'goldsilverpurchase/party_form.html'
    success_url = reverse_lazy('gsp:purchaselist')


class PartyDeleteView(LoginRequiredMixin, DeleteView):
    model = Party
    template_name = 'goldsilverpurchase/party_confirm_delete.html'
    success_url = reverse_lazy('gsp:purchaselist')


# ---------------------- Customer Purchases ----------------------
class CustomerPurchaseListView(LoginRequiredMixin, ListView):
    model = CustomerPurchase
    template_name = 'goldsilverpurchase/customer_purchase_list.html'
    context_object_name = 'customer_purchases'
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.GET.get('search')
        date_str = self.request.GET.get('date')
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        metal_type = self.request.GET.get('metal_type')

        if search:
            qs = qs.filter(
                Q(sn__icontains=search) |
                Q(customer_name__icontains=search) |
                Q(location__icontains=search) |
                Q(phone_no__icontains=search) |
                Q(ornament_name__icontains=search)
            )

        if date_str:
            try:
                y, m, d = map(int, date_str.split('-'))
                qs = qs.filter(purchase_date=ndt.date(y, m, d))
            except ValueError:
                pass
        elif start_date_str and end_date_str:
            try:
                y1, m1, d1 = map(int, start_date_str.split('-'))
                y2, m2, d2 = map(int, end_date_str.split('-'))
                qs = qs.filter(purchase_date__range=(ndt.date(y1, m1, d1), ndt.date(y2, m2, d2)))
            except ValueError:
                pass

        if metal_type:
            qs = qs.filter(metal_type=metal_type)

        # Order by SN in descending numeric order
        qs = qs.annotate(sn_numeric=Cast('sn', IntegerField())).order_by('-sn_numeric')
        
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        ctx['total_weight'] = qs.aggregate(total=Sum('weight'))['total'] or Decimal('0')
        ctx['total_refined_weight'] = qs.aggregate(total=Sum('refined_weight'))['total'] or Decimal('0')
        ctx['total_final_weight'] = qs.aggregate(total=Sum('final_weight'))['total'] or Decimal('0')
        ctx['total_profit_weight'] = qs.aggregate(total=Sum('profit_weight'))['total'] or Decimal('0')
        ctx['total_amount'] = qs.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        ctx['total_profit'] = qs.aggregate(total=Sum('profit'))['total'] or Decimal('0')
        ctx['date'] = self.request.GET.get('date', '')
        ctx['start_date'] = self.request.GET.get('start_date', '')
        ctx['end_date'] = self.request.GET.get('end_date', '')
        ctx['search'] = self.request.GET.get('search', '')
        ctx['metal_type'] = self.request.GET.get('metal_type', '')
        
        # Separate gold and silver for tabs
        ctx['gold_purchases'] = qs.filter(metal_type='gold')
        ctx['silver_purchases'] = qs.filter(metal_type='silver')
        
        # Calculate stats for each metal type
        ctx['gold_total_weight'] = ctx['gold_purchases'].aggregate(total=Sum('weight'))['total'] or Decimal('0')
        ctx['gold_total_amount'] = ctx['gold_purchases'].aggregate(total=Sum('amount'))['total'] or Decimal('0')
        ctx['gold_total_profit'] = ctx['gold_purchases'].aggregate(total=Sum('profit'))['total'] or Decimal('0')
        
        ctx['silver_total_weight'] = ctx['silver_purchases'].aggregate(total=Sum('weight'))['total'] or Decimal('0')
        ctx['silver_total_amount'] = ctx['silver_purchases'].aggregate(total=Sum('amount'))['total'] or Decimal('0')
        ctx['silver_total_profit'] = ctx['silver_purchases'].aggregate(total=Sum('profit'))['total'] or Decimal('0')
        
        return ctx


class CustomerPurchaseCreateView(LoginRequiredMixin, CreateView):
    model = CustomerPurchase
    form_class = CustomerPurchaseForm
    template_name = 'goldsilverpurchase/customer_purchase_form.html'
    success_url = reverse_lazy('gsp:customer_purchase_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = 'Create Customer Purchase'
        return ctx

    def form_valid(self, form):
        """Handle form submission - metal stock is managed by signals"""
        response = super().form_valid(form)
        purchase = self.object
        
        # Determine which weight was used based on refined_status
        if purchase.refined_status == 'yes':
            weight_used = purchase.refined_weight
            status_text = 'Refined'
        elif purchase.refined_status == 'no':
            weight_used = purchase.weight
            status_text = 'Raw'
        else:
            return response
        
        messages.success(self.request, f"Customer purchase created: {weight_used}g of {status_text} {purchase.metal_type} ({purchase.purity}) added to metal stock.")
        
        return response


class CustomerPurchaseUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomerPurchase
    form_class = CustomerPurchaseForm
    template_name = 'goldsilverpurchase/customer_purchase_form.html'
    success_url = reverse_lazy('gsp:customer_purchase_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = 'Update Customer Purchase'
        return ctx

    def form_valid(self, form):
        """Handle form submission - metal stock is managed by signals"""
        response = super().form_valid(form)
        purchase = self.object
        
        # Determine which weight was used based on refined_status
        if purchase.refined_status == 'yes':
            weight_used = purchase.refined_weight
            status_text = 'Refined'
        elif purchase.refined_status == 'no':
            weight_used = purchase.weight
            status_text = 'Raw'
        else:
            return response
        
        messages.success(self.request, f"Customer purchase updated: {weight_used}g of {status_text} {purchase.metal_type} ({purchase.purity}) metal stock adjusted.")
        
        return response


class CustomerPurchaseDeleteView(LoginRequiredMixin, DeleteView):
    model = CustomerPurchase
    template_name = 'goldsilverpurchase/customer_purchase_confirm_delete.html'
    success_url = reverse_lazy('gsp:customer_purchase_list')



@login_required(login_url='/accounts/login/')
def print_view(request):
    view = PurchaseListView()
    view.request = request  # attach request
    purchases = view.get_queryset()

    total_amount = purchases.aggregate(total=Sum('amount'))['total'] or 0
    return render(request, "goldsilverpurchase/print_view.html", {
        "purchases": purchases,
        "total_amount": total_amount
    })


@login_required(login_url='/accounts/login/')
def export_excel(request):
    view = PurchaseListView()
    view.request = request  # attach request
    purchases = view.get_queryset()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases"

    headers = [
        "Bill No", "Bill Date (BS)", "Party Name","Pan No", "Metal", "Purity",
        "Particular", "Qty", "Rate", "Rate Unit", "Wages", "Discount","Amount", "Payment","Paid", "Remarks"
    ]
    ws.append(headers)

    for p in purchases:
        ws.append([
            p.bill_no,
            str(p.bill_date),
            p.party.party_name,
            p.party.panno,
            p.metal_type,
            p.purity,
            p.particular,
            p.quantity,
            p.rate,
            p.rate_unit,
            p.wages,
            p.discount,
            p.amount,
            p.payment_mode,
            p.is_paid,
            p.remarks
        ])

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Create virtual file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="purchases.xlsx"'

    return response


@login_required(login_url='/accounts/login/')
def export_customer_excel(request):
    view = CustomerPurchaseListView()
    view.request = request
    purchases = view.get_queryset()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CustomerPurchases"

    headers = [
        "Purchase Date (BS)", "Customer Name", "Location", "Phone",
        "Metal", "Ornament Name", "Weight (तोल)", "Percentage", "Final Weight (तोल)", "Refined Weight (तोल)", "Rate (Per तोल)", "Amount"
    ]
    ws.append(headers)

    for p in purchases:
        ws.append([
            str(p.purchase_date),
            p.customer_name,
            p.location,
            p.phone_no,
            p.get_metal_type_display(),
            p.ornament_name,
            p.weight,
            p.percentage,
            p.final_weight,
            p.refined_weight,
            p.rate,
            p.amount,
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="customer_purchases.xlsx"'

    return response



def to_decimal(val):
    if val is None or val == "":
        return Decimal("0")
    return Decimal(str(val))   # SAFE conversion from float → Decimal


@login_required(login_url='/accounts/login/')
def import_excel(request):
    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("gsp:gsp_import_excel")

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            imported = 0
            skipped = 0

            for row in ws.iter_rows(min_row=2, values_only=True):

                if not any(row):   # skip empty rows
                    continue

                try:
                    (
                        bill_no,
                        bill_date_bs,
                        party_name,
                        party_pan,
                        metal_type,
                        purity,
                        particular,
                        qty,
                        rate,
                        rate_unit,
                        wages,
                        discount,
                        amount,
                        payment_mode,
                        is_paid,
                        remarks
                    ) = row
                except Exception:
                    messages.error(request, "Excel format is incorrect. Columns mismatch.")
                    return redirect("gsp:gsp_import_excel")

                # =============== 1️⃣ Duplicate Bill Check ===============
                if GoldSilverPurchase.objects.filter(bill_no=str(bill_no)).exists():
                    skipped += 1
                    continue

                # =============== 2️⃣ Find/Create Party ===============
                party = None

                if party_pan:
                    party = Party.objects.filter(panno=str(party_pan)).first()

                if not party:
                    party = Party.objects.create(
                        party_name=party_name or "Unknown",
                        panno=str(party_pan) if party_pan else "",
                    )

                # =============== 3️⃣ Convert BS Date ===============
                try:
                    y, m, d = map(int, str(bill_date_bs).split("-"))
                    bill_date = ndt.date(y, m, d)
                except:
                    bill_date = ndt.date.today()

                # =============== 4️⃣ Convert all decimals safely ===============
                qty = to_decimal(qty)
                rate = to_decimal(rate)
                wages = to_decimal(wages)
                amount = to_decimal(amount)
                discount=to_decimal(discount)

                # =============== 5️⃣ Create Purchase ===============
                purchase = GoldSilverPurchase.objects.create(
                    bill_no=str(bill_no),
                    bill_date=bill_date,
                    party=party,
                    metal_type=metal_type,
                    purity=purity or '24K',
                    particular=particular,
                    quantity=qty,
                    rate=rate,
                    rate_unit=rate_unit or 'tola',
                    wages=wages,
                    amount=amount,
                    discount=discount,
                    payment_mode=payment_mode,
                    is_paid=bool(is_paid),
                    remarks=remarks
                )

                # --- MetalStock logic ---
                # Determine stock_type: 'raw' if 'raw' in particular, else error
                stock_type_name = 'raw' if particular and 'raw' in particular.lower() else None
                if not stock_type_name:
                    messages.error(request, f"Purchase {purchase.bill_no}: Only 'raw' gold is supported for stock. Add 'raw' in particular.")
                    continue
                stock_type = MetalStockType.objects.filter(name=stock_type_name).first()
                if not stock_type:
                    stock_type = MetalStockType.objects.create(name=stock_type_name)
                # Always create MetalStock if not exists for metal_type, purity, and stock_type
                metal_stock, created = MetalStock.objects.get_or_create(
                    metal_type=metal_type,
                    purity=purity or '24K',
                    stock_type=stock_type,
                    defaults={
                        'quantity': 0,
                        'unit_cost': 0,
                        'rate_unit': rate_unit or 'tola',
                    }
                )
                # If it exists, ensure rate_unit is set (for legacy rows)
                if not created and not metal_stock.rate_unit:
                    metal_stock.rate_unit = rate_unit or 'tola'
                    metal_stock.save()
                # --- MetalStockMovement logic ---
                MetalStockMovement.objects.create(
                    metal_stock=metal_stock,
                    movement_type='in',
                    quantity=qty,
                    rate=rate,
                    reference_type='GoldSilverPurchase',
                    reference_id=purchase.bill_no,
                    notes=remarks,
                    movement_date=bill_date
                )
                imported += 1

            messages.success(
                request,
                f"Imported: {imported} | Skipped duplicates: {skipped}"
            )
            return redirect("gsp:purchaselist")

        except Exception as e:
            messages.error(request, f"Error while importing: {e}")
            return redirect("gsp:gsp_import_excel")

    return render(request, "goldsilverpurchase/import_excel.html")


@login_required(login_url='/accounts/login/')
def import_customer_excel(request):
    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("gsp:customer_import_excel")

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            imported = 0
            skipped = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                try:
                    (
                        purchase_date_bs,
                        customer_name,
                        location,
                        phone_no,
                        metal_type,
                        ornament_name,
                        weight,
                        percentage,
                        final_weight,
                        refined_weight,
                        rate,
                        amount,
                    ) = row
                except Exception:
                    messages.error(request, "Excel format is incorrect. Columns mismatch.")
                    return redirect("gsp:customer_import_excel")

                try:
                    y, m, d = map(int, str(purchase_date_bs).split("-"))
                    purchase_date = ndt.date(y, m, d)
                except Exception:
                    purchase_date = None

                metal_value = str(metal_type or CustomerPurchase.MetalType.GOLD).strip().lower()
                if metal_value not in [choice[0] for choice in CustomerPurchase.MetalType.choices]:
                    metal_value = CustomerPurchase.MetalType.GOLD

                CustomerPurchase.objects.create(
                    purchase_date=purchase_date,
                    customer_name=customer_name or "",
                    location=location or "",
                    phone_no=phone_no or "",
                    metal_type=metal_value,
                    ornament_name=ornament_name or "",
                    weight=D(weight),
                    percentage=D(percentage),
                    final_weight=D(final_weight),
                    refined_weight=D(refined_weight),
                    rate=D(rate),
                    amount=D(amount),
                )
                imported += 1

            messages.success(
                request,
                f"Imported: {imported} records"
            )
            return redirect("gsp:customer_purchase_list")

        except Exception as e:
            messages.error(request, f"Error while importing: {e}")
            return redirect("gsp:customer_import_excel")

    return render(request, "goldsilverpurchase/customer_import_excel.html")


@login_required(login_url='/accounts/login/')
def data_settings(request):
    """Settings page with data export and import controls."""
    if request.method == "POST":
        if "import_file" in request.FILES:
            return import_all_data(request)
        elif "delete_all" in request.POST:
            return delete_all_data(request)
        elif "delete_customer_purchases" in request.POST:
            return delete_customer_purchases(request)
    return render(request, "goldsilverpurchase/data_settings.html")


@login_required(login_url='/accounts/login/')
def export_full_db_dump(request):
    """Export a full database dump as JSON (Django dumpdata)."""
    from django.core.management import call_command
    import io
    from django.conf import settings
    import os

    default_db = settings.DATABASES.get("default", {})
    engine = default_db.get("ENGINE", "")
    if engine.endswith("sqlite3"):
        db_path = default_db.get("NAME")
        if db_path and os.path.exists(db_path):
            with open(db_path, "rb") as db_file:
                response = HttpResponse(db_file.read(), content_type="application/x-sqlite3")
            response["Content-Disposition"] = "attachment; filename=db_backup.sqlite3"
            return response

    out = io.StringIO()
    call_command(
        "dumpdata",
        "--exclude",
        "contenttypes",
        "--exclude",
        "auth.permission",
        stdout=out,
    )

    response = HttpResponse(out.getvalue(), content_type="application/json")
    response["Content-Disposition"] = "attachment; filename=full_db_dump.json"
    return response


@login_required(login_url='/accounts/login/')
def import_full_db_dump(request):
    """Import a full database dump from SQLite or JSON (wipe then restore)."""
    if request.method != "POST" or "import_file" not in request.FILES:
        return redirect("gsp:data_settings")

    confirm_text = request.POST.get("confirm_full_restore", "")
    if confirm_text != "RESTORE_FULL_DB":
        messages.error(request, "Confirmation text does not match. Full restore cancelled.")
        return redirect("gsp:data_settings")

    file = request.FILES.get("import_file")
    if not file:
        messages.error(request, "Please upload a SQLite or JSON dump file.")
        return redirect("gsp:data_settings")

    from django.core.management import call_command
    from django.db import transaction, connections
    from django.conf import settings
    import tempfile
    import shutil
    import os

    tmp_path = None
    try:
        # Detect file type by extension
        filename = file.name.lower()
        is_sqlite = filename.endswith(('.sqlite3', '.db', '.sqlite'))
        
        if is_sqlite:
            # Handle SQLite database file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".sqlite3") as tmp:
                for chunk in file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            
            # Close all database connections
            connections.close_all()
            
            # Get current database path
            default_db = settings.DATABASES.get("default", {})
            db_path = default_db.get("NAME")
            
            if db_path and os.path.exists(db_path):
                # Backup current database
                backup_path = f"{db_path}.bak"
                shutil.copy2(str(db_path), backup_path)
                
                # Replace database with uploaded file
                shutil.copy2(tmp_path, db_path)
                
                # Reconnect to database
                connections.close_all()
                connections['default'].ensure_connection()
            
            messages.success(request, "Full database restore completed from SQLite file.")
        else:
            # Handle JSON dump file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as tmp:
                for chunk in file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            with transaction.atomic():
                call_command("flush", "--noinput")
                call_command("loaddata", tmp_path)

            messages.success(request, "Full database restore completed from JSON file.")
        
        return redirect("gsp:data_settings")
    except Exception as e:
        messages.error(request, f"Full restore failed: {str(e)}")
        return redirect("gsp:data_settings")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@login_required(login_url='/accounts/login/')
def export_all_data(request):
    from openpyxl import Workbook
    from django.http import HttpResponse
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    def convert_to_excel_value(val):
        """Convert any value to Excel-safe format (convert dates to strings)"""
        if val is None:
            return ""
        # Convert nepali_datetime.date objects to string
        if hasattr(val, 'strftime'):
            return str(val)
        return val

    def add_sheet(title, headers, rows):
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for row in rows:
            # Convert all values in row to Excel-safe format
            safe_row = tuple(convert_to_excel_value(v) for v in row)
            ws.append(safe_row)

    # --- Finance Models ---
    from finance.models import EmployeeSalary, SundryCreditor, SundryDebtor, Loan, CreditorTransaction, DebtorTransaction, Expense, Employee
    
    # Employee sheet
    employee_rows = [
        (e.id, e.first_name, e.last_name, e.email, e.phone, e.position, e.base_salary, e.hire_date, e.is_active, e.created_at, e.updated_at) for e in Employee.objects.all().order_by("created_at")
    ]
    add_sheet("Employee", ["ID", "First Name", "Last Name", "Email", "Phone", "Position", "Base Salary", "Hire Date", "Is Active", "Created At", "Updated At"], employee_rows)

    salary_rows = [
        (s.id, s.employee.first_name + ' ' + s.employee.last_name if s.employee else "", s.month, s.total_salary, s.amount_paid, s.status, s.created_at, s.updated_at) for s in EmployeeSalary.objects.select_related('employee').all().order_by("created_at")
    ]
    add_sheet("EmployeeSalary", ["ID", "Employee Name", "Month", "Total Salary", "Amount Paid", "Status", "Created At", "Updated At"], salary_rows)

    # Expense sheet
    expense_rows = [
        (ex.id, ex.category, ex.description, ex.amount, ex.expense_date, ex.notes, ex.created_at, ex.updated_at) for ex in Expense.objects.all().order_by("created_at")
    ]
    add_sheet("Expense", ["ID", "Category", "Description", "Amount", "Expense Date", "Notes", "Created At", "Updated At"], expense_rows)

    creditor_rows = [
        (c.id, c.name, c.bs_date, c.opening_balance, c.current_balance, c.is_paid, c.created_at, c.updated_at) for c in SundryCreditor.objects.all().order_by("created_at")
    ]
    add_sheet("SundryCreditor", ["ID", "Name", "BS Date", "Opening Balance", "Current Balance", "Is Paid", "Created At", "Updated At"], creditor_rows)

    debtor_rows = [
        (d.id, d.name, d.bs_date, d.opening_balance, d.current_balance, d.is_paid, d.created_at, d.updated_at) for d in SundryDebtor.objects.all().order_by("created_at")
    ]
    add_sheet("SundryDebtor", ["ID", "Name", "BS Date", "Opening Balance", "Current Balance", "Is Paid", "Created At", "Updated At"], debtor_rows)

    loan_rows = [
        (l.id, l.bank_name, l.amount, l.interest_rate, l.start_date, l.notes, l.created_at, l.updated_at) for l in Loan.objects.all().order_by("created_at")
    ]
    add_sheet("Loan", ["ID", "Bank Name", "Amount", "Interest Rate", "Start Date", "Notes", "Created At", "Updated At"], loan_rows)

    creditor_tx_rows = [
        (tx.id, tx.creditor.name if tx.creditor else "", tx.transaction_type, tx.reference_no, tx.amount, tx.transaction_date, tx.due_date, tx.created_at) for tx in CreditorTransaction.objects.select_related('creditor').all().order_by("created_at")
    ]
    add_sheet("CreditorTransaction", ["ID", "Creditor Name", "Transaction Type", "Reference No", "Amount", "Transaction Date", "Due Date", "Created At"], creditor_tx_rows)

    # DebtorTransaction model may not exist, check if it's defined
    try:
        debtor_tx_rows = [
            (tx.id, tx.debtor.name if tx.debtor else "", tx.transaction_type, tx.reference_no, tx.amount, tx.transaction_date, tx.due_date, tx.created_at) for tx in DebtorTransaction.objects.select_related('debtor').all().order_by("created_at")
        ]
        add_sheet("DebtorTransaction", ["ID", "Debtor Name", "Transaction Type", "Reference No", "Amount", "Transaction Date", "Due Date", "Created At"], debtor_tx_rows)
    except Exception:
        pass  # Skip if DebtorTransaction doesn't exist or has different structure

    # --- GoldSilverPurchase Models ---
    from goldsilverpurchase.models import GoldSilverPurchase, CustomerPurchase, MetalStock, MetalStockMovement, Party, MetalStockType
    
    # Party sheet
    party_rows = [
        (p.id, p.party_name, p.panno) for p in Party.objects.all().order_by("id")
    ]
    add_sheet("Party", ["ID", "Party Name", "PAN No"], party_rows)

    # MetalStockType sheet
    metal_type_rows = [
        (mt.id, mt.name, mt.description) for mt in MetalStockType.objects.all().order_by("id")
    ]
    add_sheet("MetalStockType", ["ID", "Name", "Description"], metal_type_rows)

    gsp_rows = [
        (p.bill_no, p.bill_date, p.party.party_name if p.party else "", p.particular, p.metal_type, p.purity, p.quantity, p.rate, p.rate_unit, p.wages, p.discount, p.amount, p.payment_mode, p.is_paid, p.remarks, p.created_at, p.updated_at) for p in GoldSilverPurchase.objects.select_related("party").order_by("created_at")
    ]
    add_sheet("GoldSilverPurchase", ["Bill No", "Bill Date", "Party", "Particular", "Metal Type", "Purity", "Quantity", "Rate", "Rate Unit", "Wages", "Discount", "Amount", "Payment Mode", "Is Paid", "Remarks", "Created At", "Updated At"], gsp_rows)

    customer_rows = [
        (cp.sn, cp.purchase_date, cp.customer_name, cp.location, cp.phone_no, cp.metal_type, cp.ornament_name, cp.weight, cp.refined_weight, cp.rate, cp.amount, cp.created_at, cp.updated_at) for cp in CustomerPurchase.objects.order_by("created_at")
    ]
    add_sheet("CustomerPurchase", ["SN", "Purchase Date", "Customer Name", "Location", "Phone", "Metal Type", "Ornament", "Weight", "Refined Weight", "Rate", "Amount", "Created At", "Updated At"], customer_rows)

    metal_stock_rows = [
        (ms.id, ms.metal_type, ms.stock_type.name if ms.stock_type else "", ms.purity, ms.quantity, ms.unit_cost, ms.rate_unit, ms.total_cost, ms.location, ms.remarks, ms.last_updated, ms.created_at) for ms in MetalStock.objects.select_related("stock_type").order_by("created_at")
    ]
    add_sheet("MetalStock", ["ID", "Metal Type", "Stock Type", "Purity", "Quantity", "Unit Cost", "Rate Unit", "Total Cost", "Location", "Remarks", "Last Updated", "Created At"], metal_stock_rows)

    metal_movement_rows = [
        (msm.id, msm.metal_stock_id, msm.movement_type, msm.quantity, msm.rate, msm.reference_type, msm.reference_id, msm.notes, msm.movement_date, msm.created_at) for msm in MetalStockMovement.objects.select_related("metal_stock").order_by("created_at")
    ]
    add_sheet("MetalStockMovement", ["ID", "Metal Stock ID", "Movement Type", "Quantity", "Rate", "Reference Type", "Reference ID", "Notes", "Movement Date", "Created At"], metal_movement_rows)

    # --- Main Models ---
    from main.models import Stock, DailyRate
    
    stock_rows = [
        (s.id, s.year, s.diamond, s.gold, s.silver, s.jardi, s.wages, s.gold_silver_rate_unit, s.diamond_rate, s.gold_rate, s.silver_rate, s.created_at, s.updated_at) for s in Stock.objects.order_by("year")
    ]
    add_sheet("Stock", ["ID", "Year", "Diamond", "Gold", "Silver", "Jardi", "Wages", "Rate Unit", "Diamond Rate", "Gold Rate", "Silver Rate", "Created At", "Updated At"], stock_rows)

    daily_rate_rows = [
        (dr.id, dr.bs_date, dr.gold_rate, dr.silver_rate, dr.gold_rate_10g, dr.silver_rate_10g, dr.created_at, dr.updated_at) for dr in DailyRate.objects.order_by("-created_at")
    ]
    add_sheet("DailyRate", ["ID", "BS Date", "Gold Rate", "Silver Rate", "Gold Rate 10g", "Silver Rate 10g", "Created At", "Updated At"], daily_rate_rows)

    # --- Ornament Models ---
    from ornament.models import Stone, Motimala, Potey, MainCategory, SubCategory, Kaligar, Kaligar_Ornaments, Kaligar_CashAccount, Kaligar_GoldAccount, Ornament
    
    # Stone sheet
    stone_rows = [
        (st.id, st.name, st.cost_per_carat, st.carat, st.cost_price, st.sales_per_carat, st.sales_price, st.profit, st.created_at, st.updated_at) for st in Stone.objects.all().order_by("created_at")
    ]
    add_sheet("Stone", ["ID", "Name", "Cost Per Carat", "Carat", "Cost Price", "Sales Per Carat", "Sales Price", "Profit", "Created At", "Updated At"], stone_rows)

    # Motimala sheet
    motimala_rows = [
        (m.id, m.name, m.cost_per_mala, m.quantity, m.cost_price, m.sales_per_mala, m.sales_price, m.profit, m.created_at, m.updated_at) for m in Motimala.objects.all().order_by("created_at")
    ]
    add_sheet("Motimala", ["ID", "Name", "Cost Per Mala", "Quantity", "Cost Price", "Sales Per Mala", "Sales Price", "Profit", "Created At", "Updated At"], motimala_rows)

    # Potey sheet
    potey_rows = [
        (p.id, p.name, p.loon, p.cost_per_loon, p.cost_price, p.sales_per_loon, p.sales_price, p.profit, p.created_at, p.updated_at) for p in Potey.objects.all().order_by("created_at")
    ]
    add_sheet("Potey", ["ID", "Name", "Loon", "Cost Per Loon", "Cost Price", "Sales Per Loon", "Sales Price", "Profit", "Created At", "Updated At"], potey_rows)

    # MainCategory sheet
    main_cat_rows = [
        (mc.id, mc.name) for mc in MainCategory.objects.all().order_by("id")
    ]
    add_sheet("MainCategory", ["ID", "Name"], main_cat_rows)

    # SubCategory sheet
    sub_cat_rows = [
        (sc.id, sc.name) for sc in SubCategory.objects.all().order_by("id")
    ]
    add_sheet("SubCategory", ["ID", "Name"], sub_cat_rows)

    # Ornament sheet
    ornament_rows = [
        (o.id, o.code, o.ornament_name, o.type, o.ornament_type, o.metal_type, o.weight, o.status, o.created_at, o.updated_at) for o in Ornament.objects.all().order_by("created_at")
    ]
    add_sheet("Ornament", ["ID", "Code", "Name", "Type", "Ornament Type", "Metal Type", "Weight", "Status", "Created At", "Updated At"], ornament_rows)

    # Kaligar sheet
    kaligar_rows = [
        (k.id, k.name, k.address, k.phone_no) for k in Kaligar.objects.all().order_by("id")
    ]
    add_sheet("Kaligar", ["ID", "Name", "Address", "Phone"], kaligar_rows)

    # Kaligar_Ornaments sheet
    kaligar_orn_rows = [
        (ko.id, ko.kaligar.name if ko.kaligar else "", ko.date, ko.gold_given, ko.ornament_weight, ko.jarti, ko.gold_return, ko.gold_loss, ko.gold_purity) for ko in Kaligar_Ornaments.objects.select_related("kaligar").all().order_by("id")
    ]
    add_sheet("Kaligar_Ornaments", ["ID", "Kaligar", "Date", "Gold Given", "Ornament Weight", "Jarti", "Gold Return", "Gold Loss", "Gold Purity"], kaligar_orn_rows)

    # Kaligar_CashAccount sheet
    kaligar_cash_rows = [
        (kc.id, kc.kaligar.name if kc.kaligar else "", kc.date, kc.particular, kc.amount_taken, kc.to_pay, kc.provided_by) for kc in Kaligar_CashAccount.objects.select_related("kaligar").all().order_by("id")
    ]
    add_sheet("Kaligar_CashAccount", ["ID", "Kaligar", "Date", "Particular", "Amount Taken", "To Pay", "Provided By"], kaligar_cash_rows)

    # Kaligar_GoldAccount sheet
    kaligar_gold_rows = [
        (kg.id, kg.kaligar.name if kg.kaligar else "", kg.date, kg.gold_deposit, kg.gold_loss, kg.gold_remaining) for kg in Kaligar_GoldAccount.objects.select_related("kaligar").all().order_by("id")
    ]
    add_sheet("Kaligar_GoldAccount", ["ID", "Kaligar", "Date", "Gold Deposit", "Gold Loss", "Gold Remaining"], kaligar_gold_rows)

    # --- Order Models ---
    from order.models import Order, OrderMetalStock, OrderPayment, OrderOrnament, DebtorPayment
    
    order_rows = [
        (o.sn, o.order_date, o.deliver_date, o.customer_name, o.phone_number, o.status, o.order_type, o.description, o.discount, o.amount, o.subtotal, o.tax, o.total, o.remaining_amount, o.created_at, o.updated_at) for o in Order.objects.order_by("created_at")
    ]
    add_sheet("Order", ["SN", "Order Date", "Deliver Date", "Customer Name", "Phone", "Status", "Order Type", "Description", "Discount", "Amount", "Subtotal", "Tax", "Total", "Remaining Amount", "Created At", "Updated At"], order_rows)

    order_metal_rows = [
        (oms.order_id, oms.stock_type.name if oms.stock_type else "", oms.metal_type, oms.purity, oms.quantity, oms.rate_per_gram, oms.rate_unit, oms.line_amount, oms.remarks, oms.created_at, oms.updated_at) for oms in OrderMetalStock.objects.select_related("order", "stock_type").order_by("created_at")
    ]
    add_sheet("OrderMetalStock", ["Order ID", "Stock Type", "Metal Type", "Purity", "Quantity", "Rate Per Gram", "Rate Unit", "Line Amount", "Remarks", "Created At", "Updated At"], order_metal_rows)

    payment_rows = [
        (op.order_id, op.payment_mode, op.amount, op.created_at, op.updated_at) for op in OrderPayment.objects.select_related("order").order_by("created_at")
    ]
    add_sheet("OrderPayment", ["Order ID", "Payment Mode", "Amount", "Created At", "Updated At"], payment_rows)

    order_ornament_rows = [
        (oo.order_id, oo.ornament_id, oo.gold_rate, oo.diamond_rate, oo.zircon_rate, oo.stone_rate, oo.jarti, oo.jyala, oo.line_amount, oo.created_at, oo.updated_at) for oo in OrderOrnament.objects.select_related("order", "ornament").order_by("created_at")
    ]
    add_sheet("OrderOrnament", ["Order ID", "Ornament ID", "Gold Rate", "Diamond Rate", "Zircon Rate", "Stone Rate", "Jarti", "Jyala", "Line Amount", "Created At", "Updated At"], order_ornament_rows)

    # DebtorPayment sheet
    debtor_payment_rows = [
        (dp.id, dp.order_payment.order_id, dp.debtor.name if dp.debtor else "", dp.transaction_type, dp.created_at, dp.updated_at) for dp in DebtorPayment.objects.select_related("order_payment__order", "debtor").all().order_by("created_at")
    ]
    add_sheet("DebtorPayment", ["ID", "Order ID", "Debtor Name", "Transaction Type", "Created At", "Updated At"], debtor_payment_rows)

    # --- Sales Models ---
    from sales.models import Sale, SalesMetalStock
    sale_rows = [
        (s.order_id, s.bill_no, s.sale_date, s.created_at, s.updated_at) for s in Sale.objects.select_related("order").order_by("created_at")
    ]
    add_sheet("Sale", ["Order ID", "Bill No", "Sale Date", "Created At", "Updated At"], sale_rows)

    sales_metal_rows = [
        (sms.sale_id, sms.stock_type.name if sms.stock_type else "", sms.metal_type, sms.purity, sms.quantity, sms.rate_per_gram, sms.rate_unit, sms.line_amount, sms.remarks, sms.created_at, sms.updated_at) for sms in SalesMetalStock.objects.select_related("sale", "stock_type").order_by("created_at")
    ]
    add_sheet("SalesMetalStock", ["Sale ID", "Stock Type", "Metal Type", "Purity", "Quantity", "Rate Per Gram", "Rate Unit", "Line Amount", "Remarks", "Created At", "Updated At"], sales_metal_rows)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=all_data.xlsx"
    wb.save(response)
    return response


@login_required(login_url='/accounts/login/')
def import_all_data(request):
    """Import all data from a multi-sheet Excel file."""
    if request.method != "POST" or "import_file" not in request.FILES:
        return redirect("gsp:data_settings")

    file = request.FILES.get("import_file")
    if not file:
        messages.error(request, "Please upload a file.")
        return redirect("gsp:data_settings")

    try:
        wb = openpyxl.load_workbook(file)
        imported_count = {}
        errors = []
        ornament_id_map = {}
        def to_date(val):
            if not val:
                return None
            try:
                if isinstance(val, str):
                    y, m, d = map(int, val.split("-"))
                    return ndt.date(y, m, d) if ndt else None
                return val
            except Exception:
                return None

        # ========== Import GoldSilverPurchase ==========
        if "GoldSilverPurchase" in wb.sheetnames:
            ws = wb["GoldSilverPurchase"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    (
                        bill_no,
                        bill_date,
                        party_name,
                        particular,
                        metal_type,
                        purity,
                        quantity,
                        rate,
                        rate_unit,
                        wages,
                        discount,
                        amount,
                        payment_mode,
                        is_paid,
                        remarks,
                        created_at,
                        updated_at,
                    ) = row[:17]

                    if GoldSilverPurchase.objects.filter(bill_no=str(bill_no)).exists():
                        continue

                    party = None
                    if party_name:
                        party = Party.objects.filter(party_name=party_name).first()
                        if not party:
                            party = Party.objects.create(
                                party_name=party_name,
                                panno="000000000",
                            )

                    GoldSilverPurchase.objects.create(
                        bill_no=str(bill_no),
                        bill_date=to_date(bill_date),
                        party=party,
                        particular=particular,
                        metal_type=metal_type or "gold",
                        purity=purity or "22K",
                        quantity=to_decimal(quantity),
                        rate=to_decimal(rate),
                        rate_unit=rate_unit or "tola",
                        wages=to_decimal(wages),
                        discount=to_decimal(discount),
                        amount=to_decimal(amount),
                        payment_mode=payment_mode or "cash",
                        is_paid=bool(is_paid),
                        remarks=remarks,
                    )
                    count += 1
                except Exception as e:
                    errors.append(f"GoldSilverPurchase row error: {e}")
            imported_count["GoldSilverPurchase"] = count

        # ========== Import CustomerPurchase ==========
        if "CustomerPurchase" in wb.sheetnames:
            ws = wb["CustomerPurchase"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    (
                        sn,
                        purchase_date,
                        customer_name,
                        location,
                        phone_no,
                        metal_type,
                        ornament_name,
                        weight,
                        refined_weight,
                        rate,
                        amount,
                        created_at,
                        updated_at,
                    ) = row[:13]

                    if CustomerPurchase.objects.filter(sn=str(sn)).exists():
                        continue

                    CustomerPurchase.objects.create(
                        sn=str(sn),
                        purchase_date=to_date(purchase_date),
                        customer_name=customer_name or "",
                        location=location or "",
                        phone_no=phone_no or "0000000000",
                        metal_type=metal_type or "gold",
                        ornament_name=ornament_name or "",
                        weight=to_decimal(weight),
                        refined_weight=to_decimal(refined_weight),
                        rate=to_decimal(rate),
                        amount=to_decimal(amount),
                    )
                    count += 1
                except Exception as e:
                    errors.append(f"CustomerPurchase row error: {e}")
            imported_count["CustomerPurchase"] = count

        # ========== Import Orders ==========
        if "Orders" in wb.sheetnames:
            ws = wb["Orders"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    (
                        sn,
                        order_date,
                        deliver_date,
                        customer_name,
                        phone_number,
                        status,
                        order_type,
                        description,
                        discount,
                        amount,
                        subtotal,
                        tax,
                        total,
                        payment_mode,
                        payment_amount,
                        remaining_amount,
                        created_at,
                        updated_at,
                    ) = row[:18]

                    if Order.objects.filter(sn=sn).exists():
                        continue

                    Order.objects.create(
                        sn=sn,
                        order_date=to_date(order_date),
                        deliver_date=to_date(deliver_date),
                        customer_name=customer_name or "",
                        phone_number=phone_number or "0000000000",
                        status=status or "order",
                        order_type=order_type or "custom",
                        description=description or "",
                        discount=to_decimal(discount),
                        amount=to_decimal(amount),
                        subtotal=to_decimal(subtotal),
                        tax=to_decimal(tax),
                        total=to_decimal(total),
                        payment_mode=payment_mode or "cash",
                        payment_amount=to_decimal(payment_amount),
                        remaining_amount=to_decimal(remaining_amount),
                    )
                    count += 1
                except Exception as e:
                    errors.append(f"Orders row error: {e}")
            imported_count["Orders"] = count

        # ========== Import OrderPayments ==========
        if "OrderPayments" in wb.sheetnames:
            ws = wb["OrderPayments"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    (
                        order_sn,
                        payment_mode,
                        amount,
                        created_at,
                        updated_at,
                    ) = row[:5]

                    try:
                        order = Order.objects.get(sn=order_sn)
                        OrderPayment.objects.create(
                            order=order,
                            payment_mode=payment_mode or "cash",
                            amount=to_decimal(amount),
                        )
                        count += 1
                    except Order.DoesNotExist:
                        pass
                except Exception as e:
                    errors.append(f"OrderPayments row error: {e}")
            imported_count["OrderPayments"] = count

        # ========== Import Ornaments ==========
        if "Ornaments" in wb.sheetnames:
            ws = wb["Ornaments"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    (
                        orn_id,
                        code,
                        ornament_name,
                        metal_type,
                        karat,
                        ornament_type,
                        main_category,
                        sub_category,
                        kaligar_name,
                        weight,
                        gross_weight,
                        diamond_weight,
                        zircon_weight,
                        stone_weight,
                        stone_total_price,
                        jarti,
                        jyala,
                        ornament_date,
                        description,
                        image_url,
                        created_at,
                        updated_at,
                    ) = row[:22]

                    orn_id = int(orn_id) if orn_id else None

                    if code:
                        existing = Ornament.objects.filter(code=str(code)).first()
                        if existing:
                            ornament_id_map[orn_id] = existing.id
                            continue

                    main_cat = None
                    if main_category:
                        main_cat, _ = MainCategory.objects.get_or_create(name=main_category)

                    sub_cat = None
                    if sub_category:
                        sub_cat, _ = SubCategory.objects.get_or_create(name=sub_category)

                    kaligar = None
                    if kaligar_name:
                        kaligar, _ = Kaligar.objects.get_or_create(
                            name=kaligar_name,
                            defaults={"panno": "000000000"},
                        )
                    else:
                        kaligar = Kaligar.objects.first()
                        if not kaligar:
                            kaligar = Kaligar.objects.create(
                                name="Default",
                                panno="000000000",
                            )

                    new_ornament = Ornament.objects.create(
                        code=str(code) if code else None,
                        ornament_name=ornament_name or "",
                        metal_type=metal_type or "Gold",
                        type=karat or "24KARAT",
                        ornament_type=ornament_type or "stock",
                        maincategory=main_cat,
                        subcategory=sub_cat,
                        kaligar=kaligar,
                        weight=to_decimal(weight),
                        gross_weight=to_decimal(gross_weight),
                        diamond_weight=to_decimal(diamond_weight),
                        zircon_weight=to_decimal(zircon_weight),
                        stone_weight=to_decimal(stone_weight),
                        stone_totalprice=to_decimal(stone_total_price),
                        jarti=to_decimal(jarti),
                        jyala=to_decimal(jyala),
                        ornament_date=to_date(ornament_date),
                        description=description or "",
                    )
                    ornament_id_map[orn_id] = new_ornament.id
                    count += 1
                except Exception as e:
                    errors.append(f"Ornaments row error: {e}")
            imported_count["Ornaments"] = count

        # ========== Import OrderOrnaments ==========
        if "OrderOrnaments" in wb.sheetnames:
            ws = wb["OrderOrnaments"]
            count = 0
            row_num = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                if not any(row):
                    continue
                try:
                    (
                        order_sn,
                        ornament_id,
                        gold_rate,
                        diamond_rate,
                        zircon_rate,
                        stone_rate,
                        jarti,
                        jyala,
                        line_amount,
                        created_at,
                        updated_at,
                    ) = row[:11]

                    try:
                        order = Order.objects.get(sn=order_sn)
                        ornament_id = int(ornament_id) if ornament_id else None
                        mapped_ornament_id = ornament_id_map.get(ornament_id)
                        if mapped_ornament_id:
                            ornament = Ornament.objects.get(id=mapped_ornament_id)
                            OrderOrnament.objects.create(
                                order=order,
                                ornament=ornament,
                                gold_rate=to_decimal(gold_rate),
                                diamond_rate=to_decimal(diamond_rate),
                                zircon_rate=to_decimal(zircon_rate),
                                stone_rate=to_decimal(stone_rate),
                                jarti=to_decimal(jarti),
                                jyala=to_decimal(jyala),
                                line_amount=to_decimal(line_amount),
                            )
                            count += 1
                        else:
                            errors.append(
                                f"OrderOrnaments row {row_num}: Ornament ID {ornament_id} not found in mapping"
                            )
                    except Order.DoesNotExist:
                        errors.append(f"OrderOrnaments row {row_num}: Order SN {order_sn} not found")
                    except Ornament.DoesNotExist:
                        errors.append(
                            f"OrderOrnaments row {row_num}: Ornament ID {mapped_ornament_id} not found after mapping"
                        )
                except Exception as e:
                    errors.append(f"OrderOrnaments row error: {e}")
            imported_count["OrderOrnaments"] = count

        # ========== Import Sales ==========
        if "Sales" in wb.sheetnames:
            ws = wb["Sales"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    (
                        order_sn,
                        bill_no,
                        sale_date,
                        created_at,
                        updated_at,
                    ) = row[:5]

                    try:
                        order = Order.objects.get(sn=order_sn)
                        if not Sale.objects.filter(order=order).exists():
                            Sale.objects.create(
                                order=order,
                                bill_no=bill_no or None,
                                sale_date=to_date(sale_date),
                            )
                            count += 1
                    except Order.DoesNotExist:
                        pass
                except Exception as e:
                    errors.append(f"Sales row error: {e}")
            imported_count["Sales"] = count

        summary_msg = "Import completed: " + " | ".join(
            [f"{k}: {v}" for k, v in imported_count.items() if v > 0]
        )
        summary_msg += f" | Ornament ID mappings: {len(ornament_id_map)}"
        messages.success(request, summary_msg)

        if errors:
            error_summary = f"{len(errors)} row(s) had errors:\n" + "\n".join(errors[:20])
            if len(errors) > 20:
                error_summary += f"\n... and {len(errors) - 20} more errors"
            messages.warning(request, error_summary)
            import logging

            logger = logging.getLogger(__name__)
            for error in errors:
                logger.error(error)

        return redirect("gsp:data_settings")

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect("gsp:data_settings")


@login_required(login_url='/accounts/login/')
def delete_customer_purchases(request):
    """Delete all customer purchases with password protection."""
    if request.method != "POST":
        return redirect("gsp:data_settings")

    # Verify password
    password = request.POST.get("customer_purchase_password", "")
    confirm_text = request.POST.get("confirm_customer_delete", "")

    # Password for deleting customer purchases (set to "admin123" - can be changed)
    DELETE_PASSWORD = "admin123"

    if password != DELETE_PASSWORD:
        messages.error(request, "Incorrect password. Customer purchases were not deleted.")
        return redirect("gsp:data_settings")

    if confirm_text != "DELETE_CUSTOMER_PURCHASES":
        messages.error(request, "Confirmation text does not match. Customer purchases were not deleted.")
        return redirect("gsp:data_settings")

    try:
        customer_purchase_count = CustomerPurchase.objects.count()
        CustomerPurchase.objects.all().delete()

        messages.success(
            request,
            f"Successfully deleted {customer_purchase_count} customer purchases. This action cannot be undone."
        )
        return redirect("gsp:customer_purchase_list")

    except Exception as e:
        messages.error(request, f"Error while deleting customer purchases: {e}")
        return redirect("gsp:data_settings")


@login_required(login_url='/accounts/login/')
def export_all_data_json(request):
    """Export all data as JSON for faster backups - includes all models from all apps."""
    import json
    from decimal import Decimal
    from datetime import datetime

    def json_serializer(obj):
        """Custom JSON serializer for Decimal and date objects."""
        if isinstance(obj, Decimal):
            return str(obj)
        if hasattr(obj, 'strftime'):  # Handles both datetime and nepali_datetime.date
            return str(obj)
        return str(obj)

    try:
        data = {}

        def serialize_queryset(queryset):
            rows = []
            for obj in queryset:
                row = {}
                for field in obj._meta.fields:
                    row[field.name] = field.value_from_object(obj)
                rows.append(row)
            return rows

        # --- Finance Models ---
        from finance.models import EmployeeSalary, SundryCreditor, SundryDebtor, Loan, CreditorTransaction, DebtorTransaction, Expense, Employee

        # --- GoldSilverPurchase Models ---
        from goldsilverpurchase.models import GoldSilverPurchase, CustomerPurchase, MetalStock, MetalStockMovement, Party, MetalStockType

        # --- Main Models ---
        from main.models import Stock, DailyRate

        # --- Ornament Models ---
        from ornament.models import Stone, Motimala, Potey, MainCategory, SubCategory, Kaligar, Kaligar_Ornaments, Kaligar_CashAccount, Kaligar_GoldAccount, Ornament

        # --- Order Models ---
        from order.models import Order, OrderMetalStock, OrderPayment, OrderOrnament, DebtorPayment

        # --- Sales Models ---
        from sales.models import Sale, SalesMetalStock

        model_map = [
            ("Employee", Employee),
            ("EmployeeSalary", EmployeeSalary),
            ("Expense", Expense),
            ("SundryCreditor", SundryCreditor),
            ("SundryDebtor", SundryDebtor),
            ("Loan", Loan),
            ("CreditorTransaction", CreditorTransaction),
            ("DebtorTransaction", DebtorTransaction),
            ("Party", Party),
            ("MetalStockType", MetalStockType),
            ("GoldSilverPurchase", GoldSilverPurchase),
            ("CustomerPurchase", CustomerPurchase),
            ("MetalStock", MetalStock),
            ("MetalStockMovement", MetalStockMovement),
            ("Stock", Stock),
            ("DailyRate", DailyRate),
            ("Stone", Stone),
            ("Motimala", Motimala),
            ("Potey", Potey),
            ("MainCategory", MainCategory),
            ("SubCategory", SubCategory),
            ("Kaligar", Kaligar),
            ("Kaligar_Ornaments", Kaligar_Ornaments),
            ("Kaligar_CashAccount", Kaligar_CashAccount),
            ("Kaligar_GoldAccount", Kaligar_GoldAccount),
            ("Ornament", Ornament),
            ("Order", Order),
            ("OrderMetalStock", OrderMetalStock),
            ("OrderPayment", OrderPayment),
            ("OrderOrnament", OrderOrnament),
            ("DebtorPayment", DebtorPayment),
            ("Sale", Sale),
            ("SalesMetalStock", SalesMetalStock),
        ]

        for key, model in model_map:
            data[key] = serialize_queryset(model.objects.all())

        # Create JSON response
        json_str = json.dumps(data, default=json_serializer, indent=2)
        response = HttpResponse(json_str, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename=all_data_backup.json'
        return response

    except Exception as e:
        messages.error(request, f"Error exporting JSON: {str(e)}")
        return HttpResponse(f"Export failed: {str(e)}", status=500)


@login_required(login_url='/accounts/login/')
def import_all_data_json(request):
    """Import all data from a JSON file for faster restoration."""
    import json
    from datetime import date, datetime
    from django.db import models as dj_models, transaction
    from django.utils.dateparse import parse_date, parse_datetime
    from nepali_datetime_field.models import NepaliDateField
    
    if request.method != "POST" or "import_file" not in request.FILES:
        return redirect("gsp:data_settings")

    file = request.FILES.get("import_file")
    if not file:
        messages.error(request, "Please upload a JSON file.")
        return redirect("gsp:data_settings")

    try:
        # Read and parse JSON
        file_content = file.read().decode('utf-8')
        data = json.loads(file_content)
        imported_count = {}
        errors = []

        def parse_nepali_date(value):
            if not value:
                return None
            if hasattr(value, "strftime"):
                return value
            try:
                y, m, d = map(int, str(value).split("-"))
                return ndt.date(y, m, d) if ndt else None
            except Exception:
                return None

        def is_int_like(value):
            try:
                int(str(value))
                return True
            except Exception:
                return False

        def parse_field_value(field, value):
            if value in (None, ""):
                return None
            if isinstance(field, dj_models.DecimalField):
                return to_decimal(value)
            if isinstance(field, NepaliDateField):
                return parse_nepali_date(value)
            if isinstance(field, dj_models.DateTimeField):
                return parse_datetime(str(value)) or None
            if isinstance(field, dj_models.DateField):
                return parse_date(str(value)) or None
            if isinstance(field, dj_models.BooleanField):
                if isinstance(value, bool):
                    return value
                return str(value).lower() in {"1", "true", "yes", "y"}
            if isinstance(field, (dj_models.IntegerField, dj_models.AutoField, dj_models.BigIntegerField, dj_models.PositiveIntegerField)):
                try:
                    return int(str(value))
                except Exception:
                    return None
            return value

        def resolve_fk_id(model_key, field, row):
            raw = row.get(field.name)
            if raw in (None, ""):
                raw = row.get(f"{field.name}_id")

            if raw in (None, ""):
                if field.name == "employee":
                    raw = row.get("employee_name")
                    if raw:
                        parts = str(raw).strip().split()
                        first = parts[0] if parts else ""
                        last = " ".join(parts[1:]) if len(parts) > 1 else ""
                        employee = Employee.objects.filter(first_name=first, last_name=last).first()
                        return employee.pk if employee else None
                if field.name == "creditor":
                    raw = row.get("creditor_name")
                if field.name == "debtor":
                    raw = row.get("debtor_name")
                if field.name == "party":
                    raw = row.get("party_name")
                if field.name == "kaligar":
                    raw = row.get("kaligar") or row.get("kaligar_name")
                if field.name == "stock_type":
                    raw = row.get("stock_type")
                if field.name == "order_payment":
                    raw = row.get("order_payment_id")
                    if raw in (None, ""):
                        order_id = row.get("order_id") or row.get("order")
                        if order_id:
                            payment = OrderPayment.objects.filter(order_id=order_id).order_by("id").first()
                            return payment.pk if payment else None
                if field.name == "order":
                    raw = row.get("order_id")
                if field.name == "ornament":
                    raw = row.get("ornament_id") or row.get("ornament_code")

            if raw in (None, ""):
                return None

            if is_int_like(raw):
                return int(str(raw))

            related_model = field.remote_field.model
            if hasattr(related_model, "party_name"):
                obj = related_model.objects.filter(party_name=raw).first()
                return obj.pk if obj else None
            if hasattr(related_model, "name"):
                obj = related_model.objects.filter(name=raw).first()
                return obj.pk if obj else None
            if hasattr(related_model, "code"):
                obj = related_model.objects.filter(code=raw).first()
                return obj.pk if obj else None
            return None

        lookup_fields = {
            "Employee": [("email",), ("first_name", "last_name", "position")],
            "EmployeeSalary": [("employee", "month")],
            "Expense": [("category", "description", "expense_date", "amount")],
            "SundryCreditor": [("name",)],
            "SundryDebtor": [("name",)],
            "Loan": [("bank_name", "start_date", "amount", "interest_rate")],
            "CreditorTransaction": [("creditor", "transaction_type", "reference_no", "amount", "transaction_date")],
            "DebtorTransaction": [("debtor", "transaction_type", "reference_no", "amount", "transaction_date")],
            "Party": [("party_name", "panno")],
            "MetalStockType": [("name",)],
            "MetalStock": [("metal_type", "stock_type", "purity", "location")],
            "GoldSilverPurchase": [("bill_no",)],
            "CustomerPurchase": [("sn",)],
            "MetalStockMovement": [("metal_stock", "movement_type", "quantity", "rate", "reference_type", "reference_id", "movement_date")],
            "Stock": [("year",)],
            "DailyRate": [("bs_date",)],
            "Stone": [("name",)],
            "Motimala": [("name",)],
            "Potey": [("name",)],
            "MainCategory": [("name",)],
            "SubCategory": [("name",)],
            "Kaligar": [("name", "panno")],
            "Kaligar_Ornaments": [("kaligar", "date", "ornament_weight", "gold_given")],
            "Kaligar_CashAccount": [("kaligar", "date", "particular", "amount_taken", "to_pay")],
            "Kaligar_GoldAccount": [("kaligar", "date", "gold_deposit")],
            "Ornament": [("code",)],
            "Order": [("sn",)],
            "OrderMetalStock": [("order", "stock_type", "metal_type", "purity", "quantity", "rate_per_gram")],
            "OrderPayment": [("order", "payment_mode", "amount")],
            "OrderOrnament": [("order", "ornament", "line_amount")],
            "DebtorPayment": [("order_payment",)],
            "Sale": [("order",)],
            "SalesMetalStock": [("sale", "stock_type", "metal_type", "purity", "quantity", "rate_per_gram")],
        }

        def build_lookup(model_key, model, row):
            pk_field = model._meta.pk
            pk_name = pk_field.name
            if pk_name in row and row.get(pk_name) not in (None, ""):
                return {pk_name: parse_field_value(pk_field, row.get(pk_name))}
            for fields in lookup_fields.get(model_key, []):
                values = {}
                ok = True
                for field_name in fields:
                    field = model._meta.get_field(field_name)
                    if isinstance(field, dj_models.ForeignKey):
                        value = resolve_fk_id(model_key, field, row)
                    else:
                        value = parse_field_value(field, row.get(field_name))
                    if value in (None, ""):
                        ok = False
                        break
                    values[field_name] = value
                if ok:
                    return values
            return None

        def upsert_rows(model_key, model, rows):
            count = 0
            for row in rows:
                try:
                    lookup = build_lookup(model_key, model, row)
                    if not lookup:
                        errors.append(f"{model_key}: missing lookup for row")
                        continue

                    defaults = {}
                    for field in model._meta.fields:
                        if field.primary_key:
                            continue
                        if getattr(field, "auto_now", False) or getattr(field, "auto_now_add", False):
                            continue
                        if isinstance(field, dj_models.ForeignKey):
                            fk_id = resolve_fk_id(model_key, field, row)
                            if fk_id is not None:
                                defaults[f"{field.name}_id"] = fk_id
                            else:
                                defaults[field.name] = None
                        else:
                            if field.name in row:
                                defaults[field.name] = parse_field_value(field, row.get(field.name))

                    model.objects.update_or_create(defaults=defaults, **lookup)
                    count += 1
                except Exception as exc:
                    errors.append(f"{model_key}: {exc}")
            imported_count[model_key] = count

        # --- Import models in dependency order ---
        from finance.models import EmployeeSalary, SundryCreditor, SundryDebtor, Loan, CreditorTransaction, DebtorTransaction, Expense, Employee
        from goldsilverpurchase.models import GoldSilverPurchase, CustomerPurchase, MetalStock, MetalStockMovement, Party, MetalStockType
        from main.models import Stock, DailyRate
        from ornament.models import Stone, Motimala, Potey, MainCategory, SubCategory, Kaligar, Kaligar_Ornaments, Kaligar_CashAccount, Kaligar_GoldAccount, Ornament
        from order.models import Order, OrderMetalStock, OrderPayment, OrderOrnament, DebtorPayment
        from sales.models import Sale, SalesMetalStock

        model_order = [
            ("Employee", Employee),
            ("SundryCreditor", SundryCreditor),
            ("SundryDebtor", SundryDebtor),
            ("Loan", Loan),
            ("Expense", Expense),
            ("EmployeeSalary", EmployeeSalary),
            ("CreditorTransaction", CreditorTransaction),
            ("DebtorTransaction", DebtorTransaction),
            ("Party", Party),
            ("MetalStockType", MetalStockType),
            ("MetalStock", MetalStock),
            ("GoldSilverPurchase", GoldSilverPurchase),
            ("CustomerPurchase", CustomerPurchase),
            ("Kaligar", Kaligar),
            ("MetalStockMovement", MetalStockMovement),
            ("Stock", Stock),
            ("DailyRate", DailyRate),
            ("MainCategory", MainCategory),
            ("SubCategory", SubCategory),
            ("Stone", Stone),
            ("Motimala", Motimala),
            ("Potey", Potey),
            ("Kaligar_Ornaments", Kaligar_Ornaments),
            ("Kaligar_CashAccount", Kaligar_CashAccount),
            ("Kaligar_GoldAccount", Kaligar_GoldAccount),
            ("Order", Order),
            ("Ornament", Ornament),
            ("OrderMetalStock", OrderMetalStock),
            ("OrderPayment", OrderPayment),
            ("OrderOrnament", OrderOrnament),
            ("DebtorPayment", DebtorPayment),
            ("Sale", Sale),
            ("SalesMetalStock", SalesMetalStock),
        ]

        with transaction.atomic():
            for key, model in model_order:
                rows = data.get(key, [])
                if rows:
                    upsert_rows(key, model, rows)

        summary = " | ".join([f"{k}: {v}" for k, v in imported_count.items() if v > 0])
        if summary:
            messages.success(request, f"JSON import completed: {summary}")
        else:
            messages.success(request, "JSON import completed.")

        if errors:
            error_preview = "; ".join(errors[:5])
            extra = f" ... and {len(errors) - 5} more" if len(errors) > 5 else ""
            messages.warning(request, f"Some rows were skipped: {error_preview}{extra}")

        return redirect("gsp:data_settings")

    except json.JSONDecodeError:
        messages.error(request, "Invalid JSON file format.")
        return redirect("gsp:data_settings")
    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect("gsp:data_settings")


@login_required(login_url='/accounts/login/')
def delete_all_data(request):
    """Delete all data from ALL tables across all apps (with confirmation)."""
    if request.method != "POST":
        return redirect("gsp:data_settings")

    # Verify confirmation token to prevent accidental deletion
    confirm_token = request.POST.get("confirm_delete")
    if confirm_token != "DELETE_ALL_DATA_CONFIRMED":
        messages.error(request, "Deletion not confirmed. Data was not deleted.")
        return redirect("gsp:data_settings")

    try:
        # Import all models
        from finance.models import (
            EmployeeSalary, Expense, Employee,
            CreditorTransaction, DebtorTransaction,
            SundryCreditor, SundryDebtor, Loan
        )
        from goldsilverpurchase.models import (
            MetalStockMovement, MetalStock, CustomerPurchase, 
            GoldSilverPurchase, Party, MetalStockType
        )
        from main.models import DailyRate, Stock
        from ornament.models import (
            Kaligar_CashAccount, Kaligar_GoldAccount, Kaligar_Ornaments,
            Ornament, Kaligar, SubCategory, MainCategory,
            Potey, Motimala, Stone
        )
        from order.models import (
            DebtorPayment, OrderPayment, OrderOrnament, OrderMetalStock, Order
        )
        from sales.models import SalesMetalStock, Sale
        
        deleted_counts = {}
        
        # Delete in reverse dependency order (children first, parents last)
        
        # Sales (dependent on Order)
        deleted_counts['SalesMetalStock'] = SalesMetalStock.objects.count()
        SalesMetalStock.objects.all().delete()
        
        deleted_counts['Sale'] = Sale.objects.count()
        Sale.objects.all().delete()
        
        # Order dependencies
        deleted_counts['DebtorPayment'] = DebtorPayment.objects.count()
        DebtorPayment.objects.all().delete()
        
        deleted_counts['OrderPayment'] = OrderPayment.objects.count()
        OrderPayment.objects.all().delete()
        
        deleted_counts['OrderOrnament'] = OrderOrnament.objects.count()
        OrderOrnament.objects.all().delete()
        
        deleted_counts['OrderMetalStock'] = OrderMetalStock.objects.count()
        OrderMetalStock.objects.all().delete()
        
        deleted_counts['Order'] = Order.objects.count()
        Order.objects.all().delete()
        
        # Ornament models
        deleted_counts['Kaligar_CashAccount'] = Kaligar_CashAccount.objects.count()
        Kaligar_CashAccount.objects.all().delete()
        
        deleted_counts['Kaligar_GoldAccount'] = Kaligar_GoldAccount.objects.count()
        Kaligar_GoldAccount.objects.all().delete()
        
        deleted_counts['Kaligar_Ornaments'] = Kaligar_Ornaments.objects.count()
        Kaligar_Ornaments.objects.all().delete()
        
        deleted_counts['Ornament'] = Ornament.objects.count()
        Ornament.objects.all().delete()
        
        deleted_counts['SubCategory'] = SubCategory.objects.count()
        SubCategory.objects.all().delete()
        
        deleted_counts['MainCategory'] = MainCategory.objects.count()
        MainCategory.objects.all().delete()
        
        deleted_counts['Kaligar'] = Kaligar.objects.count()
        Kaligar.objects.all().delete()
        
        deleted_counts['Potey'] = Potey.objects.count()
        Potey.objects.all().delete()
        
        deleted_counts['Motimala'] = Motimala.objects.count()
        Motimala.objects.all().delete()
        
        deleted_counts['Stone'] = Stone.objects.count()
        Stone.objects.all().delete()
        
        # Main models
        deleted_counts['DailyRate'] = DailyRate.objects.count()
        DailyRate.objects.all().delete()
        
        deleted_counts['Stock'] = Stock.objects.count()
        Stock.objects.all().delete()
        
        # GoldSilverPurchase models
        deleted_counts['MetalStockMovement'] = MetalStockMovement.objects.count()
        MetalStockMovement.objects.all().delete()
        
        deleted_counts['MetalStock'] = MetalStock.objects.count()
        MetalStock.objects.all().delete()
        
        deleted_counts['CustomerPurchase'] = CustomerPurchase.objects.count()
        CustomerPurchase.objects.all().delete()
        
        deleted_counts['GoldSilverPurchase'] = GoldSilverPurchase.objects.count()
        GoldSilverPurchase.objects.all().delete()
        
        deleted_counts['MetalStockType'] = MetalStockType.objects.count()
        MetalStockType.objects.all().delete()
        
        deleted_counts['Party'] = Party.objects.count()
        Party.objects.all().delete()
        
        # Finance models (transactions before parent records)
        deleted_counts['CreditorTransaction'] = CreditorTransaction.objects.count()
        CreditorTransaction.objects.all().delete()
        
        deleted_counts['DebtorTransaction'] = DebtorTransaction.objects.count()
        DebtorTransaction.objects.all().delete()
        
        deleted_counts['EmployeeSalary'] = EmployeeSalary.objects.count()
        EmployeeSalary.objects.all().delete()
        
        deleted_counts['SundryCreditor'] = SundryCreditor.objects.count()
        SundryCreditor.objects.all().delete()
        
        deleted_counts['SundryDebtor'] = SundryDebtor.objects.count()
        SundryDebtor.objects.all().delete()
        
        deleted_counts['Loan'] = Loan.objects.count()
        Loan.objects.all().delete()
        
        deleted_counts['Expense'] = Expense.objects.count()
        Expense.objects.all().delete()
        
        deleted_counts['Employee'] = Employee.objects.count()
        Employee.objects.all().delete()
        
        # Calculate total
        total_deleted = sum(deleted_counts.values())
        
        # Build detailed message
        details = " | ".join([f"{model}: {count}" for model, count in deleted_counts.items() if count > 0])
        
        messages.success(
            request,
            f"All data deleted successfully! Total records removed: {total_deleted}. Details: {details}"
        )
        
    except Exception as e:
        messages.error(request, f"Error deleting data: {str(e)}")

    return redirect("gsp:data_settings")

class MetalStockListView(LoginRequiredMixin, ListView):
    """View to display all metal stocks (raw and refined)"""
    model = MetalStock
    template_name = 'goldsilverpurchase/metalstock_list.html'
    context_object_name = 'metal_stocks'
    paginate_by = 15

    def get_queryset(self):
        queryset = super().get_queryset().order_by('-last_updated', '-created_at')
        
        # Filter by metal type
        metal_type = self.request.GET.get('metal_type')
        if metal_type:
            queryset = queryset.filter(metal_type=metal_type)
        
        # Filter by stock type (raw/refined)
        stock_type = self.request.GET.get('stock_type')
        if stock_type:
            queryset = queryset.filter(stock_type__name=stock_type)
        
        # Filter by purity
        purity = self.request.GET.get('purity')
        if purity:
            queryset = queryset.filter(purity=purity)
        
        # Search by location
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(location__icontains=search) |
                Q(remarks__icontains=search)
            )
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get the filtered base queryset
        filtered_queryset = self.get_queryset()
        
        # Summary statistics (always use unfiltered totals for dashboard)
        context['total_gold_quantity'] = (
            MetalStock.objects.filter(metal_type='gold')
            .aggregate(total=Sum('quantity'))['total'] or Decimal('0.000')
        )
        context['total_silver_quantity'] = (
            MetalStock.objects.filter(metal_type='silver')
            .aggregate(total=Sum('quantity'))['total'] or Decimal('0.000')
        )
        context['total_platinum_quantity'] = (
            MetalStock.objects.filter(metal_type='platinum')
            .aggregate(total=Sum('quantity'))['total'] or Decimal('0.000')
        )
        
        # Total value
        gold_value = MetalStock.objects.filter(metal_type='gold').aggregate(
            total=Sum(models.F('quantity') * models.F('unit_cost'), output_field=models.DecimalField())
        )['total'] or Decimal('0.00')
        
        silver_value = MetalStock.objects.filter(metal_type='silver').aggregate(
            total=Sum(models.F('quantity') * models.F('unit_cost'), output_field=models.DecimalField())
        )['total'] or Decimal('0.00')
        
        context['total_gold_value'] = gold_value
        context['total_silver_value'] = silver_value
        context['total_value'] = gold_value + silver_value
        
        # Breakdown by stock type
        context['raw_stocks'] = MetalStock.objects.filter(
            stock_type__name='raw'
        ).aggregate(total=Sum('quantity'))['total'] or Decimal('0.000')
        
        context['refined_stocks'] = MetalStock.objects.filter(
            stock_type__name='refined'
        ).aggregate(total=Sum('quantity'))['total'] or Decimal('0.000')
        
        # Low stock alerts (from filtered queryset)
        context['low_stock_items'] = [
            item for item in filtered_queryset if item.is_low_stock
        ]
        
        # Split filtered queryset into gold and silver for tabs
        gold_stocks = filtered_queryset.filter(metal_type='gold')
        silver_stocks = filtered_queryset.filter(metal_type='silver')
        
        context['gold_stocks'] = gold_stocks
        context['silver_stocks'] = silver_stocks
        
        # Average unit cost rates from MetalStock (weighted by quantity)
        # 1 tola = 11.6643 grams
        TOLA_TO_GRAM = Decimal('11.6643')
        
        # Function to calculate weighted average cost in tola
        def calculate_weighted_average_cost(stock_qs):
            """Calculate quantity-weighted average unit cost in per-tola"""
            total_value_per_gram = Decimal('0.00')
            total_quantity = Decimal('0.00')
            
            for stock in stock_qs:
                # Skip if no cost or quantity
                if not stock.unit_cost or not stock.quantity:
                    continue
                    
                # Convert unit cost to per-gram based on rate_unit
                if stock.rate_unit == 'gram':
                    per_gram = stock.unit_cost
                elif stock.rate_unit == '10gram':
                    per_gram = stock.unit_cost / Decimal('10')
                elif stock.rate_unit == 'tola':
                    per_gram = stock.unit_cost / TOLA_TO_GRAM
                else:
                    continue
                
                # Accumulate: (cost_per_gram × quantity)
                total_value_per_gram += per_gram * stock.quantity
                total_quantity += stock.quantity
            
            if total_quantity == 0:
                return Decimal('0.00')
            
            # Weighted average per-gram = total_value / total_quantity
            avg_per_gram = total_value_per_gram / total_quantity
            # Convert per-gram to per-tola
            avg_per_tola = avg_per_gram * TOLA_TO_GRAM
            return avg_per_tola
        
        gold_avg_rate_tola = calculate_weighted_average_cost(gold_stocks)
        silver_avg_rate_tola = calculate_weighted_average_cost(silver_stocks)
        
        context['avg_gold_purchase_rate_tola'] = gold_avg_rate_tola
        context['avg_silver_purchase_rate_tola'] = silver_avg_rate_tola
        
        return context


@login_required(login_url='/accounts/login/')
def metal_stock_detail(request, pk):
    """View stock details and its movement history"""
    # Always re-fetch the MetalStock from the DB to get the latest quantity
    from django.db.models import Sum
    metal_stock = MetalStock.objects.get(pk=pk)
    # Recalculate quantity from all movements (defensive, in case of any desync)
    total_in = metal_stock.movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
    total_out = metal_stock.movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
    total_adj = metal_stock.movements.filter(movement_type='adjustment').aggregate(total=Sum('quantity'))['total'] or 0
    metal_stock.quantity = total_in - total_out + total_adj
    metal_stock.save()
    movements = MetalStockMovement.objects.filter(metal_stock=metal_stock).order_by('-movement_date')
    # Calculate average unit cost from 'in' movements
    in_movements = movements.filter(movement_type='in')
    total_qty = 0
    total_cost = 0
    for m in in_movements:
        if m.rate and m.quantity:
            total_qty += float(m.quantity)
            total_cost += float(m.rate) * float(m.quantity)
    avg_unit_cost = (total_cost / total_qty) if total_qty else 0
    context = {
        'metal_stock': metal_stock,
        'movements': movements,
        'avg_unit_cost_from_movements': avg_unit_cost,
    }
    return render(request, 'goldsilverpurchase/metalstock_detail.html', context)


class MetalStockCreateView(LoginRequiredMixin, CreateView):
    """Create a new metal stock"""
    model = MetalStock
    form_class = MetalStockForm
    template_name = 'goldsilverpurchase/metalstock_form.html'
    success_url = reverse_lazy('gsp:metal_stock_list')

    def form_valid(self, form):
        # Always save MetalStock first to get a primary key
        self.object = form.save(commit=False)
        self.object.save()  # Ensure PK exists
        add_quantity = form.cleaned_data.get('add_quantity')
        movement_notes = form.cleaned_data.get('movement_notes')
        if add_quantity not in (None, '') and add_quantity > 0:
            # Update quantity and save again
            self.object.quantity += add_quantity
            self.object.save()
            MetalStockMovement.objects.create(
                metal_stock=self.object,
                movement_type='in',
                quantity=add_quantity,
                rate=self.object.unit_cost,
                reference_type='Manual',
                reference_id=None,
                notes=movement_notes or 'Initial stock added on creation.'
            )
        messages.success(self.request, f"Metal stock for {form.cleaned_data['metal_type']} created successfully!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Metal Stock'
        context['button_text'] = 'Create Stock'
        return context


class MetalStockUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing metal stock"""
    model = MetalStock
    form_class = MetalStockForm
    template_name = 'goldsilverpurchase/metalstock_form.html'
    success_url = reverse_lazy('gsp:metal_stock_list')

    def form_valid(self, form):
        messages.success(self.request, f"Metal stock updated successfully!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Update Metal Stock'
        context['button_text'] = 'Update Stock'
        return context


class MetalStockDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a metal stock"""
    model = MetalStock
    template_name = 'goldsilverpurchase/metalstock_confirm_delete.html'
    success_url = reverse_lazy('gsp:metal_stock_list')

    def delete(self, request, *args, **kwargs):
        metal_stock = self.get_object()
        metal_type = metal_stock.get_metal_type_display()
        messages.success(request, f"Metal stock for {metal_type} has been deleted successfully!")
        return super().delete(request, *args, **kwargs)

from django.urls import reverse
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from .models import MetalStockMovement

from .forms import MetalStockForm
from .forms_movement import MetalStockMovementForm


class MetalStockMovementCreateView(LoginRequiredMixin, CreateView):
    model = MetalStockMovement
    form_class = MetalStockMovementForm
    template_name = 'goldsilverpurchase/metalstockmovement_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.metal_stock = MetalStock.objects.get(pk=kwargs['stock_pk'])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.metal_stock = self.metal_stock
        response = super().form_valid(form)
        # Update MetalStock quantity
        if form.instance.movement_type == 'in':
            self.metal_stock.quantity += form.instance.quantity
        elif form.instance.movement_type == 'out':
            self.metal_stock.quantity -= form.instance.quantity
        self.metal_stock.save()
        return response

    def get_success_url(self):
        return reverse('gsp:metal_stock_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['metal_stock'] = self.metal_stock
        return context



class MetalStockMovementUpdateView(LoginRequiredMixin, UpdateView):
    model = MetalStockMovement
    form_class = MetalStockMovementForm
    template_name = 'goldsilverpurchase/metalstockmovement_form.html'

    def form_valid(self, form):
        response = super().form_valid(form)
        # After saving, recalculate the stock quantity from all movements
        metal_stock = self.object.metal_stock
        total_in = metal_stock.movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        total_out = metal_stock.movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        total_adj = metal_stock.movements.filter(movement_type='adjustment').aggregate(total=Sum('quantity'))['total'] or 0
        metal_stock.quantity = total_in - total_out + total_adj
        metal_stock.save()
        return response

    def get_success_url(self):
        return reverse('gsp:metal_stock_detail', kwargs={'pk': self.object.metal_stock_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['metal_stock'] = self.object.metal_stock
        context['is_edit'] = True
        return context


class MetalStockMovementDeleteView(LoginRequiredMixin, DeleteView):
    model = MetalStockMovement
    template_name = 'goldsilverpurchase/metalstockmovement_confirm_delete.html'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        metal_stock_id = self.object.metal_stock_id
        response = super().delete(request, *args, **kwargs)
        # After deletion, recalculate the stock quantity from all remaining movements
        from .models import MetalStock
        from django.db.models import Sum
        metal_stock = MetalStock.objects.get(pk=metal_stock_id)
        metal_stock_id = self.get_object().metal_stock_id
        response = super().delete(request, *args, **kwargs)
        from .models import MetalStock
        from django.db.models import Sum
        metal_stock = MetalStock.objects.get(pk=metal_stock_id)
        movements = metal_stock.movements.all()
        if not movements.exists():
            metal_stock.quantity = 0
        else:
            total_in = movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0
            total_out = movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0
            total_adj = movements.filter(movement_type='adjustment').aggregate(total=Sum('quantity'))['total'] or 0
            metal_stock.quantity = total_in - total_out + total_adj
        metal_stock.save()
        from django.urls import reverse
        return redirect(reverse('gsp:metal_stock_detail', kwargs={'pk': metal_stock_id}))

    def get_success_url(self):
        return reverse('gsp:metal_stock_detail', kwargs={'pk': self.object.metal_stock_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['metal_stock'] = self.object.metal_stock
        return context


@login_required(login_url='/accounts/login/')
def import_wizard(request):
    """Display the import wizard interface."""
    return render(request, 'goldsilverpurchase/import_wizard.html')


@login_required(login_url='/accounts/login/')
def import_wizard_process(request):
    """Process the import with progress tracking."""
    import json
    import openpyxl
    from django.http import JsonResponse
    
    if request.method != "POST" or "import_file" not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file uploaded'})

    file = request.FILES.get("import_file")
    skip_existing = request.POST.get('skip_existing') == 'on'
    
    if not file:
        return JsonResponse({'success': False, 'error': 'No file provided'})

    imported_count = {}
    skipped_count = {}
    errors = []

    try:
        # Determine file type
        file_ext = file.name.split('.')[-1].lower()
        
        if file_ext == 'json':
            # JSON Import
            file_content = file.read().decode('utf-8')
            data = json.loads(file_content)
            
            # Import in dependency order
            import_order = [
                # Finance (no dependencies)
                'Employee', 'Expense',
                # Finance (with Employee FK)
                'EmployeeSalary',
                # Finance (independent)
                'SundryCreditor', 'SundryDebtor', 'Loan',
                # Finance transactions
                'CreditorTransaction', 'DebtorTransaction',
                # GoldSilverPurchase
                'Party', 'MetalStockType',
                'GoldSilverPurchase', 'CustomerPurchase',
                'MetalStock', 'MetalStockMovement',
                # Main
                'Stock', 'DailyRate',
                # Ornament
                'MainCategory', 'SubCategory', 'Stone', 'Motimala', 'Potey',
                'Kaligar', 'Ornament',
                'Kaligar_Ornaments', 'Kaligar_CashAccount', 'Kaligar_GoldAccount',
                # Orders
                'Order', 'OrderMetalStock', 'OrderPayment', 'OrderOrnament',
                'DebtorPayment',
                # Sales
                'Sale', 'SalesMetalStock',
            ]
            
            for model_name in import_order:
                if model_name not in data:
                    continue
                    
                count, skipped, model_errors = import_model_json(model_name, data[model_name], skip_existing)
                if count > 0:
                    imported_count[model_name] = count
                if skipped > 0:
                    skipped_count[model_name] = skipped
                errors.extend(model_errors)
        
        elif file_ext == 'xlsx':
            # XLSX Import
            wb = openpyxl.load_workbook(file)
            
            # Import in dependency order (same as JSON)
            import_order = [
                'Employee', 'Expense', 'EmployeeSalary',
                'SundryCreditor', 'SundryDebtor', 'Loan',
                'CreditorTransaction', 'DebtorTransaction',
                'Party', 'MetalStockType',
                'GoldSilverPurchase', 'CustomerPurchase',
                'MetalStock', 'MetalStockMovement',
                'Stock', 'DailyRate',
                'MainCategory', 'SubCategory', 'Stone', 'Motimala', 'Potey',
                'Kaligar', 'Ornament',
                'Kaligar_Ornaments', 'Kaligar_CashAccount', 'Kaligar_GoldAccount',
                'Order', 'OrderMetalStock', 'OrderPayment', 'OrderOrnament',
                'DebtorPayment',
                'Sale', 'SalesMetalStock',
            ]
            
            for sheet_name in import_order:
                if sheet_name not in wb.sheetnames:
                    continue
                    
                count, skipped, model_errors = import_model_xlsx(sheet_name, wb[sheet_name], skip_existing)
                if count > 0:
                    imported_count[sheet_name] = count
                if skipped > 0:
                    skipped_count[sheet_name] = skipped
                errors.extend(model_errors)
        
        else:
            return JsonResponse({'success': False, 'error': 'Unsupported file format'})

        return JsonResponse({
            'success': True,
            'imported': imported_count,
            'skipped': skipped_count,
            'errors': errors
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def import_model_json(model_name, records, skip_existing=True):
    """Import a single model from JSON data."""
    from decimal import Decimal
    from datetime import datetime
    from finance.models import Employee, Expense, EmployeeSalary, SundryCreditor, SundryDebtor, Loan, CreditorTransaction, DebtorTransaction
    from goldsilverpurchase.models import Party, MetalStockType, GoldSilverPurchase, CustomerPurchase, MetalStock, MetalStockMovement
    from main.models import Stock, DailyRate
    from ornament.models import Stone, Motimala, Potey, MainCategory, SubCategory, Kaligar, Ornament, Kaligar_Ornaments, Kaligar_CashAccount, Kaligar_GoldAccount
    from order.models import Order, OrderMetalStock, OrderPayment, OrderOrnament, DebtorPayment
    from sales.models import Sale, SalesMetalStock
    
    count = 0
    skipped = 0
    errors = []
    
    # Model mapping
    model_map = {
        'Employee': Employee,
        'Expense': Expense,
        'EmployeeSalary': EmployeeSalary,
        'SundryCreditor': SundryCreditor,
        'SundryDebtor': SundryDebtor,
        'Loan': Loan,
        'CreditorTransaction': CreditorTransaction,
        'DebtorTransaction': DebtorTransaction,
        'Party': Party,
        'MetalStockType': MetalStockType,
        'GoldSilverPurchase': GoldSilverPurchase,
        'CustomerPurchase': CustomerPurchase,
        'MetalStock': MetalStock,
        'MetalStockMovement': MetalStockMovement,
        'Stock': Stock,
        'DailyRate': DailyRate,
        'Stone': Stone,
        'Motimala': Motimala,
        'Potey': Potey,
        'MainCategory': MainCategory,
        'SubCategory': SubCategory,
        'Kaligar': Kaligar,
        'Ornament': Ornament,
        'Kaligar_Ornaments': Kaligar_Ornaments,
        'Kaligar_CashAccount': Kaligar_CashAccount,
        'Kaligar_GoldAccount': Kaligar_GoldAccount,
        'Order': Order,
        'OrderMetalStock': OrderMetalStock,
        'OrderPayment': OrderPayment,
        'OrderOrnament': OrderOrnament,
        'DebtorPayment': DebtorPayment,
        'Sale': Sale,
        'SalesMetalStock': SalesMetalStock,
    }
    
    if model_name not in model_map:
        return 0, 0, [f"Unknown model: {model_name}"]
    
    Model = model_map[model_name]
    
    for record in records:
        try:
            # Convert Decimal strings back to Decimal
            for key, value in record.items():
                if isinstance(value, str):
                    try:
                        record[key] = Decimal(value) if '.' in value else value
                    except:
                        pass
            
            # Check if record exists (by ID)
            if skip_existing and 'id' in record:
                if Model.objects.filter(id=record['id']).exists():
                    skipped += 1
                    continue
            
            # Create the record
            obj, created = Model.objects.update_or_create(
                id=record.get('id'),
                defaults={k: v for k, v in record.items() if k != 'id'}
            )
            if created:
                count += 1
            else:
                skipped += 1
        except Exception as e:
            errors.append(f"{model_name} - Record {record.get('id', '?')}: {str(e)}")
    
    return count, skipped, errors


def import_model_xlsx(sheet_name, worksheet, skip_existing=True):
    """Import a single model from XLSX sheet."""
    from decimal import Decimal
    from datetime import datetime
    from finance.models import Employee, Expense, EmployeeSalary, SundryCreditor, SundryDebtor, Loan, CreditorTransaction, DebtorTransaction
    from goldsilverpurchase.models import Party, MetalStockType, GoldSilverPurchase, CustomerPurchase, MetalStock, MetalStockMovement
    from main.models import Stock, DailyRate
    from ornament.models import Stone, Motimala, Potey, MainCategory, SubCategory, Kaligar, Ornament, Kaligar_Ornaments, Kaligar_CashAccount, Kaligar_GoldAccount
    from order.models import Order, OrderMetalStock, OrderPayment, OrderOrnament, DebtorPayment
    from sales.models import Sale, SalesMetalStock
    
    count = 0
    skipped = 0
    errors = []
    
    # Model mapping
    model_map = {
        'Employee': Employee,
        'Expense': Expense,
        'EmployeeSalary': EmployeeSalary,
        'SundryCreditor': SundryCreditor,
        'SundryDebtor': SundryDebtor,
        'Loan': Loan,
        'CreditorTransaction': CreditorTransaction,
        'DebtorTransaction': DebtorTransaction,
        'Party': Party,
        'MetalStockType': MetalStockType,
        'GoldSilverPurchase': GoldSilverPurchase,
        'CustomerPurchase': CustomerPurchase,
        'MetalStock': MetalStock,
        'MetalStockMovement': MetalStockMovement,
        'Stock': Stock,
        'DailyRate': DailyRate,
        'Stone': Stone,
        'Motimala': Motimala,
        'Potey': Potey,
        'MainCategory': MainCategory,
        'SubCategory': SubCategory,
        'Kaligar': Kaligar,
        'Ornament': Ornament,
        'Kaligar_Ornaments': Kaligar_Ornaments,
        'Kaligar_CashAccount': Kaligar_CashAccount,
        'Kaligar_GoldAccount': Kaligar_GoldAccount,
        'Order': Order,
        'OrderMetalStock': OrderMetalStock,
        'OrderPayment': OrderPayment,
        'OrderOrnament': OrderOrnament,
        'DebtorPayment': DebtorPayment,
        'Sale': Sale,
        'SalesMetalStock': SalesMetalStock,
    }
    
    if sheet_name not in model_map:
        return 0, 0, [f"Unknown model: {sheet_name}"]
    
    Model = model_map[sheet_name]
    
    # Get headers from first row
    headers = []
    for cell in worksheet[1]:
        if cell.value is not None:
            headers.append(cell.value)
    
    # Process data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        try:
            record = {}
            for col_idx, header in enumerate(headers):
                value = row[col_idx] if col_idx < len(row) else None
                
                # Convert value types
                if value is not None:
                    if isinstance(value, str):
                        try:
                            # Try to convert to Decimal if it's a number
                            if '.' in value or value.isdigit():
                                record[header] = Decimal(value)
                            else:
                                record[header] = value
                        except:
                            record[header] = value
                    else:
                        record[header] = value
            
            if not record:
                continue
            
            # Check if record exists (by ID)
            if skip_existing and 'ID' in record:
                if Model.objects.filter(id=record['ID']).exists():
                    skipped += 1
                    continue
            
            # Create the record
            record_id = record.pop('ID', None)
            obj, created = Model.objects.update_or_create(
                id=record_id if record_id else None,
                defaults=record
            )
            if created:
                count += 1
            else:
                skipped += 1
        except Exception as e:
            errors.append(f"{sheet_name} - Row {row_idx}: {str(e)}")
    
    return count, skipped, errors
