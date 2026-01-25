from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView
from django.views.generic.edit import DeleteView
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.db.models import Q, Sum, F
from decimal import Decimal
from datetime import timedelta
import json
import nepali_datetime as ndt
from io import BytesIO
from django.contrib import messages
import openpyxl
from openpyxl.utils import get_column_letter

from .models import Order, OrderOrnament, OrderPayment, OrderMetalStock
from sales.models import Sale
from .forms import OrderForm, OrnamentFormSet, MetalStockFormSet
from ornament.models import Ornament, Kaligar
from goldsilverpurchase.models import MetalStock, MetalStockMovement

app_name = 'order'

class SearchOrnamentsAPI(View):
    def get(self, request):
        query = request.GET.get('q', '').strip()

        # Show only stock ornaments, never show order-type ornaments
        ornament_types = [Ornament.OrnamentCategory.STOCK]

        ornaments = Ornament.objects.filter(
            ornament_type__in=ornament_types,
        ).select_related('order').order_by('-id')
        
        if query:
            ornaments = ornaments.filter(
                Q(ornament_name__icontains=query) |
                Q(code__icontains=query) |
                Q(weight__icontains=query)
            )
        
        data = []
        for ornament in ornaments[:50]:  # Increased limit to 50 results
            data.append({
                'id': ornament.id,
                'code': ornament.code or 'N/A',
                'name': ornament.ornament_name,
                'metal_type': ornament.metal_type,
                'type': ornament.type,
                'weight': float(ornament.weight or 0),
                'jarti': float(ornament.jarti or 0),
                'jyala': float(ornament.jyala or 0),
                'gross_weight': float(ornament.gross_weight or 0),
                'stone_weight': float(ornament.stone_weight or 0),
                'diamond_weight': float(ornament.diamond_weight or 0),
                'zircon_weight': float(ornament.zircon_weight or 0),
                'amount': float(ornament.stone_totalprice or 0),
                'diamond_rate': 0.0,
                'zircon_rate': 0.0,
                'stone_rate': 0.0,
                'gold_rate': 0.0,
                'ornament_type': ornament.ornament_type,
                'order_id': ornament.order.pk if ornament.order else None,
                'order_sn': ornament.order.sn if ornament.order else None,
                'order_customer': getattr(ornament.order, 'customer_name', None) if ornament.order else None,
            })
        
        return JsonResponse({'ornaments': data})


class CreateOrnamentInlineView(View):
    """Create an ornament quickly while creating an order (AJAX)."""

    def post(self, request):
        data = request.POST

        def D(val):
            try:
                return Decimal(str(val))
            except Exception:
                return Decimal("0")

        kaligar_id = data.get("kaligar")
        try:
            kaligar = Kaligar.objects.get(pk=kaligar_id)
        except Kaligar.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Kaligar is required."}, status=400)

        ornament = Ornament.objects.create(
            ornament_date=ndt.date.today(),
            code=data.get("code") or None,
            ornament_name=data.get("ornament_name") or "",
            metal_type=data.get("metal_type") or Ornament.MetalTypeCategory.GOLD,
            ornament_type=Ornament.OrnamentCategory.ORDER,
            weight=D(data.get("weight")),
            diamond_weight=D(data.get("diamond_weight")),
            zircon_weight=D(data.get("zircon_weight")),
            stone_weight=D(data.get("stone_weight")),
            jarti=D(data.get("jarti")),
            jyala=D(data.get("jyala")),
            kaligar=kaligar,
        )

        return JsonResponse(
            {
                "ok": True,
                "ornament": {
                    "id": ornament.id,
                    "code": ornament.code or "N/A",
                    "name": ornament.ornament_name,
                    "metal_type": ornament.metal_type,
                    "weight": float(ornament.weight or 0),
                    "jarti": float(ornament.jarti or 0),
                    "jyala": float(ornament.jyala or 0),
                    "gross_weight": float(ornament.gross_weight or 0),
                    "stone_weight": float(ornament.stone_weight or 0),
                    "diamond_weight": float(ornament.diamond_weight or 0),
                    "zircon_weight": float(ornament.zircon_weight or 0),
                    "amount": float(ornament.stone_totalprice or 0),
                    "gold_rate": 0.0,
                    "diamond_rate": 0.0,
                    "zircon_rate": 0.0,
                    "stone_rate": 0.0,
                },
            }
        )


class OrnamentCreateView(View):
    template_name = 'ornament/ornament_form.html'
    success_url = reverse_lazy('ornament:list')

    def get(self, request):
        formset = OrnamentFormSet(queryset=Ornament.objects.none())
        return render(request, self.template_name, {"formset": formset})

    def post(self, request):
        formset = OrnamentFormSet(request.POST)

        if formset.is_valid():
            formset.save()
            return redirect(self.success_url)

        return render(request, self.template_name, {"formset": formset})

class OrderListView(ListView):
    model = Order
    template_name = 'order/order_list.html'
    context_object_name = 'orders'
    ordering = ['-order_date', '-sn']
    paginate_by = 25

    def get_queryset(self):
        """Show only orders that have not yet been converted to sales."""
        qs = super().get_queryset()
        return (
            qs.filter(sale__isnull=True)
            .annotate(total_weight=Sum('order_ornaments__ornament__weight'))
            .prefetch_related(
                'order_ornaments__ornament',
                'payments',
                'ornaments',  # ornament FK on Ornament model if present
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Get unpaginated queryset for aggregations and calculations
        qs = self.get_queryset()  # Fresh queryset without pagination

        # Group orders by status
        orders_by_status = {}
        status_choices = {choice[0]: choice[1] for choice in Order.STATUS_CHOICES}
        
        for status_key in status_choices:
            orders_by_status[status_key] = qs.filter(status=status_key).order_by('-order_date', '-sn')
        
        ctx['orders_by_status'] = orders_by_status
        ctx['status_choices'] = status_choices

        aggregates = qs.aggregate(
            total_remaining=Sum('remaining_amount'),
            total_amount=Sum('amount'),
            total_total=Sum('total'),
        )

        purity_factors = {
            '24KARAT': Decimal('1.00'),
            '23KARAT': Decimal('0.99'),
            '22KARAT': Decimal('0.98'),
            '18KARAT': Decimal('0.75'),
            '14KARAT': Decimal('0.58'),
        }

        gold_lines = (
            OrderOrnament.objects.filter(order__in=qs)
            .select_related('ornament')
            .filter(ornament__metal_type=Ornament.MetalTypeCategory.GOLD)
        )

        total_24k_weight = Decimal('0')
        for line in gold_lines:
            weight = line.ornament.weight or Decimal('0')
            factor = purity_factors.get(line.ornament.type, Decimal('1.00'))
            total_24k_weight += weight * factor

        ctx['total_24k_weight'] = float(total_24k_weight)
        ctx['order_count'] = qs.count()

        # Profit proxy: total - amount (net over line base) across orders
        total_total = aggregates.get('total_total') or 0
        total_amount = aggregates.get('total_amount') or 0
        ctx['total_profit'] = float(total_total - total_amount)

        ctx['total_remaining'] = float(aggregates.get('total_remaining') or 0)

        # Calculate 24K weight for each status
        status_24k_weights = {}
        for status_key, orders_list in orders_by_status.items():
            status_24k = Decimal('0')
            status_gold_lines = (
                OrderOrnament.objects.filter(order__in=orders_list)
                .select_related('ornament')
                .filter(ornament__metal_type=Ornament.MetalTypeCategory.GOLD)
            )
            for line in status_gold_lines:
                weight = line.ornament.weight or Decimal('0')
                factor = purity_factors.get(line.ornament.type, Decimal('1.00'))
                status_24k += weight * factor
            status_24k_weights[status_key] = float(status_24k)
        
        ctx['status_24k_weights'] = status_24k_weights

        return ctx


class OrderCreateView(CreateView):
    model = Order
    form_class = OrderForm
    template_name = 'order/order_form.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['kaligars'] = Kaligar.objects.all()
        ctx['payment_choices'] = Order.PAYMENT_CHOICES
        ctx['initial_payments_json'] = json.dumps([])
        
        # Add metal stock formset for adding raw metals
        if self.request.POST:
            # For formset, we need a saved instance, but for CREATE we don't have one yet
            # So create a temporary unsaved instance for the formset
            if hasattr(self, 'object') and self.object and self.object.pk:
                instance = self.object
            else:
                instance = None
            ctx['metal_stock_formset'] = MetalStockFormSet(self.request.POST, instance=instance)
        else:
            ctx['metal_stock_formset'] = MetalStockFormSet(instance=None)
        
        # Pass formset prefix to template for use in JavaScript
        ctx['metal_formset_prefix'] = ctx['metal_stock_formset'].prefix
        
        return ctx

    def get_initial(self):
        """Provide default dates for new orders.

        - order_date: today's Nepali date
        - deliver_date: 7 days after today's Nepali date
        """
        initial = super().get_initial()
        # Only set when not already provided (e.g., from query params or other logic)
        if 'order_date' not in initial or initial['order_date'] in (None, ''):
            try:
                initial['order_date'] = ndt.date.today()
            except Exception:
                pass

        if 'deliver_date' not in initial or initial['deliver_date'] in (None, ''):
            try:
                initial['deliver_date'] = ndt.date.today() + timedelta(days=7)
            except Exception:
                pass

        return initial

    def get_success_url(self):
        return reverse_lazy('order:list')

    def form_valid(self, form):
        # Save the order first
        print(f"[DEBUG] form_valid called")
        print(f"[DEBUG] form.cleaned_data keys: {form.cleaned_data.keys()}")
        self.object = form.save()
        print(f"[DEBUG] Order saved with sn={self.object.sn}")


        # Create per-line OrderOrnament entries from JSON payload
        order_lines_raw = form.cleaned_data.get('order_lines_json') or '[]'
        try:
            lines = json.loads(order_lines_raw)
        except (TypeError, ValueError):
            lines = []

        for line in lines:
            ornament_id = line.get('ornament_id')
            if not ornament_id:
                continue
            try:
                ornament = Ornament.objects.get(pk=ornament_id)
            except Ornament.DoesNotExist:
                continue

            OrderOrnament.objects.create(
                order=self.object,
                ornament=ornament,
                gold_rate=Decimal(str(line.get('gold_rate', 0) or 0)),
                diamond_rate=Decimal(str(line.get('diamond_rate', 0) or 0)),
                zircon_rate=Decimal(str(line.get('zircon_rate', 0) or 0)),
                stone_rate=Decimal(str(line.get('stone_rate', 0) or 0)),
                jarti=Decimal(str(line.get('jarti', 0) or 0)),
                jyala=Decimal(str(line.get('jyala', 0) or 0)),
                line_amount=Decimal(str(line.get('line_amount', 0) or 0)),
            )

            # Also mark ornament as belonging to this order and flag as ORDER type
            ornament.order = self.object
            try:
                ornament.ornament_type = Ornament.OrnamentCategory.ORDER
            except AttributeError:
                ornament.ornament_type = 'order'
            ornament.save()

        # Save metal stock formset - NOW with the saved order instance
        print(f"\n=== ORDER CREATE VIEW - Processing metal formset for order {self.object.sn} ===")
        print(f"POST data keys with 'metal': {[k for k in self.request.POST.keys() if 'metal' in k][:5]}")
        metal_stock_formset = MetalStockFormSet(self.request.POST, instance=self.object)
        print(f"Formset initialized, is_valid={metal_stock_formset.is_valid()}")
        
        if metal_stock_formset.is_valid():
            # Only save forms that have data (non-empty)
            saved_metals = []
            formset_forms = list(metal_stock_formset)
            print(f"DEBUG: Formset has {len(formset_forms)} forms")
            for idx, form in enumerate(formset_forms):
                # Check if form marked for deletion
                is_delete = form.cleaned_data.get('DELETE', False)
                metal_type_val = form.cleaned_data.get('metal_type')
                print(f"  Form {idx}: metal_type={metal_type_val}, DELETE={is_delete}")
                
                # NOW check metal_type again inside the loop
                metal_type_inside = form.cleaned_data.get('metal_type')
                print(f"    >> Inside loop: metal_type={metal_type_inside}")
                
                if is_delete:
                    # Handle deleted forms - restore to stock
                    if form.instance.pk:
                        order_metal = form.instance
                        quantity = order_metal.quantity
                        metal_type = order_metal.metal_type
                        
                        try:
                            from goldsilverpurchase.models import MetalStockType
                            raw_stock_type = MetalStockType.objects.get(name__icontains='Raw')
                            
                            metal_stock = MetalStock.objects.get(
                                metal_type=metal_type,
                                purity=order_metal.purity,
                                stock_type=raw_stock_type
                            )
                            
                            # Restore quantity to metal stock
                            metal_stock.quantity += quantity
                            metal_stock.save()
                            
                            # Create a reversal movement record
                            MetalStockMovement.objects.create(
                                metal_stock=metal_stock,
                                movement_type='in',
                                quantity=quantity,
                                reference_type='Order',
                                reference_id=f"Order-{self.object.sn}",
                                notes=f"Metal removed from order {self.object.sn}"
                            )
                        except Exception as e:
                            messages.warning(self.request, f"Error restoring metal stock: {str(e)}")
                        
                        order_metal.delete()
                    continue
                
                # Check if form has actual data (metal_type is selected AND quantity is provided)
                metal_type = form.cleaned_data.get('metal_type')
                quantity = form.cleaned_data.get('quantity')
                print(f"      After deletion check, metal_type={metal_type}, quantity={quantity}...")
                
                # Skip empty rows - must have both metal_type and non-zero quantity
                if not metal_type or not quantity or quantity == 0:
                    print(f"      Skipping empty/incomplete row: metal_type={metal_type}, quantity={quantity}")
                    # If this is an existing record being cleared, delete it
                    if form.instance.pk and not metal_type:
                        form.instance.delete()
                    continue
                
                print(f"      *** ENTERED IF BLOCK FOR FORM - metal_type={metal_type}, qty={quantity} ***")
                # Check if this is a new record (no pk) before saving
                is_new = not form.instance.pk
                print(f"        is_new={is_new}")
                
                # Save the order metal entry
                print(f"        About to save metal form...")
                order_metal = form.save()
                print(f"        SAVED! order_metal.pk={order_metal.pk}")
                saved_metals.append(order_metal)
                print(f"        Appended to saved_metals")
                
                # Deduct from existing metal stock if this is a new entry
                if is_new:
                    print(f"        >> Starting stock deduction for is_new={is_new}")
                    # This is a new metal addition to the order
                    quantity = order_metal.quantity
                    metal_type = order_metal.metal_type
                    
                    # Find matching metal stock (Raw type with same metal and purity)
                    try:
                        stock_type = order_metal.stock_type
                        if not stock_type:
                            from goldsilverpurchase.models import MetalStockType
                            stock_type = MetalStockType.objects.get(name__icontains='Raw')
                        metal_stock = MetalStock.objects.get(
                            metal_type=metal_type,
                            purity=order_metal.purity,
                            stock_type=stock_type
                        )
                        # Deduct quantity from metal stock
                        metal_stock.quantity -= quantity
                        if metal_stock.quantity < 0:
                            messages.warning(
                                self.request,
                                f"Warning: {metal_stock.get_metal_type_display()} stock ({metal_stock.purity}) is now negative: {metal_stock.quantity}g"
                            )
                        metal_stock.save()
                        # Create a movement record
                        MetalStockMovement.objects.create(
                            metal_stock=metal_stock,
                            movement_type='out',
                            quantity=quantity,
                            reference_type='Order',
                            reference_id=f"Order-{self.object.sn}",
                            notes=f"Metal added to order {self.object.sn} for customer {self.object.customer_name}"
                        )
                    except MetalStock.DoesNotExist:
                        messages.warning(
                            self.request,
                            f"No matching metal stock found for {metal_type} ({order_metal.purity}, {stock_type}). Please add to inventory first."
                        )
                    except Exception as e:
                        messages.warning(self.request, f"Error updating metal stock: {str(e)}")
            
            # Delete any remaining empty instances if they exist
            for form in metal_stock_formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()
            
            if saved_metals:
                messages.success(self.request, f"Successfully added {len(saved_metals)} metal records to order")
        else:
            # Show validation errors for debugging
            if metal_stock_formset.errors:
                messages.warning(self.request, f"Metal form errors: {metal_stock_formset.errors}")
            non_form_errors = metal_stock_formset.non_form_errors()
            if non_form_errors:
                messages.warning(self.request, f"Metal stock issues: {non_form_errors}")

        # Persist payment breakdown
        payment_lines_raw = form.cleaned_data.get('payment_lines_json') or '[]'
        try:
            payments_data = json.loads(payment_lines_raw)
        except (TypeError, ValueError) as e:
            payments_data = []

        # Fallback: if cleaned_data is empty but POST has data, re-parse directly from POST
        if not payments_data:
            raw_post_val = self.request.POST.get('payment_lines_json')
            if raw_post_val:
                try:
                    payments_data = json.loads(raw_post_val)
                except Exception:
                    pass

        self.object.payments.all().delete()
        
        # If no payments or all payments have 0 amount, auto-fill with order total
        if not payments_data:
            if self.object.total > 0:
                payments_data = [{'payment_mode': 'cash', 'amount': float(self.object.total)}]
        elif all(Decimal(str(p.get('amount', 0) or 0)) <= 0 for p in payments_data):
            raw_mode = payments_data[0].get('payment_mode') or payments_data[0].get('mode') or 'cash'
            mode = raw_mode if raw_mode in dict(Order.PAYMENT_CHOICES) else 'cash'
            if self.object.total > 0:
                payments_data = [{'payment_mode': mode, 'amount': float(self.object.total)}]
        # If payments provided with positive amounts, keep them as-is (no redistribution)

        for payment in payments_data:
            amount = Decimal(str(payment.get('amount', 0) or 0))
            mode = payment.get('payment_mode') or payment.get('mode') or 'cash'
            if amount <= 0:
                continue
            if mode not in dict(Order.PAYMENT_CHOICES):
                mode = 'cash'
            OrderPayment.objects.create(
                order=self.object,
                payment_mode=mode,
                amount=amount,
            )

        # Recompute order totals from created lines (amount/subtotal/total/remaining)
        self.object.recompute_totals_from_lines()

        return redirect('order:list')
        
        
class OrderUpdateView(UpdateView):
    model = Order
    form_class = OrderForm
    template_name = 'order/order_form.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['kaligars'] = Kaligar.objects.all()
        ctx['payment_choices'] = Order.PAYMENT_CHOICES
        payments = list(self.object.payments.values('payment_mode', 'amount')) if self.object.pk else []
        for payment in payments:
            payment['amount'] = float(payment.get('amount') or 0)
        if not payments:
            # If no payments exist, provide an empty default
            payments = []
        ctx['initial_payments_json'] = json.dumps(payments)
        
        # Add metal stock formset for editing
        if self.request.POST:
            ctx['metal_stock_formset'] = MetalStockFormSet(self.request.POST, instance=self.object)
        else:
            ctx['metal_stock_formset'] = MetalStockFormSet(instance=self.object)
        
        # Pass formset prefix to template for use in JavaScript
        ctx['metal_formset_prefix'] = ctx['metal_stock_formset'].prefix
        
        return ctx

    def get_success_url(self):
        return reverse_lazy('order:list')

    def form_valid(self, form):
        # Save the order
        self.object = form.save()
        
        # DEBUG: Check if metal stock data is in POST
        print(f"DEBUG UPDATE: TOTAL_FORMS = {self.request.POST.get('ordermetal_stock_set-TOTAL_FORMS')}")
        print(f"DEBUG UPDATE: POST keys with ordermetal: {[k for k in self.request.POST.keys() if 'ordermetal' in k][:10]}")

        # Rebuild per-line OrderOrnament entries from JSON payload
        order_lines_raw = form.cleaned_data.get('order_lines_json') or '[]'
        try:
            lines = json.loads(order_lines_raw)
        except (TypeError, ValueError):
            lines = []

        # Clear existing lines
        self.object.order_ornaments.all().delete()

        # Clear existing payments (will be rebuilt below)
        self.object.payments.all().delete()

        new_ornament_ids = []
        for line in lines:
            ornament_id = line.get('ornament_id')
            if not ornament_id:
                continue
            try:
                ornament = Ornament.objects.get(pk=ornament_id)
            except Ornament.DoesNotExist:
                continue

            OrderOrnament.objects.create(
                order=self.object,
                ornament=ornament,
                gold_rate=Decimal(str(line.get('gold_rate', 0) or 0)),
                diamond_rate=Decimal(str(line.get('diamond_rate', 0) or 0)),
                zircon_rate=Decimal(str(line.get('zircon_rate', 0) or 0)),
                stone_rate=Decimal(str(line.get('stone_rate', 0) or 0)),
                jarti=Decimal(str(line.get('jarti', 0) or 0)),
                jyala=Decimal(str(line.get('jyala', 0) or 0)),
                line_amount=Decimal(str(line.get('line_amount', 0) or 0)),
            )

            # Track and mark ornaments linked to this order
            new_ornament_ids.append(ornament.id)
            ornament.order = self.object
            try:
                ornament.ornament_type = Ornament.OrnamentCategory.ORDER
            except AttributeError:
                ornament.ornament_type = 'order'
            ornament.save()

        # Detach ornaments no longer present in this order
        if new_ornament_ids:
            # Revert ornament_type back to STOCK for removed ornaments
            Ornament.objects.filter(order=self.object).exclude(id__in=new_ornament_ids).update(
                order=None,
                ornament_type='stock',
            )

        # Save metal stock formset - NOW with the saved order instance
        metal_stock_formset = MetalStockFormSet(self.request.POST, instance=self.object)
        
        if metal_stock_formset.is_valid():
            # Only save forms that have data (non-empty)
            saved_metals = []
            for idx, form in enumerate(metal_stock_formset):
                # Check if form marked for deletion
                is_delete = form.cleaned_data.get('DELETE', False)
                if is_delete:
                    # Handle deleted forms - restore to stock
                    if form.instance.pk:
                        order_metal = form.instance
                        quantity = order_metal.quantity
                        metal_type = order_metal.metal_type
                        
                        try:
                            from goldsilverpurchase.models import MetalStockType
                            raw_stock_type = MetalStockType.objects.get(name__icontains='Raw')
                            
                            metal_stock = MetalStock.objects.get(
                                metal_type=metal_type,
                                purity=order_metal.purity,
                                stock_type=raw_stock_type
                            )
                            
                            # Restore quantity to metal stock
                            metal_stock.quantity += quantity
                            metal_stock.save()
                            
                            # Create a reversal movement record
                            MetalStockMovement.objects.create(
                                metal_stock=metal_stock,
                                movement_type='in',
                                quantity=quantity,
                                reference_type='Order',
                                reference_id=f"Order-{self.object.sn}",
                                notes=f"Metal removed from order {self.object.sn}"
                            )
                        except Exception as e:
                            messages.warning(self.request, f"Error restoring metal stock: {str(e)}")
                        
                        order_metal.delete()
                    continue
                
                # Check if form has actual data (metal_type is selected AND quantity is provided)
                metal_type = form.cleaned_data.get('metal_type')
                quantity = form.cleaned_data.get('quantity')
                
                # Skip empty rows - must have both metal_type and non-zero quantity
                if not metal_type or not quantity or quantity == 0:
                    # If this is an existing record being cleared, delete it
                    if form.instance.pk and not metal_type:
                        form.instance.delete()
                    continue
                
                # Check if this is a new record (no pk) before saving
                is_new = not form.instance.pk
                
                # Save the order metal entry
                order_metal = form.save()
                saved_metals.append(order_metal)
                
                # Deduct from existing metal stock if this is a new entry
                if is_new:
                    # This is a new metal addition to the order
                    quantity = order_metal.quantity
                    metal_type = order_metal.metal_type
                    
                    # Find matching metal stock (Raw type with same metal and purity)
                    try:
                        from goldsilverpurchase.models import MetalStockType
                        raw_stock_type = MetalStockType.objects.get(name='raw')
                        
                        metal_stock = MetalStock.objects.get(
                            metal_type=metal_type,
                            purity=order_metal.purity,
                            stock_type=raw_stock_type
                        )
                        
                        # Deduct quantity from metal stock
                        metal_stock.quantity -= quantity
                        if metal_stock.quantity < 0:
                            messages.warning(
                                self.request,
                                f"Warning: {metal_stock.get_metal_type_display()} stock ({metal_stock.purity}) is now negative: {metal_stock.quantity}g"
                            )
                        metal_stock.save()
                        
                        # Create a movement record
                        MetalStockMovement.objects.create(
                            metal_stock=metal_stock,
                            movement_type='out',
                            quantity=quantity,
                            reference_type='Order',
                            reference_id=f"Order-{self.object.sn}",
                            notes=f"Metal added to order {self.object.sn} for customer {self.object.customer_name}"
                        )
                    except MetalStock.DoesNotExist:
                        messages.warning(
                            self.request,
                            f"No matching metal stock found for {metal_type} ({order_metal.purity}). Please add to inventory first."
                        )
                    except Exception as e:
                        messages.warning(self.request, f"Error updating metal stock: {str(e)}")
            
            # Delete any remaining empty instances if they exist
            for form in metal_stock_formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()
            
            if saved_metals:
                messages.success(self.request, f"Successfully added {len(saved_metals)} metal records to order")
        else:
            # Show validation errors for debugging
            if metal_stock_formset.errors:
                messages.warning(self.request, f"Metal form errors: {metal_stock_formset.errors}")
            non_form_errors = metal_stock_formset.non_form_errors()
            if non_form_errors:
                messages.warning(self.request, f"Metal stock issues: {non_form_errors}")

        # Persist payment breakdown
        payment_lines_raw = form.cleaned_data.get('payment_lines_json') or '[]'
        try:
            payments_data = json.loads(payment_lines_raw)
        except (TypeError, ValueError):
            payments_data = []

        # Fallback: if cleaned_data is empty but POST has data, re-parse directly from POST
        if not payments_data:
            raw_post_val = self.request.POST.get('payment_lines_json')
            if raw_post_val:
                try:
                    payments_data = json.loads(raw_post_val)
                except Exception:
                    pass

        # Clear existing payments first
        self.object.payments.all().delete()
        
        # If no payments or all payments have 0 amount, auto-fill with order total
        if not payments_data:
            if self.object.total > 0:
                payments_data = [{'payment_mode': 'cash', 'amount': float(self.object.total)}]
        elif all(Decimal(str(p.get('amount', 0) or 0)) <= 0 for p in payments_data):
            raw_mode = payments_data[0].get('payment_mode') or payments_data[0].get('mode') or 'cash'
            mode = raw_mode if raw_mode in dict(Order.PAYMENT_CHOICES) else 'cash'
            if self.object.total > 0:
                payments_data = [{'payment_mode': mode, 'amount': float(self.object.total)}]
        # If payments include positive amounts, keep them as-is (no redistribution)
        
        for payment in payments_data:
            amount = Decimal(str(payment.get('amount', 0) or 0))
            mode = payment.get('payment_mode') or payment.get('mode') or 'cash'
            if amount <= 0:
                continue
            if mode not in dict(Order.PAYMENT_CHOICES):
                mode = 'cash'
            
            order_payment = OrderPayment.objects.create(
                order=self.object,
                payment_mode=mode,
                amount=amount,
            )

        # Recompute order totals from updated lines
        self.object.recompute_totals_from_lines()

        return redirect('order:list')


class OrderDeleteView(DeleteView):
    model = Order
    template_name = 'order/order_confirm_delete.html'
    success_url = reverse_lazy('order:list')


class CreateSaleFromOrderView(View):
    """(Deprecated here) moved to sales.views.CreateSaleFromOrderView."""

    def dispatch(self, request, *args, **kwargs):
        from sales.views import CreateSaleFromOrderView as _Create

        return _Create.as_view()(request, *args, **kwargs)


class SalesListView(ListView):
    """(Deprecated here) moved to sales.views.SalesListView."""

    def get(self, request, *args, **kwargs):
        from sales.views import SalesListView as _List

        return _List.as_view()(request, *args, **kwargs)


class SaleUpdateView(UpdateView):
    """(Deprecated here) moved to sales.views.SaleUpdateView."""

    def dispatch(self, request, *args, **kwargs):
        from sales.views import SaleUpdateView as _Update

        return _Update.as_view()(request, *args, **kwargs)


class SaleDeleteView(DeleteView):
    """(Deprecated here) moved to sales.views.SaleDeleteView."""

    def dispatch(self, request, *args, **kwargs):
        from sales.views import SaleDeleteView as _Delete

        return _Delete.as_view()(request, *args, **kwargs)


def order_print_view(request):
    """Printable view of orders (same queryset as OrderListView)."""

    view = OrderListView()
    view.request = request
    orders = view.get_queryset()

    return render(request, "order/print_view.html", {"orders": orders})


def order_export_excel(request):
    """Export orders, ornament lines, and payments to Excel (multiple sheets)."""

    view = OrderListView()
    view.request = request
    orders = view.get_queryset()

    wb = openpyxl.Workbook()
    ws_orders = wb.active
    ws_orders.title = "Orders"

    order_headers = [
        "Order No",
        "Order Date (BS)",
        "Deliver Date (BS)",
        "Customer",
        "Phone",
        "Status",
        "Amount",
        "Discount",
        "Subtotal",
        "Tax",
        "Total",
        "Payment Mode",
        "Paid Amount",
        "Remaining Amount",
    ]
    ws_orders.append(order_headers)

    for o in orders:
        ws_orders.append([
            o.sn,
            str(o.order_date) if o.order_date else "",
            str(o.deliver_date) if o.deliver_date else "",
            o.customer_name,
            o.phone_number,
            o.get_status_display(),
            o.amount,
            o.discount,
            o.subtotal,
            o.tax,
            o.total,
            o.get_payment_mode_display(),
            o.payment_amount,
            o.remaining_amount,
        ])

    # Sheet 2: Order Ornaments (line items)
    ws_lines = wb.create_sheet(title="OrderOrnaments")
    line_headers = [
        "Order No",
        "Customer",
        "Ornament Code",
        "Ornament Name",
        "Metal",
        "Weight",
        "Diamond Wt",
        "Zircon Wt",
        "Stone Wt",
        "Gold Rate",
        "Diamond Rate",
        "Zircon Rate",
        "Stone Rate",
        "Jarti",
        "Jyala",
        "Line Amount",
    ]
    ws_lines.append(line_headers)

    for o in orders.prefetch_related("order_ornaments__ornament"):
        for line in o.order_ornaments.all():
            orn = line.ornament
            ws_lines.append([
                o.sn,
                o.customer_name,
                orn.code if orn else "",
                orn.ornament_name if orn else "",
                orn.metal_type if orn else "",
                float(orn.weight or 0) if orn else 0,
                float(orn.diamond_weight or 0) if orn else 0,
                float(getattr(orn, "zircon_weight", 0) or 0) if orn else 0,
                float(orn.stone_weight or 0) if orn else 0,
                line.gold_rate,
                line.diamond_rate,
                line.zircon_rate,
                line.stone_rate,
                line.jarti,
                line.jyala,
                line.line_amount,
            ])

    # Auto column width helper
    def autosize(sheet):
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2

    autosize(ws_orders)
    autosize(ws_lines)

    # Sheet 3: Order Payments
    ws_pay = wb.create_sheet(title="OrderPayments")
    pay_headers = [
        "Order No",
        "Customer",
        "Payment Mode",
        "Amount",
        "Created",
    ]
    ws_pay.append(pay_headers)

    for o in orders.prefetch_related("payments"):
        for p in o.payments.all():
            ws_pay.append([
                o.sn,
                o.customer_name,
                p.get_payment_mode_display() if hasattr(p, "get_payment_mode_display") else p.payment_mode,
                p.amount,
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            ])

    autosize(ws_pay)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'

    return response


def order_ornaments_export_excel(request):
    """Export order line items (OrderOrnament) to Excel."""
    view = OrderListView()
    view.request = request
    orders = view.get_queryset()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OrderOrnaments"

    headers = [
        "Order No",
        "Customer",
        "Ornament Code",
        "Ornament Name",
        "Metal",
        "Weight",
        "Diamond Wt",
        "Zircon Wt",
        "Stone Wt",
        "Gold Rate",
        "Diamond Rate",
        "Zircon Rate",
        "Stone Rate",
        "Jarti",
        "Jyala",
        "Line Amount",
    ]
    ws.append(headers)

    for o in orders.prefetch_related("order_ornaments__ornament"):
        for line in o.order_ornaments.all():
            orn = line.ornament
            ws.append([
                o.sn,
                o.customer_name,
                orn.code if orn else "",
                orn.ornament_name if orn else "",
                orn.metal_type if orn else "",
                float(orn.weight or 0) if orn else 0,
                float(orn.diamond_weight or 0) if orn else 0,
                float(getattr(orn, "zircon_weight", 0) or 0) if orn else 0,
                float(orn.stone_weight or 0) if orn else 0,
                line.gold_rate,
                line.diamond_rate,
                line.zircon_rate,
                line.stone_rate,
                line.jarti,
                line.jyala,
                line.line_amount,
            ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="order_ornaments.xlsx"'
    return response


def order_payments_export_excel(request):
    """Export order payments to Excel."""
    view = OrderListView()
    view.request = request
    orders = view.get_queryset()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OrderPayments"

    headers = [
        "Order No",
        "Customer",
        "Payment Mode",
        "Amount",
        "Created",
    ]
    ws.append(headers)

    for o in orders.prefetch_related("payments"):
        for p in o.payments.all():
            ws.append([
                o.sn,
                o.customer_name,
                p.get_payment_mode_display() if hasattr(p, "get_payment_mode_display") else p.payment_mode,
                p.amount,
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="order_payments.xlsx"'
    return response


def order_ornaments_import_excel(request):
    """Import order ornament lines from an Excel file.

    Expected columns (matching export):
    Order No | Customer | Ornament Code | Ornament Name | Metal | Weight | Diamond Wt | Zircon Wt | Stone Wt | Gold Rate | Diamond Rate | Zircon Rate | Stone Rate | Jarti | Jyala | Line Amount
    """
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("order:list")

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            created = 0
            skipped = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    (
                        order_no,
                        _customer,
                        ornament_code,
                        _ornament_name,
                        _metal,
                        _weight,
                        _diamond_wt,
                        _zircon_wt,
                        _stone_wt,
                        gold_rate,
                        diamond_rate,
                        zircon_rate,
                        stone_rate,
                        jarti,
                        jyala,
                        line_amount,
                    ) = row
                except Exception:
                    skipped += 1
                    continue

                # Find order
                order_obj = None
                try:
                    order_obj = Order.objects.filter(sn=int(order_no)).first()
                except Exception:
                    order_obj = None
                if not order_obj:
                    skipped += 1
                    continue

                # Find ornament by code
                ornament_obj = None
                if ornament_code:
                    ornament_obj = Ornament.objects.filter(code=str(ornament_code)).first()
                if not ornament_obj:
                    skipped += 1
                    continue

                OrderOrnament.objects.create(
                    order=order_obj,
                    ornament=ornament_obj,
                    gold_rate=Decimal(str(gold_rate or 0)),
                    diamond_rate=Decimal(str(diamond_rate or 0)),
                    zircon_rate=Decimal(str(zircon_rate or 0)),
                    stone_rate=Decimal(str(stone_rate or 0)),
                    jarti=Decimal(str(jarti or 0)),
                    jyala=Decimal(str(jyala or 0)),
                    line_amount=Decimal(str(line_amount or 0)),
                )
                created += 1

            messages.success(request, f"Imported {created} order ornaments, skipped {skipped}.")
            return redirect("order:list")
        except Exception as exc:  # noqa: BLE001
            messages.error(request, f"Failed to import order ornaments: {exc}")
            return redirect("order:list")

    return render(request, "order/import_excel.html")


def order_payments_import_excel(request):
    """Import order payments from an Excel file.

    Expected columns (matching export):
    Order No | Customer | Payment Mode | Amount | Created
    """
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("order:list")

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            created = 0
            skipped = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    order_no, _customer, payment_mode, amount, _created = row
                except Exception:
                    skipped += 1
                    continue

                try:
                    order_obj = Order.objects.filter(sn=int(order_no)).first()
                except Exception:
                    order_obj = None

                if not order_obj:
                    skipped += 1
                    continue

                mode = payment_mode or "cash"
                if mode not in dict(Order.PAYMENT_CHOICES):
                    mode = "cash"

                amt = Decimal(str(amount or 0))
                if amt <= 0:
                    skipped += 1
                    continue

                OrderPayment.objects.create(
                    order=order_obj,
                    payment_mode=mode,
                    amount=amt,
                )
                created += 1

            messages.success(request, f"Imported {created} payments, skipped {skipped}.")
            return redirect("order:list")
        except Exception as exc:  # noqa: BLE001
            messages.error(request, f"Failed to import order payments: {exc}")
            return redirect("order:list")

    return render(request, "order/import_excel.html")


def order_import_excel(request):
    """Import orders from an Excel file with columns matching export."""

    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("order:import_excel")

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            imported = 0
            skipped = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                try:
                    (
                        order_no,
                        order_date_bs,
                        deliver_date_bs,
                        customer_name,
                        phone_number,
                        status,
                        amount,
                        discount,
                        subtotal,
                        tax,
                        total,
                        payment_mode,
                        paid_amount,
                        remaining_amount,
                    ) = row
                except Exception:
                    messages.error(request, "Excel format is incorrect. Columns mismatch.")
                    return redirect("order:import_excel")

                # Skip if an order with this number already exists
                if order_no and Order.objects.filter(sn=int(order_no)).exists():
                    skipped += 1
                    continue

                o = Order(
                    customer_name=customer_name or "Unknown",
                    phone_number=str(phone_number or ""),
                )

                if order_no:
                    o.sn = int(order_no)

                # Dates: NepaliDateField accepts YYYY-MM-DD strings
                if order_date_bs:
                    o.order_date = str(order_date_bs)
                if deliver_date_bs:
                    o.deliver_date = str(deliver_date_bs)

                if status in dict(Order.STATUS_CHOICES):
                    o.status = status

                o.amount = Decimal(str(amount or 0))
                o.discount = Decimal(str(discount or 0))
                o.subtotal = Decimal(str(subtotal or 0))
                o.tax = Decimal(str(tax or 0))
                o.total = Decimal(str(total or 0))
                o.payment_amount = Decimal(str(paid_amount or 0))
                o.remaining_amount = Decimal(str(remaining_amount or 0))

                if payment_mode in dict(Order.PAYMENT_CHOICES):
                    o.payment_mode = payment_mode

                o.save()
                imported += 1

            messages.success(
                request,
                f"Imported {imported} orders, skipped {skipped} duplicate/invalid rows.",
            )
            return redirect("order:list")

        except Exception as exc:  # noqa: BLE001
            messages.error(request, f"Failed to import Excel: {exc}")
            return redirect("order:import_excel")

    return render(request, "order/import_excel.html")

