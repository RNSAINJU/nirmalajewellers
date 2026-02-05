"""Order reports and analytics views"""
from django.shortcuts import render
from django.views import View
from django.db.models import Sum, Count, Q, F, DecimalField, Case, When, Value, Max, Min
from django.db.models.functions import Coalesce, TruncMonth, TruncYear
from django.http import HttpResponse
from decimal import Decimal
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Order, OrderPayment


class OrderDashboardReport(View):
    """Main order dashboard with key metrics and charts"""
    
    def get(self, request):
        # Date filters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        orders_qs = Order.objects.all().annotate(
            paid=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField())
        )
        
        if date_from:
            orders_qs = orders_qs.filter(order_date__gte=date_from)
        if date_to:
            orders_qs = orders_qs.filter(order_date__lte=date_to)
        
        # Key metrics
        total_orders = orders_qs.count()
        total_revenue = orders_qs.aggregate(
            total=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField())
        )['total']
        total_collected = orders_qs.aggregate(
            total=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField())
        )['total']
        total_pending = total_revenue - total_collected
        
        # Orders by status
        status_summary = orders_qs.values('status').annotate(
            count=Count('sn'),
            amount=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField())
        ).order_by('-count')
        
        # Orders by type
        type_summary = orders_qs.values('order_type').annotate(
            count=Count('sn'),
            amount=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField())
        ).order_by('-count')
        
        # Payment methods summary (from payments)
        payment_summary = OrderPayment.objects.values('payment_mode').annotate(
            count=Count('id'),
            amount=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
        ).order_by('-count')
        
        # Monthly trend
        monthly_data = orders_qs.annotate(
            month=TruncMonth('order_date')
        ).values('month').annotate(
            count=Count('sn'),
            revenue=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField()),
            collected=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField())
        ).order_by('month')
        
        # Top customers by revenue
        top_customers = orders_qs.values('customer_name').annotate(
            order_count=Count('sn'),
            total_amount=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField()),
            paid_amount=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField()),
            pending=F('total_amount') - F('paid_amount')
        ).order_by('-total_amount')[:10]
        
        context = {
            'total_orders': total_orders,
            'total_revenue': float(total_revenue),
            'total_collected': float(total_collected),
            'total_pending': float(total_pending),
            'collection_rate': round((float(total_collected) / float(total_revenue) * 100) if total_revenue > 0 else 0, 2),
            'status_summary': list(status_summary),
            'type_summary': list(type_summary),
            'payment_summary': list(payment_summary),
            'monthly_data': list(monthly_data),
            'top_customers': list(top_customers),
            'date_from': date_from,
            'date_to': date_to,
        }
        
        return render(request, 'order/reports/dashboard.html', context)


class OrderSalesAnalysis(View):
    """Detailed sales analysis with profit calculations"""
    
    def get(self, request):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        order_type = request.GET.get('order_type')
        
        orders_qs = Order.objects.prefetch_related('order_ornaments')
        
        if date_from:
            orders_qs = orders_qs.filter(order_date__gte=date_from)
        if date_to:
            orders_qs = orders_qs.filter(order_date__lte=date_to)
        if order_type:
            orders_qs = orders_qs.filter(order_type=order_type)
        
        # Sales analysis
        total_orders = orders_qs.count()
        total_amount = orders_qs.aggregate(
            total=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
        )['total']
        total_discount = orders_qs.aggregate(
            total=Coalesce(Sum('discount'), Decimal('0'), output_field=DecimalField())
        )['total']
        total_tax = orders_qs.aggregate(
            total=Coalesce(Sum('tax'), Decimal('0'), output_field=DecimalField())
        )['total']
        total_revenue = orders_qs.aggregate(
            total=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField())
        )['total']
        
        # Calculate profit (assuming cost is discount amount or markup)
        avg_order_value = float(total_amount) / total_orders if total_orders > 0 else 0
        avg_discount_percent = (float(total_discount) / float(total_amount) * 100) if total_amount > 0 else 0
        
        # Orders by status
        order_details = []
        for order in orders_qs.order_by('-order_date'):
            ornament_count = order.order_ornaments.count()
            metal_count = order.order_metals.count()
            pending = order.total - order.total_paid
            order_details.append({
                'sn': order.sn,
                'customer_name': order.customer_name,
                'order_date': order.order_date,
                'status': order.get_status_display(),
                'order_type': order.get_order_type_display(),
                'items_count': ornament_count + metal_count,
                'amount': float(order.amount),
                'discount': float(order.discount),
                'tax': float(order.tax),
                'total': float(order.total),
                'paid': float(order.total_paid),
                'pending': float(pending),
                'collection_percent': round((float(order.total_paid) / float(order.total) * 100) if order.total > 0 else 0, 2)
            })
        
        context = {
            'total_orders': total_orders,
            'total_amount': float(total_amount),
            'total_discount': float(total_discount),
            'total_tax': float(total_tax),
            'total_revenue': float(total_revenue),
            'avg_order_value': round(avg_order_value, 2),
            'avg_discount_percent': round(avg_discount_percent, 2),
            'order_details': order_details,
            'date_from': date_from,
            'date_to': date_to,
            'order_type': order_type,
        }
        
        return render(request, 'order/reports/sales_analysis.html', context)


class OrderPaymentAnalysis(View):
    """Payment collection and pending analysis"""
    
    def get(self, request):
        orders_qs = Order.objects.all().annotate(
            paid=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField())
        )
        
        # Payment status breakdown
        fully_paid = orders_qs.filter(remaining_amount=0).count()
        partial_paid = orders_qs.filter(remaining_amount__gt=0, paid__gt=0).count()
        unpaid = orders_qs.filter(paid=0).count()
        
        total_pending = orders_qs.aggregate(
            total=Coalesce(Sum('remaining_amount'), Decimal('0'), output_field=DecimalField())
        )['total']
        
        total_collected = orders_qs.aggregate(
            total=Coalesce(Sum('paid'), Decimal('0'), output_field=DecimalField())
        )['total']
        
        total_due = orders_qs.aggregate(
            total=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField())
        )['total']
        
        # Payment methods
        payment_methods = OrderPayment.objects.values('payment_mode').annotate(
            count=Count('id'),
            amount=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
        ).order_by('-amount')
        
        # Pending orders
        pending_orders = orders_qs.filter(
            remaining_amount__gt=0
        ).values('customer_name', 'phone_number').annotate(
            order_count=Count('sn'),
            total_pending=Coalesce(Sum('remaining_amount'), Decimal('0'), output_field=DecimalField()),
            total_due=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField()),
            paid=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField())
        ).order_by('-total_pending')
        
        # Calculate percentage paid for each customer
        for customer in pending_orders:
            if customer['total_due'] > 0:
                customer['paid_percent'] = round((float(customer['paid']) / float(customer['total_due'])) * 100, 2)
            else:
                customer['paid_percent'] = 0
        
        context = {
            'fully_paid': fully_paid,
            'partial_paid': partial_paid,
            'unpaid': unpaid,
            'total_pending': float(total_pending),
            'total_collected': float(total_collected),
            'total_due': float(total_due),
            'collection_rate': round((float(total_collected) / float(total_due) * 100) if total_due > 0 else 0, 2),
            'payment_methods': list(payment_methods),
            'pending_orders': list(pending_orders),
        }
        
        return render(request, 'order/reports/payment_analysis.html', context)


class OrderMetalAnalysis(View):
    """Metal usage and stock analysis in orders"""
    
    def get(self, request):
        from .models import OrderMetalStock
        
        # Get only orders that have metals
        orders_qs = Order.objects.filter(order_metals__isnull=False).prefetch_related('order_metals').distinct()
        
        total_orders = orders_qs.count()
        
        # Metal summary - aggregate directly from OrderMetalStock
        metal_summary = OrderMetalStock.objects.values('metal_type').annotate(
            total_quantity=Coalesce(Sum('quantity'), Decimal('0'), output_field=DecimalField()),
            total_value=Coalesce(Sum('line_amount'), Decimal('0'), output_field=DecimalField()),
            count=Count('order', distinct=True)
        ).order_by('-total_quantity')
        
        context = {
            'total_orders': total_orders,
            'metal_summary': list(metal_summary),
        }
        
        return render(request, 'order/reports/metal_analysis.html', context)


class OrderCustomerAnalysis(View):
    """Customer analysis and segmentation"""
    
    def get(self, request):
        orders_qs = Order.objects.all()
        
        # Customer metrics
        unique_customers = orders_qs.values('customer_name').distinct().count()
        repeat_customers = orders_qs.values('customer_name').annotate(
            order_count=Count('sn')
        ).filter(order_count__gt=1).count()
        
        # Calculate repeat customer percentage
        repeat_percent = round((repeat_customers / unique_customers * 100) if unique_customers > 0 else 0, 1)
        
        # Customer details
        customer_details = orders_qs.values('customer_name', 'phone_number').annotate(
            order_count=Count('sn'),
            total_spent=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField()),
            paid=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField()),
            pending=Coalesce(Sum('remaining_amount'), Decimal('0'), output_field=DecimalField()),
            avg_order_value=Coalesce(Sum('total') / Count('sn'), Decimal('0'), output_field=DecimalField())
        ).order_by('-total_spent')
        
        context = {
            'unique_customers': unique_customers,
            'repeat_customers': repeat_customers,
            'repeat_percent': repeat_percent,
            'customer_details': list(customer_details),
        }
        
        return render(request, 'order/reports/customer_analysis.html', context)


class FastSlowMoversReport(View):
    """Identify fast-moving and slow-moving ornaments by order frequency"""
    
    def get(self, request):
        from .models import OrderOrnament
        
        # Ornament movement analysis
        ornament_movement = OrderOrnament.objects.values(
            'ornament__code', 'ornament__ornament_name'
        ).annotate(
            order_count=Count('order', distinct=True),
            line_count=Count('id'),
            total_value=Coalesce(Sum('line_amount'), Decimal('0'), output_field=DecimalField()),
            last_order_date=Max('order__order_date')
        ).order_by('-order_count')
        
        # Split into fast & slow movers (threshold: median)
        movement_list = list(ornament_movement)
        total = len(movement_list)
        threshold_idx = total // 2 if total > 0 else 0
        
        fast_movers = movement_list[:threshold_idx]
        slow_movers = movement_list[threshold_idx:]
        
        context = {
            'fast_movers': fast_movers,
            'slow_movers': slow_movers,
            'total_ornaments': total,
        }
        
        return render(request, 'order/reports/fast_slow_movers.html', context)


class StockAgingReport(View):
    """Analyze ornament stock age and inventory aging"""
    
    def get(self, request):
        from ornament.models import Ornament
        from datetime import datetime, timedelta
        
        # Define aging buckets
        today = datetime.now().date()
        thirty_days_ago = today - timedelta(days=30)
        sixty_days_ago = today - timedelta(days=60)
        ninety_days_ago = today - timedelta(days=90)
        
        ornaments = Ornament.objects.all()
        
        # Categorize by age
        very_old = ornaments.filter(updated_at__lt=ninety_days_ago).count()
        old = ornaments.filter(updated_at__range=[sixty_days_ago, ninety_days_ago]).count()
        aged = ornaments.filter(updated_at__range=[thirty_days_ago, sixty_days_ago]).count()
        recent = ornaments.filter(updated_at__gte=thirty_days_ago).count()
        
        context = {
            'very_old': very_old,
            'old': old,
            'aged': aged,
            'recent': recent,
            'total': ornaments.count(),
        }
        
        return render(request, 'order/reports/stock_aging.html', context)


class MarginByCategoryReport(View):
    """Margin analysis by ornament category/type"""
    
    def get(self, request):
        from .models import OrderOrnament
        
        # Get ornaments with their category info
        margin_data = OrderOrnament.objects.values(
            'ornament__metal_type'
        ).annotate(
            total_sales=Coalesce(Sum('line_amount'), Decimal('0'), output_field=DecimalField()),
            order_count=Count('order', distinct=True),
            avg_order_value=Coalesce(Sum('line_amount') / Count('order', distinct=True), Decimal('0'), output_field=DecimalField())
        ).order_by('-total_sales')
        
        context = {
            'margin_data': list(margin_data),
        }
        
        return render(request, 'order/reports/margin_by_category.html', context)


class PaymentMixDiscountsReport(View):
    """Payment modes mix and discount analysis"""
    
    def get(self, request):
        # Payment mode distribution
        payment_modes = OrderPayment.objects.values('payment_mode').annotate(
            count=Count('id'),
            total_amount=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField()),
            avg_amount=Coalesce(Sum('amount') / Count('id'), Decimal('0'), output_field=DecimalField())
        ).order_by('-total_amount')
        
        # Discount analysis
        orders_with_discount = Order.objects.filter(discount__gt=0)
        total_discount = orders_with_discount.aggregate(
            total=Coalesce(Sum('discount'), Decimal('0'), output_field=DecimalField())
        )['total']
        
        avg_discount = orders_with_discount.aggregate(
            avg=Coalesce(Sum('discount') / Count('sn'), Decimal('0'), output_field=DecimalField())
        )['avg'] if orders_with_discount.count() > 0 else Decimal('0')
        
        # Calculate discount percentage
        total_sales = Order.objects.aggregate(
            total=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField())
        )['total']
        
        discount_percent = (float(total_discount) / float(total_sales) * 100) if total_sales > 0 else 0
        
        context = {
            'payment_modes': list(payment_modes),
            'orders_with_discount': orders_with_discount.count(),
            'total_discount': float(total_discount),
            'avg_discount': float(avg_discount),
            'discount_percent': round(discount_percent, 2),
        }
        
        return render(request, 'order/reports/payment_mix_discounts.html', context)


class DebtorAgingReport(View):
    """Analyze debtor accounts aging and collections"""
    
    def get(self, request):
        # Pending customers with aging info
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # Use today's date in a timezone-aware way
        today = timezone.now().date()
        thirty_days_ago = today - timedelta(days=30)
        sixty_days_ago = today - timedelta(days=60)
        ninety_days_ago = today - timedelta(days=90)
        
        # Get customers with outstanding balance
        debtors = Order.objects.filter(
            remaining_amount__gt=0
        ).values('customer_name', 'phone_number').annotate(
            total_due=Coalesce(Sum('remaining_amount'), Decimal('0'), output_field=DecimalField()),
            total_orders=Count('sn'),
            oldest_order_date=Min('order_date'),
            paid=Coalesce(Sum('payments__amount'), Decimal('0'), output_field=DecimalField()),
            total_amount=Coalesce(Sum('total'), Decimal('0'), output_field=DecimalField())
        ).order_by('-total_due')
        
        # Categorize by age
        current = []
        thirty = []
        sixty = []
        ninety_plus = []
        
        for debtor in debtors:
            oldest_date = debtor.get('oldest_order_date')
            if oldest_date:
                # Convert NepaliDate to Python date if needed
                try:
                    # If it's a NepaliDate object, convert to Python date
                    if hasattr(oldest_date, 'to_date_object'):
                        oldest_python_date = oldest_date.to_date_object()
                    else:
                        oldest_python_date = oldest_date
                    
                    days_old = (today - oldest_python_date).days
                except (AttributeError, TypeError):
                    # Skip if conversion fails
                    continue
                
                if days_old < 30:
                    current.append(debtor)
                elif days_old < 60:
                    thirty.append(debtor)
                elif days_old < 90:
                    sixty.append(debtor)
                else:
                    ninety_plus.append(debtor)
        
        context = {
            'current': current,
            'thirty': thirty,
            'sixty': sixty,
            'ninety_plus': ninety_plus,
            'total_debtors': len(list(debtors)),
            'total_pending': sum(float(d['total_due']) for d in debtors),
        }
        
        return render(request, 'order/reports/debtor_aging.html', context)


class MonthlySalesReport(View):
    """Monthly sales report showing aggregated sales data by month"""
    
    def get(self, request):
        from sales.models import Sale
        from ornament.models import Ornament
        import nepali_datetime as ndt
        
        # Get all sales
        sales = Sale.objects.select_related('order').prefetch_related(
            'order__order_ornaments__ornament',
            'sale_metals',
            'order__payments'
        ).exclude(sale_date__isnull=True).order_by('-sale_date', 'bill_no', 'order__sn')
        
        # Group sales by month/year
        monthly_data = {}
        
        purity_factors = {
            Ornament.TypeCategory.TWENTYFOURKARAT: Decimal("1.00"),
            Ornament.TypeCategory.TWENTHREEKARAT: Decimal("0.99"),
            Ornament.TypeCategory.TWENTYTWOKARAT: Decimal("0.98"),
            Ornament.TypeCategory.EIGHTEENKARAT: Decimal("0.75"),
            Ornament.TypeCategory.FOURTEENKARAT: Decimal("0.58"),
        }
        
        for sale in sales:
            if not sale.sale_date:
                continue
            
            # Get year and month from sale_date
            try:
                year = sale.sale_date.year
                month = sale.sale_date.month
            except AttributeError:
                continue
            
            month_key = f"{year}-{month:02d}"
            
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'year': year,
                    'month': month,
                    'sales_count': 0,
                    'ornament_gold_weight': Decimal("0"),
                    'ornament_silver_weight': Decimal("0"),
                    'raw_gold_weight': Decimal("0"),
                    'raw_silver_weight': Decimal("0"),
                    'total_jarti': Decimal("0"),
                    'total_sales_amount': Decimal("0"),
                    'total_remaining': Decimal("0"),
                    'total_tax': Decimal("0"),
                    'total_profit': Decimal("0"),
                }
            
            # Update counts
            monthly_data[month_key]['sales_count'] += 1
            
            # Calculate weights, jarti, and profit
            for line in sale.order.order_ornaments.all():
                weight = line.ornament.weight or Decimal("0")
                factor = purity_factors.get(getattr(line.ornament, 'type', None), Decimal("1.00"))
                
                if getattr(line.ornament, 'metal_type', None) == getattr(Ornament.MetalTypeCategory, 'GOLD', 'gold'):
                    gold_weight_24k = weight * factor
                    monthly_data[month_key]['ornament_gold_weight'] += gold_weight_24k
                elif getattr(line.ornament, 'metal_type', None) == getattr(Ornament.MetalTypeCategory, 'SILVER', 'silver'):
                    monthly_data[month_key]['ornament_silver_weight'] += weight
                
                # Jarti (customer jarti)
                jarti_amount = line.jarti or Decimal("0")
                monthly_data[month_key]['total_jarti'] += jarti_amount
                
                # Calculate profit
                customer_jarti = line.jarti or Decimal("0")
                ornament_jarti = line.ornament.jarti or Decimal("0")
                rate = line.gold_rate or Decimal("0")
                jyala = line.jyala or Decimal("0")
                
                jarti_difference = customer_jarti - ornament_jarti
                profit = (jarti_difference / Decimal("11.664") * rate) + jyala
                monthly_data[month_key]['total_profit'] += profit
            
            # Raw metals
            for metal in sale.sale_metals.all():
                if metal.metal_type == 'gold':
                    monthly_data[month_key]['raw_gold_weight'] += metal.quantity
                elif metal.metal_type == 'silver':
                    monthly_data[month_key]['raw_silver_weight'] += metal.quantity
                monthly_data[month_key]['total_sales_amount'] += metal.line_amount or Decimal("0")
            
            # Add order totals
            monthly_data[month_key]['total_sales_amount'] += sale.order.total or Decimal("0")
            monthly_data[month_key]['total_remaining'] += sale.order.remaining_amount or Decimal("0")
            monthly_data[month_key]['total_tax'] += sale.order.tax or Decimal("0")
            
        # Convert to sorted list
        nepali_months = {
            1: 'Baishakh', 2: 'Jestha', 3: 'Ashadh', 4: 'Shrawan',
            5: 'Bhadra', 6: 'Ashwin', 7: 'Kartik', 8: 'Mangsir',
            9: 'Poush', 10: 'Magh', 11: 'Falgun', 12: 'Chaitra'
        }
        
        monthly_list = []
        for key in sorted(monthly_data.keys(), reverse=True):
            data = monthly_data[key]
            data['month_name'] = nepali_months.get(data['month'], str(data['month']))
            data['month_key'] = key
            monthly_list.append(data)
        
        context = {
            'monthly_data': monthly_list,
        }
        
        return render(request, 'order/reports/monthly_sales.html', context)


class DailyProfitLossReport(View):
    """Daily Profit & Loss report with charts based on daily rates and stock"""
    
    def get(self, request):
        from main.models import DailyRate
        from ornament.models import Ornament
        from datetime import timedelta
        
        # Get number of days to show (default 30)
        days = int(request.GET.get('days', 30))
        
        # Get daily rates for the period
        rates = DailyRate.objects.all().order_by('-created_at')[:days]
        
        # Purity factors for gold
        purity_factors = {
            Ornament.TypeCategory.TWENTYFOURKARAT: Decimal("1.00"),
            Ornament.TypeCategory.TWENTHREEKARAT: Decimal("0.99"),
            Ornament.TypeCategory.TWENTYTWOKARAT: Decimal("0.98"),
            Ornament.TypeCategory.EIGHTEENKARAT: Decimal("0.75"),
            Ornament.TypeCategory.FOURTEENKARAT: Decimal("0.58"),
        }
        
        # Get current stock ornaments
        gold_ornaments = Ornament.objects.filter(
            metal_type=Ornament.MetalTypeCategory.GOLD,
            ornament_type=Ornament.OrnamentCategory.STOCK,
            status=Ornament.StatusCategory.ACTIVE
        )
        
        silver_ornaments = Ornament.objects.filter(
            metal_type=Ornament.MetalTypeCategory.SILVER,
            ornament_type=Ornament.OrnamentCategory.STOCK,
            status=Ornament.StatusCategory.ACTIVE
        )
        
        diamond_ornaments = Ornament.objects.filter(
            metal_type=Ornament.MetalTypeCategory.DIAMOND,
            ornament_type=Ornament.OrnamentCategory.STOCK,
            status=Ornament.StatusCategory.ACTIVE
        )
        
        # Calculate total weights
        gold_24k_weight = Decimal("0")
        for ornament in gold_ornaments:
            weight = ornament.weight or Decimal("0")
            factor = purity_factors.get(ornament.type, Decimal("1.00"))
            gold_24k_weight += weight * factor
        
        silver_weight = Decimal("0")
        for ornament in silver_ornaments:
            silver_weight += ornament.weight or Decimal("0")
        
        # Diamond: metal weight + diamond weight
        diamond_metal_weight = Decimal("0")
        diamond_stone_weight = Decimal("0")
        for ornament in diamond_ornaments:
            weight = ornament.weight or Decimal("0")
            factor = purity_factors.get(ornament.type, Decimal("1.00"))
            diamond_metal_weight += weight * factor
            diamond_stone_weight += ornament.diamond_weight or Decimal("0")
        
        # Prepare chart data
        chart_data = []
        TOLA_CONVERSION = Decimal("11.664")
        
        for rate in reversed(list(rates)):  # Reverse to show oldest to newest
            # Convert per tola to per gram for calculations
            gold_rate_per_gram = (rate.gold_rate or Decimal("0")) / TOLA_CONVERSION
            silver_rate_per_gram = (rate.silver_rate or Decimal("0")) / TOLA_CONVERSION
            
            # Gold value
            gold_value = gold_24k_weight * gold_rate_per_gram
            
            # Silver value
            silver_value = silver_weight * silver_rate_per_gram
            
            # Diamond value: metal value + stone value (assuming some diamond rate)
            # For simplicity, using gold rate for metal part
            diamond_metal_value = diamond_metal_weight * gold_rate_per_gram
            # Diamond stone value at estimated rate per gram (you can adjust this)
            diamond_rate_per_gram = Decimal("5000")  # Example rate
            diamond_stone_value = diamond_stone_weight * diamond_rate_per_gram
            diamond_value = diamond_metal_value + diamond_stone_value
            
            total_value = gold_value + silver_value + diamond_value
            
            chart_data.append({
                'date': rate.bs_date,
                'gold_rate': float(rate.gold_rate),
                'silver_rate': float(rate.silver_rate),
                'gold_value': float(gold_value),
                'silver_value': float(silver_value),
                'diamond_value': float(diamond_value),
                'total_value': float(total_value),
            })
        
        # Calculate P&L if we have at least 2 data points
        profit_loss_data = []
        if len(chart_data) >= 2:
            base_value = chart_data[0]['total_value']
            for data in chart_data:
                pl = data['total_value'] - base_value
                pl_percent = (pl / base_value * 100) if base_value > 0 else 0
                profit_loss_data.append({
                    'date': data['date'],
                    'value': pl,
                    'percent': pl_percent,
                })
        
        # Summary statistics
        if chart_data:
            latest = chart_data[-1]
            oldest = chart_data[0]
            total_change = latest['total_value'] - oldest['total_value']
            total_change_percent = (total_change / oldest['total_value'] * 100) if oldest['total_value'] > 0 else 0
        else:
            latest = oldest = None
            total_change = total_change_percent = 0
        
        context = {
            'chart_data': chart_data,
            'profit_loss_data': profit_loss_data,
            'gold_weight': gold_24k_weight,
            'silver_weight': silver_weight,
            'diamond_metal_weight': diamond_metal_weight,
            'diamond_stone_weight': diamond_stone_weight,
            'gold_count': gold_ornaments.count(),
            'silver_count': silver_ornaments.count(),
            'diamond_count': diamond_ornaments.count(),
            'latest': latest,
            'oldest': oldest,
            'total_change': total_change,
            'total_change_percent': total_change_percent,
            'days': days,
        }
        
        return render(request, 'order/reports/daily_profit_loss.html', context)


class OrnamentStockReport(View):
    """Stock ornament report by category with today's valuation"""
    
    def get(self, request):
        from ornament.models import Ornament, MainCategory, Stone, Potey
        from main.models import DailyRate
        
        # Get today's gold and silver rates
        today_rate = DailyRate.objects.first()
        gold_rate = today_rate.gold_rate if today_rate else Decimal("0")
        silver_rate = today_rate.silver_rate if today_rate else Decimal("0")
        
        # Convert per tola to per gram
        TOLA_CONVERSION = Decimal("11.664")
        gold_rate_per_gram = gold_rate / TOLA_CONVERSION if gold_rate else Decimal("0")
        silver_rate_per_gram = silver_rate / TOLA_CONVERSION if silver_rate else Decimal("0")
        
        # Get all active stock ornaments
        ornaments = Ornament.objects.filter(
            ornament_type=Ornament.OrnamentCategory.STOCK,
            status=Ornament.StatusCategory.ACTIVE
        ).select_related('maincategory', 'subcategory')
        
        # Purity conversion factors to 24K equivalent
        purity_factors = {
            Ornament.TypeCategory.TWENTYFOURKARAT: Decimal("1.00"),
            Ornament.TypeCategory.TWENTHREEKARAT: Decimal("0.99"),
            Ornament.TypeCategory.TWENTYTWOKARAT: Decimal("0.98"),
            Ornament.TypeCategory.EIGHTEENKARAT: Decimal("0.75"),
            Ornament.TypeCategory.FOURTEENKARAT: Decimal("0.58"),
        }
        
        # Group ornaments by metal type and main category
        gold_category_data = {}
        silver_category_data = {}
        
        # Summary totals
        total_gold_weight = Decimal("0")
        total_gold_amount = Decimal("0")
        total_diamond_weight = Decimal("0")
        total_silver_weight = Decimal("0")
        total_silver_amount = Decimal("0")
        
        for ornament in ornaments:
            category_name = ornament.maincategory.name if ornament.maincategory else "Uncategorized"
        for ornament in ornaments:
            category_name = ornament.maincategory.name if ornament.maincategory else "Uncategorized"
            
            # Calculate weight and amount based on metal type
            if ornament.metal_type == Ornament.MetalTypeCategory.GOLD:
                if category_name not in gold_category_data:
                    gold_category_data[category_name] = {
                        'count': 0,
                        'total_weight': Decimal("0"),
                        'total_diamond': Decimal("0"),
                        'total_amount': Decimal("0"),
                        'metal_type': 'Gold',
                    }
                
                gold_category_data[category_name]['count'] += 1
                
                weight = ornament.weight or Decimal("0")
                factor = purity_factors.get(ornament.type, Decimal("1.00"))
                weight_24k = weight * factor
                
                gold_category_data[category_name]['total_weight'] += weight_24k
                
                # Calculate amount: (weight * gold_rate) + jyala + diamond value
                gold_value = weight_24k * gold_rate_per_gram
                jyala = ornament.jyala or Decimal("0")
                diamond_weight = ornament.diamond_weight or Decimal("0")
                diamond_value = diamond_weight * (ornament.diamond_rate or Decimal("0"))
                
                amount = gold_value + jyala + diamond_value
                gold_category_data[category_name]['total_amount'] += amount
                gold_category_data[category_name]['total_diamond'] += diamond_weight
                
                # Add to summary
                total_gold_weight += weight_24k
                total_gold_amount += amount
                total_diamond_weight += diamond_weight
                
            elif ornament.metal_type == Ornament.MetalTypeCategory.SILVER:
                if category_name not in silver_category_data:
                    silver_category_data[category_name] = {
                        'count': 0,
                        'total_weight': Decimal("0"),
                        'total_diamond': Decimal("0"),
                        'total_amount': Decimal("0"),
                        'metal_type': 'Silver',
                    }
                
                silver_category_data[category_name]['count'] += 1
                
                weight = ornament.weight or Decimal("0")
                silver_category_data[category_name]['total_weight'] += weight
                
                # Calculate amount
                silver_value = weight * silver_rate_per_gram
                jyala = ornament.jyala or Decimal("0")
                amount = silver_value + jyala
                
                silver_category_data[category_name]['total_amount'] += amount
                
                # Add to summary
                total_silver_weight += weight
                total_silver_amount += amount
                
            elif ornament.metal_type == Ornament.MetalTypeCategory.DIAMOND:
                if category_name not in gold_category_data:
                    gold_category_data[category_name] = {
                        'count': 0,
                        'total_weight': Decimal("0"),
                        'total_diamond': Decimal("0"),
                        'total_amount': Decimal("0"),
                        'metal_type': 'Diamond',
                    }
                
                gold_category_data[category_name]['count'] += 1
                
                weight = ornament.weight or Decimal("0")
                factor = purity_factors.get(ornament.type, Decimal("1.00"))
                weight_24k = weight * factor
                
                diamond_weight = ornament.diamond_weight or Decimal("0")
                gold_category_data[category_name]['total_weight'] += weight_24k
                gold_category_data[category_name]['total_diamond'] += diamond_weight
                
                # Calculate amount
                gold_value = weight_24k * gold_rate_per_gram
                jyala = ornament.jyala or Decimal("0")
                diamond_value = diamond_weight * (ornament.diamond_rate or Decimal("0"))
                
                amount = gold_value + jyala + diamond_value
                gold_category_data[category_name]['total_amount'] += amount
                
                # Add to summary (Diamond is counted in gold section)
                total_gold_weight += weight_24k
                total_diamond_weight += diamond_weight
                total_gold_amount += amount
        
        # Get Potey and Stones stock
        potey_stock = Potey.objects.all()
        total_potey_amount = sum(p.sales_price or Decimal("0") for p in potey_stock)
        
        stones_stock = Stone.objects.all()
        total_stones_amount = sum(s.sales_price or Decimal("0") for s in stones_stock)
        
        # Calculate grand total
        grand_total = total_gold_amount + total_silver_amount + total_potey_amount + total_stones_amount
        
        # Create separate lists for gold and silver ornaments
        gold_category_list = [
            {
                'name': name,
                'count': data['count'],
                'weight': data['total_weight'],
                'diamond': data['total_diamond'],
                'rate': gold_rate_per_gram,
                'amount': data['total_amount'],
                'metal_type': data['metal_type'],
            }
            for name, data in sorted(gold_category_data.items())
        ]
        
        silver_category_list = [
            {
                'name': name,
                'count': data['count'],
                'weight': data['total_weight'],
                'diamond': data['total_diamond'],
                'rate': silver_rate_per_gram,
                'amount': data['total_amount'],
                'metal_type': data['metal_type'],
            }
            for name, data in sorted(silver_category_data.items())
        ]
        
        context = {
            'gold_category_list': gold_category_list,
            'silver_category_list': silver_category_list,
            'total_gold_weight': total_gold_weight,
            'total_gold_amount': total_gold_amount,
            'total_diamond_weight': total_diamond_weight,
            'total_silver_weight': total_silver_weight,
            'total_silver_amount': total_silver_amount,
            'total_potey_amount': total_potey_amount,
            'total_stones_amount': total_stones_amount,
            'grand_total': grand_total,
            'gold_rate': gold_rate,
            'silver_rate': silver_rate,
            'gold_rate_per_gram': gold_rate_per_gram,
            'silver_rate_per_gram': silver_rate_per_gram,
            'today_date': today_rate.bs_date if today_rate else 'N/A',
        }
        
        return render(request, 'order/reports/ornament_stock.html', context)


def ornament_stock_export_excel(request):
    """Export ornament stock report to Excel."""
    from ornament.models import Ornament, MainCategory, Stone, Potey
    from main.models import DailyRate
    
    # Get today's gold and silver rates
    today_rate = DailyRate.objects.first()
    gold_rate = today_rate.gold_rate if today_rate else Decimal("0")
    silver_rate = today_rate.silver_rate if today_rate else Decimal("0")
    
    # Convert per tola to per gram
    TOLA_CONVERSION = Decimal("11.664")
    gold_rate_per_gram = gold_rate / TOLA_CONVERSION if gold_rate else Decimal("0")
    silver_rate_per_gram = silver_rate / TOLA_CONVERSION if silver_rate else Decimal("0")
    
    # Get all active stock ornaments
    ornaments = Ornament.objects.filter(
        ornament_type=Ornament.OrnamentCategory.STOCK,
        status=Ornament.StatusCategory.ACTIVE
    ).select_related('maincategory', 'subcategory')
    
    # Purity conversion factors to 24K equivalent
    purity_factors = {
        Ornament.TypeCategory.TWENTYFOURKARAT: Decimal("1.00"),
        Ornament.TypeCategory.TWENTHREEKARAT: Decimal("0.99"),
        Ornament.TypeCategory.TWENTYTWOKARAT: Decimal("0.98"),
        Ornament.TypeCategory.EIGHTEENKARAT: Decimal("0.75"),
        Ornament.TypeCategory.FOURTEENKARAT: Decimal("0.58"),
    }
    
    # Group ornaments by metal type and main category
    gold_category_data = {}
    silver_category_data = {}
    
    # Summary totals
    total_gold_weight = Decimal("0")
    total_gold_amount = Decimal("0")
    total_diamond_weight = Decimal("0")
    total_silver_weight = Decimal("0")
    total_silver_amount = Decimal("0")
    
    for ornament in ornaments:
        category_name = ornament.maincategory.name if ornament.maincategory else "Uncategorized"
        
        # Calculate weight and amount based on metal type
        if ornament.metal_type == Ornament.MetalTypeCategory.GOLD:
            if category_name not in gold_category_data:
                gold_category_data[category_name] = {
                    'count': 0,
                    'total_weight': Decimal("0"),
                    'total_diamond': Decimal("0"),
                    'total_amount': Decimal("0"),
                    'metal_type': 'Gold',
                }
            
            gold_category_data[category_name]['count'] += 1
            
            weight = ornament.weight or Decimal("0")
            factor = purity_factors.get(ornament.type, Decimal("1.00"))
            weight_24k = weight * factor
            
            gold_category_data[category_name]['total_weight'] += weight_24k
            
            # Calculate amount: (weight * gold_rate) + jyala + diamond value
            gold_value = weight_24k * gold_rate_per_gram
            jyala = ornament.jyala or Decimal("0")
            diamond_weight = ornament.diamond_weight or Decimal("0")
            diamond_value = diamond_weight * (ornament.diamond_rate or Decimal("0"))
            
            amount = gold_value + jyala + diamond_value
            gold_category_data[category_name]['total_amount'] += amount
            gold_category_data[category_name]['total_diamond'] += diamond_weight
            
            # Add to summary
            total_gold_weight += weight_24k
            total_gold_amount += amount
            total_diamond_weight += diamond_weight
            
        elif ornament.metal_type == Ornament.MetalTypeCategory.SILVER:
            if category_name not in silver_category_data:
                silver_category_data[category_name] = {
                    'count': 0,
                    'total_weight': Decimal("0"),
                    'total_diamond': Decimal("0"),
                    'total_amount': Decimal("0"),
                    'metal_type': 'Silver',
                }
            
            silver_category_data[category_name]['count'] += 1
            
            weight = ornament.weight or Decimal("0")
            silver_category_data[category_name]['total_weight'] += weight
            
            # Calculate amount
            silver_value = weight * silver_rate_per_gram
            jyala = ornament.jyala or Decimal("0")
            amount = silver_value + jyala
            
            silver_category_data[category_name]['total_amount'] += amount
            
            # Add to summary
            total_silver_weight += weight
            total_silver_amount += amount
            
        elif ornament.metal_type == Ornament.MetalTypeCategory.DIAMOND:
            if category_name not in gold_category_data:
                gold_category_data[category_name] = {
                    'count': 0,
                    'total_weight': Decimal("0"),
                    'total_diamond': Decimal("0"),
                    'total_amount': Decimal("0"),
                    'metal_type': 'Diamond',
                }
            
            gold_category_data[category_name]['count'] += 1
            
            weight = ornament.weight or Decimal("0")
            factor = purity_factors.get(ornament.type, Decimal("1.00"))
            weight_24k = weight * factor
            
            diamond_weight = ornament.diamond_weight or Decimal("0")
            gold_category_data[category_name]['total_weight'] += weight_24k
            gold_category_data[category_name]['total_diamond'] += diamond_weight
            
            # Calculate amount
            gold_value = weight_24k * gold_rate_per_gram
            jyala = ornament.jyala or Decimal("0")
            diamond_value = diamond_weight * (ornament.diamond_rate or Decimal("0"))
            
            amount = gold_value + jyala + diamond_value
            gold_category_data[category_name]['total_amount'] += amount
            
            # Add to summary (Diamond is counted in gold section)
            total_gold_weight += weight_24k
            total_diamond_weight += diamond_weight
            total_gold_amount += amount
    
    # Get Potey and Stones stock
    potey_stock = Potey.objects.all()
    total_potey_amount = sum(p.sales_price or Decimal("0") for p in potey_stock)
    
    stones_stock = Stone.objects.all()
    total_stones_amount = sum(s.sales_price or Decimal("0") for s in stones_stock)
    
    # Calculate grand total
    grand_total = total_gold_amount + total_silver_amount + total_potey_amount + total_stones_amount
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    summary_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary.merge_cells('A1:C1')
    ws_summary['A1'] = "Gold / Silver Stock Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # Rate info
    ws_summary['A3'] = f"Date: {today_rate.bs_date if today_rate else 'N/A'}"
    ws_summary['A4'] = f"Gold Rate: रु{float(gold_rate):.2f} per tola (रु{float(gold_rate_per_gram):.2f} per gram)"
    ws_summary['A5'] = f"Silver Rate: रु{float(silver_rate):.2f} per tola (रु{float(silver_rate_per_gram):.2f} per gram)"
    
    # Summary table headers
    ws_summary['A7'] = "Type"
    ws_summary['B7'] = "Weight (gms)"
    ws_summary['C7'] = "Amount (रु)"
    
    for cell in ['A7', 'B7', 'C7']:
        ws_summary[cell].fill = header_fill
        ws_summary[cell].font = header_font
        ws_summary[cell].alignment = Alignment(horizontal='center')
        ws_summary[cell].border = border
    
    # Summary data
    summary_data = [
        ["Gold", float(total_gold_weight), float(total_gold_amount)],
        ["Diamond", float(total_diamond_weight), ""],
        ["Silver", float(total_silver_weight), float(total_silver_amount)],
        ["Potey", "", float(total_potey_amount)],
        ["Stones", "", float(total_stones_amount)],
        ["Total", "", float(grand_total)],
    ]
    
    row = 8
    for data in summary_data:
        ws_summary[f'A{row}'] = data[0]
        ws_summary[f'B{row}'] = data[1]
        ws_summary[f'C{row}'] = data[2]
        
        for col in ['A', 'B', 'C']:
            ws_summary[f'{col}{row}'].border = border
            if row == 13:  # Total row
                ws_summary[f'{col}{row}'].fill = summary_fill
                ws_summary[f'{col}{row}'].font = summary_font
        
        row += 1
    
    # Gold Ornaments Sheet
    ws_gold = wb.create_sheet(title="Gold Ornaments")
    
    # Title
    ws_gold.merge_cells('A1:G1')
    ws_gold['A1'] = "Gold Ornament Stock Details"
    ws_gold['A1'].font = Font(bold=True, size=14)
    ws_gold['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    gold_headers = ["SN", "Particulars", "Quantity", "Weight (gms)", "Diamond (gms)", "Rate (per gm)", "Total (रु)"]
    for col_num, header in enumerate(gold_headers, 1):
        cell = ws_gold.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Gold data
    row = 4
    for idx, (name, data) in enumerate(sorted(gold_category_data.items()), 1):
        ws_gold.cell(row=row, column=1, value=idx)
        ws_gold.cell(row=row, column=2, value=name)
        ws_gold.cell(row=row, column=3, value=data['count'])
        ws_gold.cell(row=row, column=4, value=float(data['total_weight']))
        ws_gold.cell(row=row, column=5, value=float(data['total_diamond']))
        ws_gold.cell(row=row, column=6, value=float(gold_rate_per_gram))
        ws_gold.cell(row=row, column=7, value=float(data['total_amount']))
        
        for col in range(1, 8):
            ws_gold.cell(row=row, column=col).border = border
        
        row += 1
    
    # Adjust column widths
    for col in range(1, 8):
        ws_gold.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Silver Ornaments Sheet
    ws_silver = wb.create_sheet(title="Silver Ornaments")
    
    # Title
    ws_silver.merge_cells('A1:G1')
    ws_silver['A1'] = "Silver Ornament Stock Details"
    ws_silver['A1'].font = Font(bold=True, size=14)
    ws_silver['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    silver_headers = ["SN", "Particulars", "Quantity", "Weight (gms)", "Diamond (gms)", "Rate (per gm)", "Total (रु)"]
    for col_num, header in enumerate(silver_headers, 1):
        cell = ws_silver.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Silver data
    row = 4
    for idx, (name, data) in enumerate(sorted(silver_category_data.items()), 1):
        ws_silver.cell(row=row, column=1, value=idx)
        ws_silver.cell(row=row, column=2, value=name)
        ws_silver.cell(row=row, column=3, value=data['count'])
        ws_silver.cell(row=row, column=4, value=float(data['total_weight']))
        ws_silver.cell(row=row, column=5, value=float(data['total_diamond']))
        ws_silver.cell(row=row, column=6, value=float(silver_rate_per_gram))
        ws_silver.cell(row=row, column=7, value=float(data['total_amount']))
        
        for col in range(1, 8):
            ws_silver.cell(row=row, column=col).border = border
        
        row += 1
    
    # Adjust column widths
    for col in range(1, 8):
        ws_silver.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Adjust column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 20
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ornament_stock_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response
