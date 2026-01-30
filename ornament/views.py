from django.db import models
from django.urls import reverse_lazy
from .models import Ornament, Stone, Motimala, Potey
# Import ListView and CreateView for generic class-based views
from django.views.generic import ListView, CreateView
from .forms import OrnamentForm, KaligarCashAccountForm, KaligarGoldAccountForm
from django.shortcuts import get_object_or_404
# Create Kaligar_CashAccount for a Kaligar
def create_kaligar_cash_account(request, kaligar_id=None):
    kaligar = get_object_or_404(Kaligar, id=kaligar_id) if kaligar_id else None
    if request.method == 'POST':
        form = KaligarCashAccountForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ornament:kaligar_list')
    else:
        form = KaligarCashAccountForm(initial={'kaligar': kaligar})
    return render(request, 'ornament/kaligar_cashaccount_form.html', {'form': form, 'kaligar': kaligar})

# Create Kaligar_GoldAccount for a Kaligar
def create_kaligar_gold_account(request, kaligar_id=None):
    kaligar = get_object_or_404(Kaligar, id=kaligar_id) if kaligar_id else None
    if request.method == 'POST':
        form = KaligarGoldAccountForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ornament:kaligar_list')
    else:
        form = KaligarGoldAccountForm(initial={'kaligar': kaligar})
    return render(request, 'ornament/kaligar_goldaccount_form.html', {'form': form, 'kaligar': kaligar})
# Stones List and Create Views
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required

@method_decorator(login_required, name='dispatch')
class StoneListView(ListView):
    model = Stone
    template_name = 'ornament/stone_list.html'
    context_object_name = 'stones'
    ordering = ['-id']
    paginate_by = 10

@method_decorator(login_required, name='dispatch')
class StoneCreateView(CreateView):
    model = Stone
    fields = ['name', 'cost_per_carat', 'carat', 'sales_per_carat']
    template_name = 'ornament/stone_form.html'
    success_url = reverse_lazy('ornament:stone_list')

# Motimala List and Create Views
@method_decorator(login_required, name='dispatch')
class MotimalaListView(ListView):
    model = Motimala
    template_name = 'ornament/motimala_list.html'
    context_object_name = 'motimalas'
    ordering = ['-id']
    paginate_by = 10

@method_decorator(login_required, name='dispatch')
class MotimalaCreateView(CreateView):
    model = Motimala
    fields = ['name', 'cost_per_mala', 'quantity', 'sales_per_mala']
    template_name = 'ornament/motimala_form.html'
    success_url = reverse_lazy('ornament:motimala_list')

# Potey List and Create Views
@method_decorator(login_required, name='dispatch')
class PoteyListView(ListView):
    model = Potey
    template_name = 'ornament/potey_list.html'
    context_object_name = 'poteys'
    ordering = ['-id']
    paginate_by = 10

@method_decorator(login_required, name='dispatch')
class PoteyCreateView(CreateView):
    model = Potey
    fields = ['name', 'loon', 'cost_per_loon', 'sales_per_loon']
    template_name = 'ornament/potey_form.html'
    success_url = reverse_lazy('ornament:potey_list')
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.db.models import Sum, F, DecimalField
from django.db.models.functions import Coalesce
from .models import Kaligar, Ornament, MainCategory, SubCategory
from order.models import Order, OrderOrnament
from .forms import OrnamentForm
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db.models import Q
from io import BytesIO
from django.contrib import messages
from decimal import Decimal
from django.forms import modelformset_factory
import openpyxl
from openpyxl.utils import get_column_letter
import nepali_datetime as ndt
from main.models import Stock
from django.db import IntegrityError

class MainCategoryCreateView(CreateView):
    model = MainCategory
    fields = ['name']
    template_name = 'ornament/maincategory_form.html'
    success_url = reverse_lazy('ornament:list')

class SubCategoryCreateView(CreateView):
    model = SubCategory
    fields = ['name']
    template_name = 'ornament/subcategory_form.html'
    success_url = reverse_lazy('ornament:list')

class KaligarCreateView(CreateView):
    model = Kaligar
    fields = ['name','phone_no','panno','address','stamp']
    template_name = 'ornament/kaligar_form.html'
    success_url = reverse_lazy('ornament:list')


class OrnamentListView(ListView):
    model = Ornament
    template_name = 'ornament/ornament_list.html'
    context_object_name = 'ornaments'
    # Order by latest entry first
    ordering = ['-id']

    def get_paginate_by(self, queryset):
        # If both metal_type and maincategory are set, show all (no pagination)
        metal_type = self.request.GET.get('metal_type')
        maincategory = self.request.GET.get('maincategory')
        if metal_type and maincategory:
            return None
        return 10  # Default pagination

    def get_queryset(self):
        qs = super().get_queryset()
        
        # Optimize with prefetch_related for foreign keys
        qs = qs.select_related('maincategory', 'subcategory', 'kaligar')

        # Filters
        code = self.request.GET.get("code")
        name = self.request.GET.get("name")
        customer = self.request.GET.get("customer")
        type = self.request.GET.get("type")
        ornament_type = self.request.GET.get("ornament_type")
        metal_type = self.request.GET.get("metal_type")
        maincategory_id = self.request.GET.get("maincategory")
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")
        kaligar_id = self.request.GET.get('kaligar')
        search=self.request.GET.get('search')

        if search:
            qs = qs.filter(
                Q(weight__icontains=search) |
                Q(diamond_weight__icontains=search) |
                Q(ornament_name__icontains=search) |
                Q(gross_weight__icontains=search)
            )

        if code:
            qs = qs.filter(code__icontains=code)

        if name:
            qs = qs.filter(ornament_name__icontains=name)

        if type:
            qs = qs.filter(type=type)

        if ornament_type:
            qs = qs.filter(ornament_type=ornament_type)

        if metal_type:
            qs = qs.filter(metal_type=metal_type)

        if kaligar_id:
            qs = qs.filter(kaligar_id=kaligar_id)

        if maincategory_id:
            qs = qs.filter(maincategory_id=maincategory_id)

        if start_date and end_date:
            qs = qs.filter(
                ornament_date__gte=start_date,
                ornament_date__lte=end_date
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Total weight calculation for current page only
        page_ornaments = context['ornaments']
        total_weight = Decimal('0')
        for ornament in page_ornaments:
            total_weight += ornament.weight or Decimal('0')
        context['total_weight'] = total_weight

        # Filters back to template
        context['metal_type'] = self.request.GET.get('metal_type')
        context['ornament_type'] = self.request.GET.get('ornament_type')
        context['type'] = self.request.GET.get('type')
        context['maincategory'] = self.request.GET.get('maincategory')

        # Kaligar list
        context['kaligar'] = Kaligar.objects.all()
        context['selected_kaligar'] = self.request.GET.get('kaligar', '')

        # Status choices for display
        status_choices = {choice[0]: choice[1] for choice in Ornament.StatusCategory.choices}
        context['status_choices'] = status_choices
        
        # Group PAGINATED ornaments by status using Python filtering (not QuerySet)
        ornaments_by_status = {}
        for status_key in status_choices:
            # Use list comprehension since page_ornaments is a list, not a QuerySet
            ornaments_by_status[status_key] = [o for o in page_ornaments if o.status == status_key]
        
        context['ornaments_by_status'] = ornaments_by_status

        return context


class OrnamentCreateView(CreateView):
    model = Ornament
    form_class = OrnamentForm
    template_name = 'ornament/ornament_form.html'
    success_url = reverse_lazy('ornament:list')

    def get_initial(self):
        """Set initial values including today's Nepali date."""
        initial = super().get_initial()
        initial['ornament_date'] = ndt.date.today()
        # Set default values for weight fields
        initial['gross_weight'] = Decimal('0.0')
        initial['weight'] = Decimal('0.0')
        initial['diamond_weight'] = Decimal('0.0')
        initial['diamond_rate'] = Decimal('0.0')
        initial['zircon_weight'] = Decimal('0.0')
        initial['stone_weight'] = Decimal('0.0')
        initial['stone_percaratprice'] = Decimal('0.0')
        initial['stone_totalprice'] = Decimal('0.0')
        initial['jarti'] = Decimal('0.0')
        initial['jyala'] = Decimal('0.0')
        return initial

    def form_valid(self, form):
        try:
            # Cloudinary image is automatically handled by ModelForm
            result = super().form_valid(form)
            
            # Check if Cloudinary upload failed but form saved anyway
            if hasattr(form, '_cloudinary_failed') and form._cloudinary_failed:
                messages.warning(
                    self.request,
                    'Ornament created successfully, but image upload failed due to network restrictions. '
                    'Images may not be uploadable on this server (PythonAnywhere firewall).'
                )
            
            return result
        except Exception as e:
            # Handle unexpected errors
            error_msg = str(e)
            messages.error(self.request, f'Error creating ornament: {error_msg}')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['JARTI_CHOICES'] = [
            (4.0, "4%"),
            (4.5, "4.5%"),
            (5.0, "5%"),
            (6.5, "6.5%"),
            (8.0, "8%"),
        ]
        return context


class OrnamentUpdateView(UpdateView):
    model = Ornament
    form_class = OrnamentForm
    template_name = 'ornament/ornament_form.html'
    success_url = reverse_lazy('ornament:list')

    def form_valid(self, form):
        try:
            # Cloudinary image is automatically handled by ModelForm
            result = super().form_valid(form)
            
            # Check if Cloudinary upload failed but form saved anyway
            if hasattr(form, '_cloudinary_failed') and form._cloudinary_failed:
                messages.warning(
                    self.request,
                    'Ornament updated successfully, but image upload failed due to network restrictions. '
                    'Images may not be uploadable on this server (PythonAnywhere firewall).'
                )
            
            return result
        except Exception as e:
            # Handle unexpected errors
            error_msg = str(e)
            messages.error(self.request, f'Error updating ornament: {error_msg}')
            return self.form_invalid(form)

    def get_success_url(self):
        # Check if there's a 'next' parameter to redirect back to
        next_url = self.request.GET.get('next')
        if next_url:
            return next_url
        return self.success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['JARTI_CHOICES'] = [
            (4.0, "4%"),
            (4.5, "4.5%"),
            (5.0, "5%"),
            (6.5, "6.5%"),
            (8.0, "8%"),
        ]
        return context


class OrnamentDeleteView(DeleteView):
    model = Ornament
    template_name = 'ornament/ornament_confirm_delete.html'
    success_url = reverse_lazy('ornament:list')

    def post(self, request, *args, **kwargs):
        """Handle delete with protection check for orders."""
        self.object = self.get_object()
        
        # Check if this ornament is used in any orders
        related_orders = OrderOrnament.objects.filter(ornament=self.object).select_related('order')
        
        if related_orders.exists():
            # Ornament is used in orders - don't allow delete
            order_list = [f"Order {o.order.sn} - {o.order.customer_name}" for o in related_orders]
            messages.error(
                request,
                f"Cannot delete this ornament as it is referenced by {len(order_list)} order(s): {', '.join(order_list[:3])}{'...' if len(order_list) > 3 else ''}. "
                f"Please use the 'Destroy' status instead to mark it as no longer available."
            )
            return redirect('ornament:list')
        
        # No related orders, proceed with deletion
        messages.success(request, f"Ornament '{self.object.ornament_name}' has been deleted.")
        return super().post(request, *args, **kwargs)


class OrnamentDestroyView(UpdateView):
    """Change ornament status to 'destroyed'."""
    model = Ornament
    fields = []
    template_name = 'ornament/ornament_confirm_destroy.html'
    success_url = reverse_lazy('ornament:list')

    def form_valid(self, form):
        """Update status to destroyed."""
        self.object.status = Ornament.StatusCategory.DESTROYED
        self.object.save()
        messages.success(self.request, f"Ornament '{self.object.ornament_name}' marked as destroyed.")
        return super().form_valid(form)


def multiple_ornament_create(request):
    """Create multiple ornaments at once using a model formset.

    This provides a separate page where you can enter several ornaments in
    one go, similar in spirit to the order create page.
    """

    OrnamentFormSet = modelformset_factory(
        Ornament,
        form=OrnamentForm,
        extra=5,
        can_delete=False,
    )

    if request.method == "POST":
        formset = OrnamentFormSet(request.POST, request.FILES, queryset=Ornament.objects.none())
        if formset.is_valid():
            formset.save()
            return redirect('ornament:list')
    else:
        formset = OrnamentFormSet(queryset=Ornament.objects.none())

    return render(request, 'ornament/ornament_bulk_form.html', {
        'formset': formset,
    })


def print_view(request):
    ornament = Ornament.objects.filter(id__gte=1).order_by('id')

    total_weight = ornament.aggregate(
        total=Sum('weight')
    )['total'] or 0

    return render(request, "ornament/print_view.html", {
        "ornament": ornament,
        "total_weight": total_weight
    })


def export_excel(request):
    view = OrnamentListView()
    view.request = request  # attach request
    ornaments = view.get_queryset()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ornaments"

    headers = [
        "Ornament Date", "Code", "Metal Type","Type", "Ornament Type",
        "MainCategory", "SubCategory", "Ornament Name","Gross Weight",
        "Weight", "Diamond/Stones Weight","Diamond Rate","Zircon Weight","Stone Weight",
        "Stone Price Per Carat","Stone Total Price",
        "Jarti","Jyala","Kaligar","Description","Image","Order","Created at","Updated at","Status"
    ]
    ws.append(headers)

    for o in ornaments:
        ws.append([
            str(o.ornament_date),
            o.code,
            o.metal_type,
            o.type,
            o.ornament_type,
            o.maincategory.name if o.maincategory else "",
            o.subcategory.name if o.subcategory else "",
            o.ornament_name,
            o.gross_weight,
            o.weight,
            o.diamond_weight,
            o.diamond_rate,
            o.zircon_weight,
            o.stone_weight,
            o.stone_percaratprice,
            o.stone_totalprice,
            o.jarti,
            o.jyala,
            o.kaligar.name if o.kaligar else "",
            o.description if o.description else "",
            str(o.image) if o.image else "",
            str(o.order) if o.order else "",
            str(o.created_at),
            str(o.updated_at),
            "fetched"
        ])

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Create virtual file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="ornaments.xlsx"'

    return response



def to_decimal(val):
    if val is None or val == "":
        return Decimal("0")
    return Decimal(str(val))   # SAFE conversion from float → Decimal


def import_excel(request):
    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("ornament:import_excel")

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            imported = 0
            skipped = 0
            skipped_fetched = 0
            skipped_duplicate = 0
            errors = []

            expected_cols = 25
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                # Pad or trim row to expected columns
                row_list = list(row) if row else []
                if len(row_list) > expected_cols:
                    row_list = row_list[:expected_cols]
                elif len(row_list) < expected_cols:
                    row_list = row_list + [None] * (expected_cols - len(row_list))
                try:
                    (
                        ornament_date_bs,
                        code,
                        metal_type,
                        type,
                        ornament_type,
                        maincategory_name,
                        subcategory_name,
                        ornament_name,
                        gross_weight,
                        weight,
                        diamond_weight,
                        diamond_rate,
                        zircon_weight,
                        stone_weight,
                        stone_percaratprice,
                        stone_totalprice,
                        jarti,
                        jyala,
                        kaligar_name,
                        description,
                        image,
                        order,
                        created_at,
                        updated_at,
                        status
                    ) = row_list
                except Exception as e:
                    errors.append(f"Row {idx}: Column mismatch or missing data. {e}")
                    skipped += 1
                    continue

                # Skip duplicates only if code is not empty and already exists
                if code and str(code).strip() and Ornament.objects.filter(code=str(code).strip()).exists():
                    skipped += 1
                    skipped_duplicate += 1
                    continue
                # MainCategory
                maincategory = None
                if maincategory_name:
                    maincategory = MainCategory.objects.filter(name=str(maincategory_name)).first()
                if not maincategory:
                    maincategory = MainCategory.objects.create(name=maincategory_name or "Unknown")
                # SubCategory
                subcategory = None
                if subcategory_name:
                    subcategory = SubCategory.objects.filter(name=str(subcategory_name)).first()
                if not subcategory:
                    subcategory = SubCategory.objects.create(name=subcategory_name or "Unknown")
                # Kaligar
                kaligar = None
                if kaligar_name:
                    kaligar = Kaligar.objects.filter(name=str(kaligar_name)).first()
                if not kaligar:
                    # Always create a valid 9-digit PAN
                    pan = "123456789"
                    try:
                        pan = str(int(getattr(kaligar, 'panno', 123456789))).zfill(9)
                    except Exception:
                        pan = "123456789"
                    kaligar = Kaligar.objects.create(
                        name=kaligar_name or "Unknown",
                        phone_no="",
                        panno=pan,
                        address="",
                        stamp=""
                    )
                # Order
                linked_order = None
                if order:
                    try:
                        linked_order = Order.objects.filter(sn=int(str(order).strip())).first()
                    except Exception:
                        linked_order = None
                    if linked_order is None:
                        import re
                        match = re.search(r"(\d+)", str(order))
                        if match:
                            try:
                                sn_val = int(match.group(1))
                                linked_order = Order.objects.filter(sn=sn_val).first()
                            except Exception:
                                linked_order = None
                # Date
                try:
                    y, m, d = map(int, str(ornament_date_bs).split("-"))
                    ornament_date = ndt.date(y, m, d)
                except Exception:
                    ornament_date = ndt.date.today()
                # Decimals
                gross_weight = to_decimal(gross_weight)
                weight = to_decimal(weight)
                diamond_weight = to_decimal(diamond_weight)
                diamond_rate = to_decimal(diamond_rate)
                zircon_weight = to_decimal(zircon_weight)
                stone_weight = to_decimal(stone_weight)
                stone_percaratprice = to_decimal(stone_percaratprice)
                stone_totalprice = to_decimal(stone_totalprice)
                jarti = to_decimal(jarti)
                jyala = to_decimal(jyala)
                # Create
                try:
                    Ornament.objects.create(
                        ornament_date=str(ornament_date),
                        code=code,
                        metal_type=metal_type,
                        type=type,
                        ornament_type=ornament_type,
                        maincategory=maincategory,
                        subcategory=subcategory,
                        ornament_name=ornament_name,
                        gross_weight=gross_weight,
                        weight=weight,
                        diamond_weight=diamond_weight,
                        diamond_rate=diamond_rate,
                        zircon_weight=zircon_weight,
                        stone_weight=stone_weight,
                        stone_percaratprice=stone_percaratprice,
                        stone_totalprice=stone_totalprice,
                        jarti=jarti,
                        jyala=jyala,
                        kaligar=kaligar,
                        description=description,
                        image=image,
                        order=linked_order,
                        created_at=created_at,
                        updated_at=updated_at,
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {idx}: Failed to import. {e}")
                    skipped += 1
                    continue

            msg = f"Imported: {imported} | Skipped: {skipped}"
            if skipped_fetched > 0:
                msg += f" (Fetched: {skipped_fetched})"
            if skipped_duplicate > 0:
                msg += f" (Duplicates: {skipped_duplicate})"
            if errors:
                msg += f" | Errors: {len(errors)}"
                for err in errors[:10]:  # Show first 10 errors
                    messages.error(request, err)
                if len(errors) > 10:
                    messages.warning(request, f"... and {len(errors) - 10} more errors")
            messages.success(request, msg)
            return redirect("ornament:list")

        except Exception as e:
            messages.error(request, f"Error while importing: {e}")
            return redirect("ornament:import_excel")

    return render(request, "ornament/import_excel.html")
@login_required
def export_stone_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stones"

    headers = ["Name", "Cost Per Carat", "Carat", "Sales Per Carat"]
    ws.append(headers)

    for stone in Stone.objects.all().order_by('-id'):
        ws.append([
            stone.name,
            float(stone.cost_per_carat) if stone.cost_per_carat is not None else '',
            float(stone.carat) if stone.carat is not None else '',
            float(stone.sales_per_carat) if stone.sales_per_carat is not None else '',
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="stones.xlsx"'
    return response

@login_required
def import_stone_excel(request):
    if request.method != "POST":
        return redirect('ornament:stone_list')

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Please upload an Excel file.")
        return redirect('ornament:stone_list')

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        skipped = 0
        errors = []

        expected_cols = 4
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_list = list(row) if row else []
            if len(row_list) > expected_cols:
                row_list = row_list[:expected_cols]
            elif len(row_list) < expected_cols:
                row_list = row_list + [None] * (expected_cols - len(row_list))

            try:
                name, cost_per_carat, carat, sales_per_carat = row_list
            except Exception as e:
                errors.append(f"Row {idx}: {e}")
                skipped += 1
                continue

            if not name:
                skipped += 1
                continue

            Stone.objects.create(
                name=str(name).strip(),
                cost_per_carat=to_decimal(cost_per_carat),
                carat=to_decimal(carat),
                sales_per_carat=to_decimal(sales_per_carat),
            )
            imported += 1

        messages.success(request, f"Stone import completed. Imported: {imported}, Skipped: {skipped}")
        if errors:
            messages.warning(request, f"Some rows had errors: {errors[:5]}")
    except Exception as e:
        messages.error(request, f"Failed to import stones: {e}")

    return redirect('ornament:stone_list')

@login_required
def export_motimala_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Motimala"

    headers = ["Name", "Cost Per Mala", "Quantity", "Sales Per Mala"]
    ws.append(headers)

    for moti in Motimala.objects.all().order_by('-id'):
        ws.append([
            moti.name,
            float(moti.cost_per_mala) if moti.cost_per_mala is not None else '',
            moti.quantity if moti.quantity is not None else '',
            float(moti.sales_per_mala) if moti.sales_per_mala is not None else '',
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="motimala.xlsx"'
    return response

@login_required
def import_motimala_excel(request):
    if request.method != "POST":
        return redirect('ornament:motimala_list')

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Please upload an Excel file.")
        return redirect('ornament:motimala_list')

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        skipped = 0
        errors = []

        expected_cols = 4
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_list = list(row) if row else []
            if len(row_list) > expected_cols:
                row_list = row_list[:expected_cols]
            elif len(row_list) < expected_cols:
                row_list = row_list + [None] * (expected_cols - len(row_list))

            try:
                name, cost_per_mala, quantity, sales_per_mala = row_list
            except Exception as e:
                errors.append(f"Row {idx}: {e}")
                skipped += 1
                continue

            if not name:
                skipped += 1
                continue

            Motimala.objects.create(
                name=str(name).strip(),
                cost_per_mala=to_decimal(cost_per_mala),
                quantity=int(quantity) if quantity not in (None, '') else 0,
                sales_per_mala=to_decimal(sales_per_mala),
            )
            imported += 1

        messages.success(request, f"Motimala import completed. Imported: {imported}, Skipped: {skipped}")
        if errors:
            messages.warning(request, f"Some rows had errors: {errors[:5]}")
    except Exception as e:
        messages.error(request, f"Failed to import motimala: {e}")

    return redirect('ornament:motimala_list')

@login_required
def export_potey_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Potey"

    headers = ["Name", "Loon", "Cost Per Loon", "Sales Per Loon"]
    ws.append(headers)

    for potey in Potey.objects.all().order_by('-id'):
        ws.append([
            potey.name,
            potey.loon if potey.loon is not None else '',
            float(potey.cost_per_loon) if potey.cost_per_loon is not None else '',
            float(potey.sales_per_loon) if potey.sales_per_loon is not None else '',
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="potey.xlsx"'
    return response

@login_required
def import_potey_excel(request):
    if request.method != "POST":
        return redirect('ornament:potey_list')

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Please upload an Excel file.")
        return redirect('ornament:potey_list')

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        skipped = 0
        errors = []

        expected_cols = 4
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_list = list(row) if row else []
            if len(row_list) > expected_cols:
                row_list = row_list[:expected_cols]
            elif len(row_list) < expected_cols:
                row_list = row_list + [None] * (expected_cols - len(row_list))

            try:
                name, loon, cost_per_loon, sales_per_loon = row_list
            except Exception as e:
                errors.append(f"Row {idx}: {e}")
                skipped += 1
                continue

            if not name:
                skipped += 1
                continue

            Potey.objects.create(
                name=str(name).strip(),
                loon=int(loon) if loon not in (None, '') else 0,
                cost_per_loon=to_decimal(cost_per_loon),
                sales_per_loon=to_decimal(sales_per_loon),
            )
            imported += 1

        messages.success(request, f"Potey import completed. Imported: {imported}, Skipped: {skipped}")
        if errors:
            messages.warning(request, f"Some rows had errors: {errors[:5]}")
    except Exception as e:
        messages.error(request, f"Failed to import potey: {e}")

    return redirect('ornament:potey_list')


def ornament_report(request):
    """Show ornament counts grouped by metal type, then by main category."""
    from django.db.models import Count

    total_ornaments = Ornament.objects.filter(
        ornament_type__in=[
            Ornament.OrnamentCategory.STOCK,
            Ornament.OrnamentCategory.ORDER,
        ]
    ).count()

    # Totals per metal type
    metal_totals = {
        row['metal_type']: row['count']
        for row in Ornament.objects.filter(
            ornament_type__in=[
                Ornament.OrnamentCategory.STOCK,
                Ornament.OrnamentCategory.ORDER,
            ]
        ).values('metal_type').annotate(count=Count('id'))
    }

    # Category breakdown per metal
    category_rows = (
        Ornament.objects.filter(
            ornament_type__in=[
                Ornament.OrnamentCategory.STOCK,
                Ornament.OrnamentCategory.ORDER,
            ]
        )
        .values('metal_type', 'maincategory__id', 'maincategory__name')
        .annotate(count=Count('id'))
        .order_by('metal_type', 'maincategory__name')
    )

    metal_order = ['Diamond', 'Gold', 'Silver', 'Others']
    sections = []

    for metal in metal_order:
        total = metal_totals.get(metal, 0)
        if total == 0:
            continue

        categories = [
            {
                'id': row['maincategory__id'],
                'name': row['maincategory__name'] or 'Uncategorized',
                'count': row['count'],
            }
            for row in category_rows
            if row['metal_type'] == metal
        ]

        sections.append({
            'metal': metal,
            'total': total,
            'categories': categories,
        })

    context = {
        'sections': sections,
        'total_ornaments': total_ornaments,
    }
    return render(request, 'ornament/ornament_report.html', context)


def ornament_weight_report(request):
    """Show ornament total net weight by metal type."""
    from django.db.models import F, Count
    # Use latest fetched daily rates for per tola pricing
    from main.models import DailyRate

    # Only include ornaments with positive weight and status active
    base_qs = Ornament.objects.filter(
        ornament_type=Ornament.OrnamentCategory.STOCK,
        status=Ornament.StatusCategory.ACTIVE,
        weight__gt=0
    )

    # Get latest daily gold/silver per tola rates
    daily_rate = DailyRate.objects.order_by('-created_at').first()
    gold_rate = daily_rate.gold_rate if daily_rate else Decimal('0')
    silver_rate = daily_rate.silver_rate if daily_rate else Decimal('0')
    
    # Fallback to Stock rates if Daily rates not available
    if gold_rate == 0 or silver_rate == 0:
        try:
            stock = Stock.objects.latest('year')
            if gold_rate == 0:
                gold_rate = stock.gold_rate
            if silver_rate == 0:
                silver_rate = stock.silver_rate
        except Stock.DoesNotExist:
            pass

    # Group by metal type
    weight_by_metal = list(
        base_qs.values('metal_type').annotate(
            total_weight=Sum('weight'),
            total_gross_weight=Sum('gross_weight'),
            total_jarti=Sum('jarti'),
            total_jyala=Sum('jyala'),
            ornament_count=Count('id')
        ).order_by('metal_type')
    )

    grand_total_amount = Decimal('0')
    total_gold_amount_with_diamond = Decimal('0')
    total_silver_amount = Decimal('0')
    total_diamond_amount = Decimal('0')
    total_jarti_amount_all = Decimal('0')
    total_jyala_amount_all = Decimal('0')
    total_silver_jarti_amount_all = Decimal('0')
    total_silver_jyala_amount_all = Decimal('0')
    total_silver_stone_amount_all = Decimal('0')

    for metal in weight_by_metal:
        mtype = metal['metal_type']
        mqs = base_qs.filter(metal_type=mtype)
        # Karat breakdown
        karats = mqs.values('type').annotate(weight_sum=Sum('weight'))
        karat_dict = {k['type']: k['weight_sum'] or Decimal('0') for k in karats}
        k24 = karat_dict.get('24KARAT', Decimal('0'))
        k22 = karat_dict.get('22KARAT', Decimal('0'))
        k18 = karat_dict.get('18KARAT', Decimal('0'))
        k14 = karat_dict.get('14KARAT', Decimal('0'))
        eq_24k = k24 + (k22 * Decimal('0.92')) + (k18 * Decimal('0.75')) + (k14 * Decimal('0.58'))
        # Stone price
        stone_total = mqs.aggregate(total=Sum('stone_totalprice'))['total'] or Decimal('0')
        # Jarti/Jyala
        total_jarti = metal['total_jarti'] or Decimal('0')
        total_jyala = metal['total_jyala'] or Decimal('0')
        # Amount and per-metal context fields
        if mtype == 'Gold':
            metal['gold_24k'] = k24
            metal['gold_22k'] = k22
            metal['gold_18k'] = k18
            metal['gold_14k'] = k14
            metal['gold_24k_equivalent'] = eq_24k
            metal['gold_amount'] = (eq_24k / Decimal('11.664')) * gold_rate
            metal['jarti_amount'] = (total_jarti / Decimal('11.664')) * gold_rate
            metal['jyala_amount'] = total_jyala
            metal['stone_amount'] = stone_total
            amount = metal['gold_amount'] + metal['jarti_amount'] + metal['jyala_amount'] + metal['stone_amount']
            total_gold_amount_with_diamond += metal['gold_amount']
            total_jarti_amount_all += metal['jarti_amount']
            total_jyala_amount_all += metal['jyala_amount']
        elif mtype == 'Silver':
            metal['silver_24k'] = k24
            metal['silver_22k'] = k22
            metal['silver_18k'] = k18
            metal['silver_14k'] = k14
            metal['silver_24k_equivalent'] = eq_24k
            metal['silver_amount'] = (eq_24k / Decimal('11.664')) * silver_rate
            metal['jarti_amount'] = (total_jarti / Decimal('11.664')) * silver_rate
            metal['jyala_amount'] = total_jyala
            metal['stone_amount'] = stone_total
            amount = metal['silver_amount'] + metal['jarti_amount'] + metal['jyala_amount'] + metal['stone_amount']
            total_silver_amount += metal['silver_amount']
            total_silver_jarti_amount_all += metal['jarti_amount']
            total_silver_jyala_amount_all += metal['jyala_amount']
            total_silver_stone_amount_all += metal['stone_amount']
            total_jarti_amount_all += metal['jarti_amount']
            total_jyala_amount_all += metal['jyala_amount']
        elif mtype == 'Diamond':
            # Diamond ornament calculations (as per provided specification)
            # Gold net weight in diamond ornaments
            diamond_gold_net_weight = metal['total_weight'] or Decimal('0')
            # Convert to 24K equivalent using 0.59 factor
            diamond_gold_net_24k = diamond_gold_net_weight * Decimal('0.59')
            # Total gold price: 24K equivalent converted to tola (/11.664) × daily gold rate per tola
            gold_amount = (diamond_gold_net_24k / Decimal('11.664')) * gold_rate
            # Jyala: gold net weight × 1800
            jyala_amount = diamond_gold_net_weight * Decimal('1800')
            # Diamond price: total diamond weight × 35000
            diamond_weight_total = mqs.aggregate(dw=Sum('diamond_weight'))['dw'] or Decimal('0')
            diamond_weight_amount = diamond_weight_total * Decimal('35000')
            # Total (exclude jarti/stone per spec unless required)
            amount = gold_amount + jyala_amount + diamond_weight_amount
            metal['diamond_24k'] = k24
            metal['diamond_22k'] = k22
            metal['diamond_18k'] = k18
            metal['diamond_14k'] = k14
            metal['diamond_24k_equivalent'] = diamond_gold_net_24k
            metal['gold_amount'] = gold_amount
            metal['diamond_weight_total'] = diamond_weight_total
            metal['diamond_weight_amount'] = diamond_weight_amount
            metal['jarti_amount'] = Decimal('0')
            metal['jyala_amount'] = jyala_amount
            metal['stone_amount'] = Decimal('0')
            total_gold_amount_with_diamond += gold_amount
            total_diamond_amount += diamond_weight_amount
            total_jyala_amount_all += jyala_amount
        else:
            amount = Decimal('0')
        # Common fields
        metal['karat_24k'] = k24
        metal['karat_22k'] = k22
        metal['karat_18k'] = k18
        metal['karat_14k'] = k14
        metal['eq_24k'] = eq_24k
        metal['stone_total'] = stone_total
        metal['total_amount'] = amount
        grand_total_amount += amount

    # Overall totals
    totals = base_qs.aggregate(
        total_weight=Sum('weight'),
        total_gross_weight=Sum('gross_weight'),
        total_jarti=Sum('jarti'),
        total_jyala=Sum('jyala'),
        total_count=Count('id')
    )
    # Overall 24k equivalent for Gold
    gold_qs = base_qs.filter(metal_type='Gold')
    gold_karats = gold_qs.values('type').annotate(weight_sum=Sum('weight'))
    gold_karat_dict = {k['type']: k['weight_sum'] or Decimal('0') for k in gold_karats}
    gold_24k_total = gold_karat_dict.get('24KARAT', Decimal('0'))
    gold_22k_total = gold_karat_dict.get('22KARAT', Decimal('0'))
    gold_18k_total = gold_karat_dict.get('18KARAT', Decimal('0'))
    gold_14k_total = gold_karat_dict.get('14KARAT', Decimal('0'))
    total_gold_24k_equivalent = gold_24k_total + (gold_22k_total * Decimal('0.92')) + (gold_18k_total * Decimal('0.75')) + (gold_14k_total * Decimal('0.58'))
    
    # Overall 24k equivalent for Silver
    silver_qs = base_qs.filter(metal_type='Silver')
    silver_karats = silver_qs.values('type').annotate(weight_sum=Sum('weight'))
    silver_karat_dict = {k['type']: k['weight_sum'] or Decimal('0') for k in silver_karats}
    silver_24k_total = silver_karat_dict.get('24KARAT', Decimal('0'))
    silver_22k_total = silver_karat_dict.get('22KARAT', Decimal('0'))
    silver_18k_total = silver_karat_dict.get('18KARAT', Decimal('0'))
    silver_14k_total = silver_karat_dict.get('14KARAT', Decimal('0'))
    total_silver_24k_equivalent = silver_24k_total + (silver_22k_total * Decimal('0.92')) + (silver_18k_total * Decimal('0.75')) + (silver_14k_total * Decimal('0.58'))
    
    # Overall 24k equivalent for Diamond (gold content in diamond ornaments)
    diamond_qs = base_qs.filter(metal_type='Diamond')
    diamond_total_weight = diamond_qs.aggregate(total=Sum('weight'))['total'] or Decimal('0')
    total_diamond_24k_equivalent = diamond_total_weight * Decimal('0.59')

    context = {
        'weight_by_metal': weight_by_metal,
        'totals': totals,
        'gold_rate': gold_rate,
        'grand_total_amount': grand_total_amount,
        'total_gold_24k_equivalent': total_gold_24k_equivalent,
        'total_silver_24k_equivalent': total_silver_24k_equivalent,
        'total_diamond_24k_equivalent': total_diamond_24k_equivalent,
        'total_gold_amount_with_diamond': total_gold_amount_with_diamond,
        'total_silver_amount': total_silver_amount,
        'total_diamond_amount': total_diamond_amount,
        'total_jarti_amount_all': total_jarti_amount_all,
        'total_jyala_amount_all': total_jyala_amount_all,
        'total_silver_jarti_amount_all': total_silver_jarti_amount_all,
        'total_silver_jyala_amount_all': total_silver_jyala_amount_all,
        'total_silver_stone_amount_all': total_silver_stone_amount_all,
    }
    return render(request, 'ornament/ornament_weight_report.html', context)


def rates_and_stock_view(request):
    """View to display fetched rates and stock year rates with dropdown."""
    from main.models import DailyRate
    
    # Get selected rate date and stock year from request
    selected_rate_bs_date = request.GET.get('rate_date')
    selected_stock_year = request.GET.get('stock_year')
    
    # Get all daily rates (fetched rates)
    daily_rates = DailyRate.objects.all().order_by('-created_at')
    
    # Get all stock years from both ornament and main app Stock models
    stock_years = Stock.objects.all().order_by('-year')
    
    # Initialize selected rate and stock data
    selected_rate = None
    selected_stock = None
    
    if selected_rate_bs_date:
        try:
            selected_rate = DailyRate.objects.get(bs_date=selected_rate_bs_date)
        except DailyRate.DoesNotExist:
            pass
    else:
        # Show the most recent rate by default
        selected_rate = daily_rates.first()
    
    if selected_stock_year:
        try:
            selected_stock = Stock.objects.get(year=selected_stock_year)
        except Stock.DoesNotExist:
            pass
    else:
        # Show the most recent stock year by default
        selected_stock = stock_years.first()
    
    # Calculate amounts for selected stock
    stock_amounts = {}
    if selected_stock:
        stock_amounts = {
            'diamond_amount': selected_stock.diamond_amount,
            'gold_amount': selected_stock.gold_amount,
            'silver_amount': selected_stock.silver_amount,
        }
    
    context = {
        'daily_rates': daily_rates,
        'stock_years': stock_years,
        'selected_rate': selected_rate,
        'selected_stock': selected_stock,
        'stock_amounts': stock_amounts,
        'selected_rate_date': selected_rate_bs_date,
        'selected_stock_year': selected_stock_year,
    }
    
    return render(request, 'ornament/rates_and_stock.html', context)


def kaligar_list(request):
    """Display list of all kaligar with their ornament weights."""
    from django.db.models import Sum, DecimalField
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    
    # Get all kaligar
    kaligars = Kaligar.objects.all().order_by('name')
    
    # Get selected kaligar if filtering
    selected_kaligar_id = request.GET.get('kaligar_id')
    selected_kaligar = None
    selected_kaligar_ornaments = None
    ornaments_by_metal_type = {}
    metal_type_totals = {}
    overall_totals = {
        'weight': Decimal('0'),
        'jarti': Decimal('0'),
        'jyala': Decimal('0'),
        'gross_weight': Decimal('0'),
        'diamond_weight': Decimal('0'),
        'zircon_weight': Decimal('0'),
        'stone_weight': Decimal('0'),
        'stone_totalprice': Decimal('0'),
    }
    
    kaligar_cash_accounts = None
    kaligar_gold_accounts = None
    if selected_kaligar_id:
        try:
            selected_kaligar = Kaligar.objects.get(id=selected_kaligar_id)
            all_ornaments = selected_kaligar.ornaments.all().order_by('-ornament_date')
            # Group ornaments by metal type and calculate totals
            for ornament in all_ornaments:
                metal_type = ornament.get_metal_type_display()
                if metal_type not in ornaments_by_metal_type:
                    ornaments_by_metal_type[metal_type] = []
                    metal_type_totals[metal_type] = {
                        'weight': Decimal('0'),
                        'jarti': Decimal('0'),
                        'jyala': Decimal('0'),
                        'gross_weight': Decimal('0'),
                        'diamond_weight': Decimal('0'),
                        'zircon_weight': Decimal('0'),
                        'stone_weight': Decimal('0'),
                        'stone_totalprice': Decimal('0'),
                    }
                ornaments_by_metal_type[metal_type].append(ornament)
                # Add to metal type totals
                metal_type_totals[metal_type]['weight'] += ornament.weight or Decimal('0')
                metal_type_totals[metal_type]['jarti'] += ornament.jarti or Decimal('0')
                metal_type_totals[metal_type]['jyala'] += ornament.jyala or Decimal('0')
                metal_type_totals[metal_type]['gross_weight'] += ornament.gross_weight or Decimal('0')
                metal_type_totals[metal_type]['diamond_weight'] += ornament.diamond_weight or Decimal('0')
                metal_type_totals[metal_type]['zircon_weight'] += ornament.zircon_weight or Decimal('0')
                metal_type_totals[metal_type]['stone_weight'] += ornament.stone_weight or Decimal('0')
                metal_type_totals[metal_type]['stone_totalprice'] += ornament.stone_totalprice or Decimal('0')
                # Add to overall totals
                overall_totals['weight'] += ornament.weight or Decimal('0')
                overall_totals['jarti'] += ornament.jarti or Decimal('0')
                overall_totals['jyala'] += ornament.jyala or Decimal('0')
                overall_totals['gross_weight'] += ornament.gross_weight or Decimal('0')
                overall_totals['diamond_weight'] += ornament.diamond_weight or Decimal('0')
                overall_totals['zircon_weight'] += ornament.zircon_weight or Decimal('0')
                overall_totals['stone_weight'] += ornament.stone_weight or Decimal('0')
                overall_totals['stone_totalprice'] += ornament.stone_totalprice or Decimal('0')
            # Add related accounts
            kaligar_cash_accounts = selected_kaligar.cash_accounts.all().order_by('-date')
            kaligar_gold_accounts = selected_kaligar.gold_accounts.all().order_by('-date')
        except Kaligar.DoesNotExist:
            pass
    
    # Annotate each kaligar with total ornament weight
    kaligars_with_weights = []
    total_ornament_count = 0
    for kaligar in kaligars:
        total_weight = kaligar.ornaments.aggregate(
            total=Coalesce(Sum('weight'), Decimal('0'), output_field=DecimalField())
        )['total']
        ornament_count = kaligar.ornaments.count()
        total_ornament_count += ornament_count
        kaligars_with_weights.append({
            'id': kaligar.id,
            'name': kaligar.name,
            'phone_no': kaligar.phone_no,
            'panno': kaligar.panno,
            'address': kaligar.address,
            'total_weight': total_weight,
            'ornament_count': ornament_count
        })
    
    context = {
        'kaligars': kaligars_with_weights,
        'selected_kaligar': selected_kaligar,
        'selected_kaligar_ornaments': selected_kaligar_ornaments,
        'ornaments_by_metal_type': ornaments_by_metal_type,
        'metal_type_totals': metal_type_totals,
        'overall_totals': overall_totals,
        'total_ornament_count': total_ornament_count,
        'kaligar_cash_accounts': kaligar_cash_accounts,
        'kaligar_gold_accounts': kaligar_gold_accounts,
    }
    
    return render(request, 'ornament/kaligar_list.html', context)
