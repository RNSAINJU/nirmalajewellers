
from decimal import Decimal
from io import BytesIO
from datetime import date
from datetime import timedelta
from calendar import monthrange
import openpyxl
from openpyxl import Workbook

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from .models import Loan, LoanInterestPayment, DhukutiLoan, DhukutiKistaPayment, DhukutiKistaPlan, EmiLoan, GoldLoanAccount, GoldLoanInterestPayment
from .forms import LoanForm, GoldLoanAccountForm
from common.nepali_utils import ad_to_bs_date_str


def _safe_decimal(val):
    try:
        return Decimal(str(val)) if val not in (None, '') else Decimal('0')
    except Exception:
        return Decimal('0')


def _extract_date(val):
    """Extract date in YYYY-MM-DD format from various input types."""
    if not val:
        return None
    try:
        # Convert to string and extract only date part (YYYY-MM-DD)
        date_str = str(val)
        # If it has time component, split it
        if ' ' in date_str:
            date_str = date_str.split(' ')[0]
        return date_str
    except Exception:
        return None


def _compute_dhukuti_summary(received_amount, total_kista, paid_amounts, remaining_base_payment=None, received_kista_number=1, planned_amounts_by_month=None, kista_increment=None):
    """Compute Dhukuti-style payment summary from variable monthly kista amounts."""
    if received_amount < 0:
        raise ValueError('Received amount cannot be negative.')
    if total_kista <= 0:
        raise ValueError('Total kista must be greater than zero.')

    paid_rows = []
    total_paid = Decimal('0.00')
    for idx, amount in enumerate(paid_amounts, start=1):
        amt = Decimal(str(amount)).quantize(Decimal('0.01'))
        if amt <= 0:
            continue
        paid_rows.append({'month': idx, 'amount': amt})
        total_paid += amt

    paid_kista = len(paid_rows)
    remaining_kista = max(total_kista - paid_kista, 0)
    raw_diff = received_amount - total_paid
    interest_on_paid_side = raw_diff < 0
    total_interest = abs(raw_diff).quantize(Decimal('0.01'))

    # Interest period starts from the kista where whole amount is received.
    # 0 means not yet received — treat same as no interest period started.
    if received_kista_number < 0:
        received_kista_number = 0
    if received_kista_number == 0:
        elapsed_interest_kista = 0
    else:
        effective_paid_kista = paid_kista - received_kista_number + 1
        elapsed_interest_kista = effective_paid_kista if effective_paid_kista > 0 else 0

    monthly_interest = Decimal('0.00')
    if received_amount == 0:
        interest_on_paid_side = False
        total_interest = Decimal('0.00')
    elif elapsed_interest_kista > 0:
        monthly_interest = (total_interest / Decimal(str(elapsed_interest_kista))).quantize(Decimal('0.000001'))

    average_interest_rate_percent = Decimal('0.00')
    if received_amount > 0:
        average_interest_rate_percent = ((total_interest / received_amount) * Decimal('100')).quantize(Decimal('0.01'))

    average_monthly_interest_rate_percent = Decimal('0.00')
    if elapsed_interest_kista > 0 and received_amount > 0:
        average_monthly_interest_rate_percent = ((monthly_interest / received_amount) * Decimal('100')).quantize(Decimal('0.01'))

    remaining_interest = (monthly_interest * Decimal(str(remaining_kista))).quantize(Decimal('0.01'))

    if remaining_base_payment is None or remaining_base_payment == Decimal('0'):
        avg_paid = Decimal('0.00')
        if paid_kista > 0:
            avg_paid = (total_paid / Decimal(str(paid_kista))).quantize(Decimal('0.01'))
        remaining_base_payment = (avg_paid * Decimal(str(remaining_kista))).quantize(Decimal('0.01'))
    else:
        remaining_base_payment = Decimal(str(remaining_base_payment)).quantize(Decimal('0.01'))

    to_pay_with_interest_adj = (remaining_base_payment - remaining_interest).quantize(Decimal('0.01'))
    remaining_kista_amount = Decimal('0.00')
    if remaining_kista > 0:
        remaining_kista_amount = (to_pay_with_interest_adj / Decimal(str(remaining_kista))).quantize(Decimal('0.01'))

    # Build a full list of all kista slots (paid + remaining) for display.
    # For remaining slots, project amounts using the average increment across paid kista.
    paid_month_set = {row['month']: row['amount'] for row in paid_rows}
    planned_month_set = {}
    if planned_amounts_by_month:
        for month, amount in planned_amounts_by_month.items():
            if month in paid_month_set:
                continue
            amt = Decimal(str(amount)).quantize(Decimal('0.01'))
            if amt > 0:
                planned_month_set[int(month)] = amt

    # Determine increment for projecting future kistas.
    # If user provided kista_increment, use it; otherwise future kistas equal the last paid kista (increment=0).
    if kista_increment is not None:
        avg_increment = Decimal(str(kista_increment)).quantize(Decimal('0.01'))
    else:
        avg_increment = Decimal('0.00')
    all_kista_rows = []
    prev_amount = None
    for i in range(1, total_kista + 1):
        if i in paid_month_set:
            amt = paid_month_set[i]
            increment = (amt - prev_amount) if prev_amount is not None else None
            all_kista_rows.append({'month': i, 'amount': amt, 'paid': True, 'planned': False, 'increment': increment, 'projected': False})
            prev_amount = amt
        elif i in planned_month_set:
            amt = planned_month_set[i]
            increment = (amt - prev_amount) if prev_amount is not None else None
            all_kista_rows.append({'month': i, 'amount': amt, 'paid': False, 'planned': True, 'increment': increment, 'projected': False})
            prev_amount = amt
        else:
            if prev_amount is None:
                projected = remaining_kista_amount.quantize(Decimal('0.01'))
                increment = None
            else:
                projected = (prev_amount + avg_increment).quantize(Decimal('0.01'))
                if avg_increment > Decimal('0') and projected <= prev_amount:
                    projected = prev_amount + Decimal('0.01')
                increment = projected - prev_amount
            all_kista_rows.append({'month': i, 'amount': projected, 'paid': False, 'planned': False, 'increment': increment, 'projected': True})
            prev_amount = projected

    # Estimation should follow actual remaining kista rows shown to the user
    # (planned values entered by user + projected values for not-yet-provided kista).
    remaining_rows = [row for row in all_kista_rows if not row['paid']]
    estimated_remaining_total = sum((row['amount'] for row in remaining_rows), Decimal('0.00')).quantize(Decimal('0.01'))
    estimated_remaining_per_kista = Decimal('0.00')
    if remaining_kista > 0:
        estimated_remaining_per_kista = (estimated_remaining_total / Decimal(str(remaining_kista))).quantize(Decimal('0.01'))

    return {
        'paid_rows': paid_rows,
        'all_kista_rows': all_kista_rows,
        'total_kista': total_kista,
        'paid_kista': paid_kista,
        'received_kista_number': received_kista_number,
        'elapsed_interest_kista': elapsed_interest_kista,
        'remaining_kista': remaining_kista,
        'total_paid': total_paid.quantize(Decimal('0.01')),
        'received_amount': received_amount.quantize(Decimal('0.01')),
        'interest_on_paid_side': interest_on_paid_side,
        'total_interest': total_interest,
        'monthly_interest': monthly_interest,
        'average_interest_rate_percent': average_interest_rate_percent,
        'average_monthly_interest_rate_percent': average_monthly_interest_rate_percent,
        'remaining_interest': remaining_interest,
        'remaining_base_payment': remaining_base_payment,
        'to_pay_with_interest_adjustment': to_pay_with_interest_adj,
        'remaining_kista_amount': remaining_kista_amount,
        'estimated_remaining_total': estimated_remaining_total,
        'estimated_remaining_per_kista': estimated_remaining_per_kista,
    }


def _compute_emi_schedule(principal, annual_rate, tenure_months):
    """Compute EMI and amortization schedule."""
    if principal <= 0 or annual_rate < 0 or tenure_months <= 0:
        raise ValueError('Principal must be > 0, rate >= 0 and tenure > 0.')

    monthly_rate = annual_rate / Decimal('1200')
    n = int(tenure_months)

    if monthly_rate == 0:
        emi = (principal / Decimal(str(n))).quantize(Decimal('0.01'))
    else:
        factor = (Decimal('1') + monthly_rate) ** n
        emi = (principal * monthly_rate * factor / (factor - Decimal('1'))).quantize(Decimal('0.01'))

    balance = principal.quantize(Decimal('0.01'))
    total_interest = Decimal('0.00')
    total_payment = Decimal('0.00')
    schedule = []

    for month in range(1, n + 1):
        interest_component = (balance * monthly_rate).quantize(Decimal('0.01'))
        principal_component = (emi - interest_component).quantize(Decimal('0.01'))

        if month == n or principal_component > balance:
            principal_component = balance
            installment = (principal_component + interest_component).quantize(Decimal('0.01'))
        else:
            installment = emi

        closing_balance = (balance - principal_component).quantize(Decimal('0.01'))
        if closing_balance < 0:
            closing_balance = Decimal('0.00')

        schedule.append({
            'month': month,
            'opening_balance': balance,
            'installment': installment,
            'interest_component': interest_component,
            'principal_component': principal_component,
            'closing_balance': closing_balance,
        })

        total_interest += interest_component
        total_payment += installment
        balance = closing_balance

    return {
        'emi': emi,
        'total_interest': total_interest.quantize(Decimal('0.01')),
        'total_payment': total_payment.quantize(Decimal('0.01')),
        'schedule': schedule,
    }


@login_required
def loan_list(request):
    status_filter = request.GET.get('status', 'active')
    all_loans = Loan.objects.all().order_by('-start_date', '-created_at')

    if status_filter == 'settled':
        loans = all_loans.filter(is_settled=True)
    elif status_filter == 'all':
        loans = all_loans
    else:
        loans = all_loans.filter(is_settled=False)

    # Calculate total loan amount for shown loans
    total_loan_amount = loans.aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    # Calculate tentative interest totals for active loans only
    active_loans = all_loans.filter(is_settled=False)
    total_monthly_interest = Decimal('0')
    total_quarterly_interest = Decimal('0')
    total_yearly_interest = Decimal('0')
    for loan in active_loans:
        total_monthly_interest += loan.monthly_interest
        total_quarterly_interest += loan.quarterly_interest
        total_yearly_interest += loan.yearly_interest

    # Calculate bank-wise totals for shown loans
    bank_totals = {}
    for loan in loans:
        bank_name = loan.bank_name
        if bank_name not in bank_totals:
            bank_totals[bank_name] = {
                'total_amount': Decimal('0'),
                'loan_count': 0
            }
        bank_totals[bank_name]['total_amount'] += loan.amount
        bank_totals[bank_name]['loan_count'] += 1

    # Sort banks by amount (descending)
    sorted_banks = sorted(bank_totals.items(), key=lambda x: x[1]['total_amount'], reverse=True)

    # Stats
    active_count = active_loans.count()
    settled_count = all_loans.filter(is_settled=True).count()

    context = {
        'loans': loans,
        'total_loan_amount': total_loan_amount,
        'total_monthly_interest': total_monthly_interest,
        'total_quarterly_interest': total_quarterly_interest,
        'total_yearly_interest': total_yearly_interest,
        'bank_totals': sorted_banks,
        'status_filter': status_filter,
        'active_count': active_count,
        'settled_count': settled_count,
    }
    return render(request, 'finance/loan_list.html', context)


@login_required
def gold_loan_account_list(request):
    accounts = GoldLoanAccount.objects.all().prefetch_related('interest_payments').order_by('-loan_taken_date', '-created_at')

    total_loan_given = Decimal('0.00')
    total_unpaid_interest = Decimal('0.00')
    for account in accounts:
        loan_amount = Decimal(str(account.loan_amount or Decimal('0.00')))
        accrued_interest = Decimal(str(account.accrued_interest_to_date or Decimal('0.00')))
        paid_interest = sum(
            Decimal(str(payment.interest_amount or Decimal('0.00')))
            for payment in account.interest_payments.all()
        )

        unpaid_interest = accrued_interest - paid_interest
        if unpaid_interest < 0:
            unpaid_interest = Decimal('0.00')

        total_loan_given += loan_amount
        total_unpaid_interest += unpaid_interest

    total_loan_given = total_loan_given.quantize(Decimal('0.01'))
    total_unpaid_interest = total_unpaid_interest.quantize(Decimal('0.01'))

    if request.method == 'POST':
        form = GoldLoanAccountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gold loan account added successfully!')
            return redirect('finance:gold_loan_account_list')
    else:
        form = GoldLoanAccountForm()

    context = {
        'accounts': accounts,
        'form': form,
        'title': 'Gold Loan Accounts',
        'total_accounts': accounts.count(),
        'total_loan_given': total_loan_given,
        'total_unpaid_interest': total_unpaid_interest,
    }
    return render(request, 'finance/gold_loan_account_list.html', context)


@login_required
def gold_loan_account_detail(request, pk):
    account = get_object_or_404(GoldLoanAccount, pk=pk)
    monthly_interest = Decimal(str(account.effective_monthly_interest or Decimal('0.00')))
    penalty_rate = Decimal(str(account.penalty_rate or Decimal('0.00')))
    breakdown = []
    total_breakdown_interest = Decimal('0.00')
    total_paid_interest = Decimal('0.00')
    total_penalty_interest = Decimal('0.00')

    start_date = account.loan_taken_ad_date
    today = date.today()
    paid_periods = {
        (entry.period_start_ad.isoformat(), entry.period_end_ad.isoformat()): entry
        for entry in account.interest_payments.all()
    }

    def _bs_day(ad_dt):
        bs_text = ad_to_bs_date_str(ad_dt)
        try:
            return int(bs_text.split('-')[2])
        except Exception:
            return 1

    def _bs_month_start(ad_dt):
        cursor = ad_dt
        # Walk backward until BS day becomes 1.
        for _ in range(35):
            if _bs_day(cursor) == 1:
                return cursor
            cursor -= timedelta(days=1)
        return ad_dt

    def _bs_month_end(ad_dt):
        cursor = ad_dt
        # Walk forward until next BS day becomes 1.
        for _ in range(35):
            next_day = cursor + timedelta(days=1)
            if _bs_day(next_day) == 1:
                return cursor
            cursor = next_day
        return ad_dt

    if start_date and start_date <= today and monthly_interest > 0:
        cursor = start_date
        index = 1

        while cursor <= today:
            month_start = _bs_month_start(cursor)
            month_end = _bs_month_end(cursor)
            month_total_days = (month_end - month_start).days + 1

            active_start = start_date if start_date > month_start else month_start
            active_end = today if today < month_end else month_end

            if active_start <= active_end:
                active_days = (active_end - active_start).days + 1

                # Charge a full month's interest when the whole BS month is covered.
                if active_start == month_start and active_end == month_end:
                    month_interest = monthly_interest.quantize(Decimal('0.01'))
                else:
                    month_interest = (
                        monthly_interest * Decimal(str(active_days)) / Decimal(str(month_total_days))
                    ).quantize(Decimal('0.01'))

                total_breakdown_interest += month_interest

                breakdown.append({
                    'sn': index,
                    'month_label': ad_to_bs_date_str(month_start)[:7],
                    'from_date': active_start,
                    'to_date': active_end,
                    'active_days': active_days,
                    'month_total_days': month_total_days,
                    'monthly_interest': monthly_interest,
                    'interest_amount': month_interest,
                    'period_start_iso': active_start.isoformat(),
                    'period_end_iso': active_end.isoformat(),
                })
                index += 1

            cursor = month_end + timedelta(days=1)

    total_rows = len(breakdown)

    outstanding_due_with_penalty = Decimal('0.00')
    running_total_interest = Decimal('0.00')
    total_active_days = 0
    total_monthly_interest_column = Decimal('0.00')

    for row in breakdown:
        payment = paid_periods.get((row['period_start_iso'], row['period_end_iso']))
        row['is_paid'] = payment is not None
        row['paid_on'] = payment.paid_on if payment else None

        total_active_days += int(row['active_days'])
        total_monthly_interest_column += Decimal(str(row['monthly_interest']))

        running_total_interest += row['interest_amount']
        row['running_total_interest'] = running_total_interest.quantize(Decimal('0.01'))

        # Penalty style requested: every unpaid month gets 4% of
        # (previous outstanding with penalty + current month interest).
        if row['is_paid']:
            penalty_amount = Decimal('0.00')
        else:
            base_for_penalty = outstanding_due_with_penalty + row['interest_amount']
            penalty_amount = (base_for_penalty * penalty_rate / Decimal('100')).quantize(Decimal('0.01'))
            outstanding_due_with_penalty = (base_for_penalty + penalty_amount).quantize(Decimal('0.01'))

        row['penalty_amount'] = penalty_amount
        row['interest_with_penalty_cumulative'] = outstanding_due_with_penalty.quantize(Decimal('0.01'))

        if payment:
            total_paid_interest += payment.interest_amount
        else:
            total_penalty_interest += penalty_amount

    context = {
        'account': account,
        'interest_breakdown': breakdown,
        'total_active_days': total_active_days,
        'total_monthly_interest_column': total_monthly_interest_column.quantize(Decimal('0.01')),
        'total_breakdown_interest': total_breakdown_interest.quantize(Decimal('0.01')),
        'total_paid_interest': total_paid_interest.quantize(Decimal('0.01')),
        'total_unpaid_interest': (total_breakdown_interest - total_paid_interest).quantize(Decimal('0.01')),
        'total_penalty_interest': total_penalty_interest.quantize(Decimal('0.01')),
        'total_due_with_penalty': outstanding_due_with_penalty.quantize(Decimal('0.01')),
        'title': 'Gold Loan Customer Details',
    }
    return render(request, 'finance/gold_loan_account_detail.html', context)


@login_required
def gold_loan_account_update(request, pk):
    account = get_object_or_404(GoldLoanAccount, pk=pk)
    if request.method == 'POST':
        form = GoldLoanAccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gold loan account updated successfully!')
            return redirect('finance:gold_loan_account_detail', pk=account.pk)
    else:
        form = GoldLoanAccountForm(instance=account)

    context = {
        'form': form,
        'account': account,
        'title': 'Edit Gold Loan Account',
    }
    return render(request, 'finance/gold_loan_account_form.html', context)


@login_required
def gold_loan_interest_mark_paid(request, pk):
    account = get_object_or_404(GoldLoanAccount, pk=pk)
    if request.method != 'POST':
        return redirect('finance:gold_loan_account_detail', pk=account.pk)

    period_start_str = request.POST.get('period_start_ad', '')
    period_end_str = request.POST.get('period_end_ad', '')
    period_label_bs = request.POST.get('period_label_bs', '')
    interest_amount_raw = request.POST.get('interest_amount', '0')

    try:
        period_start_ad = date.fromisoformat(period_start_str)
        period_end_ad = date.fromisoformat(period_end_str)
        interest_amount = Decimal(str(interest_amount_raw)).quantize(Decimal('0.01'))
    except Exception:
        messages.error(request, 'Invalid monthly period data. Could not mark as paid.')
        return redirect('finance:gold_loan_account_detail', pk=account.pk)

    payment, created = GoldLoanInterestPayment.objects.get_or_create(
        account=account,
        period_start_ad=period_start_ad,
        period_end_ad=period_end_ad,
        defaults={
            'period_label_bs': period_label_bs,
            'interest_amount': interest_amount,
            'paid_on': _extract_date(ad_to_bs_date_str(date.today())),
        },
    )

    if created:
        messages.success(request, f'Interest for {period_label_bs} marked as paid.')
    else:
        messages.info(request, f'Interest for {period_label_bs} is already marked as paid.')

    return redirect('finance:gold_loan_account_detail', pk=account.pk)


@login_required
def gold_loan_interest_undo_paid(request, pk):
    account = get_object_or_404(GoldLoanAccount, pk=pk)
    if request.method != 'POST':
        return redirect('finance:gold_loan_account_detail', pk=account.pk)

    period_start_str = request.POST.get('period_start_ad', '')
    period_end_str = request.POST.get('period_end_ad', '')
    period_label_bs = request.POST.get('period_label_bs', '')

    try:
        period_start_ad = date.fromisoformat(period_start_str)
        period_end_ad = date.fromisoformat(period_end_str)
    except Exception:
        messages.error(request, 'Invalid monthly period data. Could not undo payment.')
        return redirect('finance:gold_loan_account_detail', pk=account.pk)

    deleted_count, _ = GoldLoanInterestPayment.objects.filter(
        account=account,
        period_start_ad=period_start_ad,
        period_end_ad=period_end_ad,
    ).delete()

    if deleted_count > 0:
        label = period_label_bs or f'{period_start_ad} to {period_end_ad}'
        messages.success(request, f'Interest payment for {label} has been undone.')
    else:
        messages.info(request, 'No paid record found for this month.')

    return redirect('finance:gold_loan_account_detail', pk=account.pk)


@login_required
def gold_loan_account_delete(request, pk):
    account = get_object_or_404(GoldLoanAccount, pk=pk)
    if request.method == 'POST':
        account.delete()
        messages.success(request, 'Gold loan account deleted successfully!')
    return redirect('finance:gold_loan_account_list')


@login_required
def loan_create(request):
    if request.method == 'POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Loan added successfully!')
            return redirect('finance:loan_list')
    else:
        form = LoanForm()
    return render(request, 'finance/loan_form.html', {'form': form, 'title': 'Add Loan'})


@login_required
def loan_update(request, pk):
    loan = get_object_or_404(Loan, pk=pk)
    if request.method == 'POST':
        form = LoanForm(request.POST, instance=loan)
        if form.is_valid():
            form.save()
            messages.success(request, 'Loan updated successfully!')
            return redirect('finance:loan_list')
    else:
        form = LoanForm(instance=loan)
    return render(request, 'finance/loan_form.html', {'form': form, 'title': 'Edit Loan', 'object': loan})


@login_required
def loan_delete(request, pk):
    loan = get_object_or_404(Loan, pk=pk)
    if request.method == 'POST':
        loan.delete()
        messages.success(request, 'Loan deleted successfully!')
        return redirect('finance:loan_list')
    return render(request, 'finance/loan_confirm_delete.html', {'object': loan})


@login_required
def loan_export(request):
    """Export all loans to XLSX file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Loans"
    
    ws.append(["bank_name", "amount", "interest_rate", "start_date", "notes"])
    for loan in Loan.objects.all().order_by('-start_date', '-created_at'):
        ws.append([
            loan.bank_name,
            float(loan.amount),
            float(loan.interest_rate),
            str(loan.start_date) if loan.start_date else '',
            loan.notes or '',
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="loans.xlsx"'
    return response


@login_required
def loan_import(request):
    """Import loans from XLSX file"""
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            updated = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                bank_name, amount, interest_rate, start_date, notes = row[:5]
                
                # Skip if bank name is empty
                if not bank_name:
                    continue
                
                start_date = _extract_date(start_date)
                if not start_date:
                    continue
                
                loan, created_flag = Loan.objects.update_or_create(
                    bank_name=bank_name,
                    start_date=start_date,
                    defaults={
                        'amount': _safe_decimal(amount),
                        'interest_rate': _safe_decimal(interest_rate),
                        'notes': notes or '',
                    }
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1
            
            messages.success(request, f'Imported {created} new loans, updated {updated} existing loans')
        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')
    
    return redirect('finance:loan_list')


@login_required
def loan_add_interest(request, pk):
    """Add a 3-month interest payment for a loan"""
    loan = get_object_or_404(Loan, pk=pk)
    if loan.is_settled and loan.interest_payments.exists():
        messages.error(request, 'Cannot add interest payment to a settled loan that already has payment records.')
        return redirect('finance:loan_list')

    # Pre-calculate 3-month (quarterly) interest
    quarterly_interest = loan.quarterly_interest

    if request.method == 'POST':
        try:
            amount = Decimal(request.POST.get('amount', '0'))
            payment_date = request.POST.get('payment_date', '')
            months_covered = Decimal(request.POST.get('months_covered', '3'))
            notes = request.POST.get('notes', '')

            if not payment_date:
                messages.error(request, 'Payment date is required.')
                return render(request, 'finance/loan_interest_payment_form.html', {
                    'loan': loan,
                    'quarterly_interest': quarterly_interest,
                    'monthly_interest': loan.monthly_interest,
                })

            LoanInterestPayment.objects.create(
                loan=loan,
                amount=amount,
                payment_date=payment_date,
                months_covered=months_covered,
                notes=notes,
            )
            messages.success(request, f'Interest payment of रु{amount} recorded for {loan.bank_name} ({months_covered} months).')
            return redirect('finance:loan_list')
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')

    # Determine the from_date for months_covered auto-calculation:
    # use last payment date if exists, else loan start date
    last_payment = loan.interest_payments.order_by('-payment_date', '-created_at').first()
    from_date = str(last_payment.payment_date) if last_payment else str(loan.start_date)

    return render(request, 'finance/loan_interest_payment_form.html', {
        'loan': loan,
        'quarterly_interest': quarterly_interest,
        'monthly_interest': loan.monthly_interest,
        'from_date': from_date,
    })


@login_required
def loan_settle(request, pk):
    """Mark a loan as fully settled/paid"""
    loan = get_object_or_404(Loan, pk=pk)

    if request.method == 'POST':
        settled_date = request.POST.get('settled_date', '')
        final_interest_paid = request.POST.get('final_interest_paid', '')
        settlement_months = request.POST.get('settlement_months', '')

        if not settled_date:
            messages.error(request, 'Settlement date is required.')
            return render(request, 'finance/loan_settle_confirm.html', {'loan': loan})

        try:
            loan.is_settled = True
            loan.settled_date = settled_date
            
            # Update final interest paid if provided
            if final_interest_paid:
                loan.final_interest_paid = Decimal(final_interest_paid)
            
            # Update settlement months if provided
            if settlement_months:
                loan.settlement_months = int(settlement_months)
            
            loan.save()
            
            # Build success message with calculated metrics
            message = f'Loan from {loan.bank_name} (रु{loan.amount}) has been marked as settled.'
            if loan.final_interest_paid:
                message += f' Final interest: रु{loan.final_interest_paid}.'
                if loan.effective_interest_rate:
                    message += f' Effective rate: {loan.effective_interest_rate:.2f}% p.a.'
            
            messages.success(request, message)
            return redirect('finance:loan_list')
        except (ValueError, TypeError) as e:
            messages.error(request, f'Error processing settlement: {str(e)}')
            return render(request, 'finance/loan_settle_confirm.html', {'loan': loan})

    return render(request, 'finance/loan_settle_confirm.html', {'loan': loan})


@login_required
def loan_interest_payment_edit(request, pk):
    """Edit a loan interest payment"""
    payment = get_object_or_404(LoanInterestPayment, pk=pk)
    loan = payment.loan

    if request.method == 'POST':
        try:
            payment.amount = Decimal(request.POST.get('amount', '0'))
            payment.payment_date = request.POST.get('payment_date', '')
            payment.months_covered = Decimal(request.POST.get('months_covered', '3'))
            payment.notes = request.POST.get('notes', '')
            if not payment.payment_date:
                messages.error(request, 'Payment date is required.')
            else:
                payment.save()
                messages.success(request, f'Interest payment updated.')
                return redirect('finance:loan_list')
        except Exception as e:
            messages.error(request, f'Error updating payment: {str(e)}')

    # Determine from_date: use the previous payment date or loan start date
    prev_payment = loan.interest_payments.filter(
        payment_date__lt=payment.payment_date
    ).order_by('-payment_date').first()
    from_date = str(prev_payment.payment_date) if prev_payment else str(loan.start_date)

    return render(request, 'finance/loan_interest_payment_form.html', {
        'loan': loan,
        'quarterly_interest': loan.quarterly_interest,
        'monthly_interest': loan.monthly_interest,
        'from_date': from_date,
        'payment': payment,
    })


@login_required
def loan_interest_payment_delete(request, pk):
    """Delete a loan interest payment"""
    payment = get_object_or_404(LoanInterestPayment, pk=pk)
    if request.method == 'POST':
        loan_pk = payment.loan.pk
        payment.delete()
        messages.success(request, 'Interest payment deleted.')
        return redirect('finance:loan_list')
    return render(request, 'finance/loan_interest_payment_confirm_delete.html', {'payment': payment})


@login_required
def loan_dhukuti_calculator(request):
    """Finance page to calculate Dhukuti-style payment summary."""
    dhukuti_loans = DhukutiLoan.objects.prefetch_related('paid_kistas', 'planned_kistas').all()
    selected_dhukuti = None
    result = None

    initial = {
        'name': '',
        'start_date': date.today().isoformat(),
        'received_amount': '',
        'total_kista': '20',
        'received_kista_number': '1',
        'remaining_base_payment': '',
        'kista_increment': '',
        'paid_amounts_text': '',
        'notes': '',
    }

    selected_dhukuti_id = request.GET.get('dhukuti') or request.POST.get('dhukuti_id')
    if selected_dhukuti_id:
        selected_dhukuti = DhukutiLoan.objects.filter(pk=selected_dhukuti_id).first()
        if selected_dhukuti:
            initial['name'] = selected_dhukuti.name
            initial['start_date'] = str(selected_dhukuti.start_date) if selected_dhukuti.start_date else ''
            initial['received_amount'] = str(selected_dhukuti.received_amount)
            initial['total_kista'] = str(selected_dhukuti.total_kista)
            initial['received_kista_number'] = str(selected_dhukuti.received_kista_number)
            initial['remaining_base_payment'] = str(selected_dhukuti.remaining_base_payment)
            initial['kista_increment'] = str(selected_dhukuti.kista_increment) if selected_dhukuti.kista_increment is not None else ''
            initial['notes'] = selected_dhukuti.notes or ''
            initial['paid_amounts_text'] = '\n'.join(
                str(p.amount) for p in selected_dhukuti.paid_kistas.all().order_by('month_number')
            )
            result = _compute_dhukuti_summary(
                received_amount=selected_dhukuti.received_amount,
                total_kista=selected_dhukuti.total_kista,
                paid_amounts=[p.amount for p in selected_dhukuti.paid_kistas.all().order_by('month_number')],
                remaining_base_payment=selected_dhukuti.remaining_base_payment,
                received_kista_number=selected_dhukuti.received_kista_number,
                planned_amounts_by_month={
                    p.month_number: p.amount for p in selected_dhukuti.planned_kistas.all().order_by('month_number')
                },
                kista_increment=selected_dhukuti.kista_increment,
            )

    if request.method == 'POST':
        initial['name'] = request.POST.get('name', '').strip()
        initial['start_date'] = request.POST.get('start_date', '').strip()
        initial['received_amount'] = request.POST.get('received_amount', '').strip()
        initial['total_kista'] = request.POST.get('total_kista', '').strip()
        initial['received_kista_number'] = request.POST.get('received_kista_number', '').strip()
        initial['remaining_base_payment'] = request.POST.get('remaining_base_payment', '').strip()
        initial['kista_increment'] = request.POST.get('kista_increment', '').strip()
        initial['paid_amounts_text'] = request.POST.get('paid_amounts_text', '').strip()
        initial['notes'] = request.POST.get('notes', '').strip()

        try:
            received_amount = Decimal(initial['received_amount'])
            start_date = date.fromisoformat(initial['start_date']) if initial['start_date'] else None
            total_kista = int(initial['total_kista'])
            received_kista_number = int(initial['received_kista_number'] or '1')

            paid_amounts = []
            for line in initial['paid_amounts_text'].splitlines():
                normalized = line.replace(',', '').strip()
                if not normalized:
                    continue
                paid_amounts.append(Decimal(normalized))

            remaining_base_payment = None
            if initial['remaining_base_payment']:
                remaining_base_payment = Decimal(initial['remaining_base_payment'])

            kista_increment = None
            if initial['kista_increment']:
                kista_increment = Decimal(initial['kista_increment'])

            result = _compute_dhukuti_summary(
                received_amount=received_amount,
                total_kista=total_kista,
                paid_amounts=paid_amounts,
                remaining_base_payment=remaining_base_payment,
                received_kista_number=received_kista_number,
                planned_amounts_by_month={
                    p.month_number: p.amount for p in selected_dhukuti.planned_kistas.all().order_by('month_number')
                } if selected_dhukuti else None,
                kista_increment=kista_increment,
            )

            record_name = initial['name'] or f"Dhukuti Loan {received_amount}"

            if request.POST.get('update_record') == '1':
                if not selected_dhukuti:
                    messages.error(request, 'No Dhukuti record selected to update. Please load a saved record first.')
                else:
                    selected_dhukuti.name = record_name
                    selected_dhukuti.start_date = start_date
                    selected_dhukuti.received_amount = received_amount
                    selected_dhukuti.total_kista = total_kista
                    selected_dhukuti.received_kista_number = received_kista_number
                    selected_dhukuti.remaining_base_payment = result['remaining_base_payment']
                    selected_dhukuti.kista_increment = kista_increment
                    selected_dhukuti.notes = initial['notes']
                    selected_dhukuti.save()

                    # Replace existing paid-kista rows with updated values from form.
                    selected_dhukuti.paid_kistas.all().delete()
                    for row in result['paid_rows']:
                        DhukutiKistaPayment.objects.create(
                            loan=selected_dhukuti,
                            month_number=row['month'],
                            amount=row['amount'],
                        )
                    # Remove planned rows that are now considered paid from textarea input.
                    selected_dhukuti.planned_kistas.filter(
                        month_number__in=[row['month'] for row in result['paid_rows']]
                    ).delete()

                    messages.success(request, f'Dhukuti loan record "{selected_dhukuti.name}" updated successfully.')
                    dhukuti_loans = DhukutiLoan.objects.prefetch_related('paid_kistas', 'planned_kistas').all()

            elif request.POST.get('save_record') == '1':
                dhukuti_loan = DhukutiLoan.objects.create(
                    name=record_name,
                    start_date=start_date,
                    received_amount=received_amount,
                    total_kista=total_kista,
                    received_kista_number=received_kista_number,
                    remaining_base_payment=result['remaining_base_payment'],
                    kista_increment=kista_increment,
                    notes=initial['notes'],
                )
                for row in result['paid_rows']:
                    DhukutiKistaPayment.objects.create(
                        loan=dhukuti_loan,
                        month_number=row['month'],
                        amount=row['amount'],
                    )

                messages.success(request, f'Dhukuti loan record "{dhukuti_loan.name}" saved successfully.')
                selected_dhukuti = dhukuti_loan
                dhukuti_loans = DhukutiLoan.objects.prefetch_related('paid_kistas', 'planned_kistas').all()
        except Exception as exc:
            messages.error(request, f'Unable to calculate Dhukuti payment: {exc}')

    context = {
        'title': 'Dhukuti Loans',
        'page_title': 'Dhukuti Loans',
        'dhukuti_loans': dhukuti_loans,
        'selected_dhukuti': selected_dhukuti,
        'initial': initial,
        'result': result,
        'dhukuti_total_received': sum((d.received_amount for d in dhukuti_loans), Decimal('0')),
        'dhukuti_total_paid': sum((d.total_paid for d in dhukuti_loans), Decimal('0')),
        'dhukuti_total_remaining': sum((d.estimated_remaining_to_pay for d in dhukuti_loans if d.received_amount > 0), Decimal('0')),
    }
    return render(request, 'finance/loan_emi_calculator.html', context)


@login_required
def loan_dhukuti_delete(request, pk):
    """Delete a saved Dhukuti loan record and its kista rows."""
    dhukuti_loan = get_object_or_404(DhukutiLoan, pk=pk)
    if request.method == 'POST':
        loan_name = dhukuti_loan.name
        dhukuti_loan.delete()
        messages.success(request, f'Dhukuti loan record "{loan_name}" deleted successfully.')
    return redirect('finance:loan_dhukuti_calculator')


@login_required
def loan_dhukuti_kista_mark_paid(request, pk):
    """Mark a specific kista number as paid for a DhukutiLoan."""
    dhukuti_loan = get_object_or_404(DhukutiLoan, pk=pk)
    if request.method == 'POST':
        month_number = int(request.POST.get('month_number', 0))
        amount = request.POST.get('amount', '0').strip() or '0'
        if month_number > 0:
            planned = DhukutiKistaPlan.objects.filter(
                loan=dhukuti_loan, month_number=month_number
            ).first()
            if planned and (not amount or amount == '0'):
                amount = str(planned.amount)
            DhukutiKistaPayment.objects.update_or_create(
                loan=dhukuti_loan,
                month_number=month_number,
                defaults={'amount': Decimal(amount)},
            )
            DhukutiKistaPlan.objects.filter(loan=dhukuti_loan, month_number=month_number).delete()
            messages.success(request, f'Kista {month_number} marked as paid (रु{amount}).') 
    return redirect(reverse('finance:loan_dhukuti_calculator') + f'?dhukuti={pk}')


@login_required
def loan_dhukuti_kista_undo_paid(request, pk):
    """Remove a paid kista entry (undo paid) for a DhukutiLoan."""
    dhukuti_loan = get_object_or_404(DhukutiLoan, pk=pk)
    if request.method == 'POST':
        month_number = int(request.POST.get('month_number', 0))
        if month_number > 0:
            deleted, _ = DhukutiKistaPayment.objects.filter(
                loan=dhukuti_loan, month_number=month_number
            ).delete()
            if deleted:
                messages.success(request, f'Kista {month_number} payment undone.')
    return redirect(reverse('finance:loan_dhukuti_calculator') + f'?dhukuti={pk}')


@login_required
def loan_dhukuti_kista_update_amount(request, pk):
    """Save amount change for a kista without forcing paid status."""
    if request.method != 'POST':
        return redirect(reverse('finance:loan_dhukuti_calculator') + f'?dhukuti={pk}')
    dhukuti_loan = get_object_or_404(DhukutiLoan, pk=pk)
    month_number = int(request.POST.get('month_number', 0))
    amount_str = request.POST.get('amount', '').strip()
    if month_number <= 0 or not amount_str:
        messages.error(request, 'Invalid kista update data.')
        return redirect(reverse('finance:loan_dhukuti_calculator') + f'?dhukuti={pk}')
    try:
        amount = Decimal(amount_str)
        updated_paid = DhukutiKistaPayment.objects.filter(
            loan=dhukuti_loan, month_number=month_number
        ).update(amount=amount)
        if updated_paid:
            messages.success(request, f'Updated paid kista {month_number} amount to रु{amount}.')
        else:
            DhukutiKistaPlan.objects.update_or_create(
                loan=dhukuti_loan,
                month_number=month_number,
                defaults={'amount': amount},
            )
            messages.success(request, f'Saved unpaid kista {month_number} amount रु{amount}.')
        return redirect(reverse('finance:loan_dhukuti_calculator') + f'?dhukuti={pk}')
    except Exception as exc:
        messages.error(request, f'Unable to update kista amount: {exc}')
        return redirect(reverse('finance:loan_dhukuti_calculator') + f'?dhukuti={pk}')


@login_required
def loan_emi_calculator(request):
    """Separate EMI loans calculator and record manager."""
    emi_loans = EmiLoan.objects.all()
    selected_emi = None
    result = None

    initial = {
        'name': '',
        'principal': '',
        'current_principal': '',
        'annual_interest_rate': '',
        'tenure_months': '12',
        'notes': '',
    }

    selected_emi_id = request.GET.get('emi') or request.POST.get('emi_id')
    if selected_emi_id:
        selected_emi = EmiLoan.objects.filter(pk=selected_emi_id).first()
        if selected_emi:
            initial['name'] = selected_emi.name
            initial['principal'] = str(selected_emi.principal)
            initial['current_principal'] = str(selected_emi.current_principal) if selected_emi.current_principal else ''
            initial['annual_interest_rate'] = str(selected_emi.annual_interest_rate)
            initial['tenure_months'] = str(selected_emi.tenure_months)
            initial['notes'] = selected_emi.notes or ''

            principal_for_calc = selected_emi.current_principal if selected_emi.current_principal else selected_emi.principal
            result = _compute_emi_schedule(
                principal=principal_for_calc,
                annual_rate=selected_emi.annual_interest_rate,
                tenure_months=selected_emi.tenure_months,
            )
            result['calculation_principal'] = principal_for_calc

    if request.method == 'POST':
        initial['name'] = request.POST.get('name', '').strip()
        initial['principal'] = request.POST.get('principal', '').strip()
        initial['current_principal'] = request.POST.get('current_principal', '').strip()
        initial['annual_interest_rate'] = request.POST.get('annual_interest_rate', '').strip()
        initial['tenure_months'] = request.POST.get('tenure_months', '').strip()
        initial['notes'] = request.POST.get('notes', '').strip()

        try:
            principal = Decimal(initial['principal'])
            current_principal = Decimal(initial['current_principal']) if initial['current_principal'] else None
            annual_rate = Decimal(initial['annual_interest_rate'])
            tenure_months = int(initial['tenure_months'])

            principal_for_calc = current_principal if current_principal else principal

            result = _compute_emi_schedule(
                principal=principal_for_calc,
                annual_rate=annual_rate,
                tenure_months=tenure_months,
            )
            result['calculation_principal'] = principal_for_calc

            record_name = initial['name'] or f"EMI Loan {principal}"
            if request.POST.get('update_record') == '1':
                if not selected_emi:
                    messages.error(request, 'No EMI record selected to update. Please load a saved record first.')
                else:
                    selected_emi.name = record_name
                    selected_emi.principal = principal
                    selected_emi.current_principal = current_principal
                    selected_emi.annual_interest_rate = annual_rate
                    selected_emi.tenure_months = tenure_months
                    selected_emi.notes = initial['notes']
                    selected_emi.save()
                    messages.success(request, f'EMI loan record "{selected_emi.name}" updated successfully.')
                    emi_loans = EmiLoan.objects.all()
            elif request.POST.get('save_record') == '1':
                selected_emi = EmiLoan.objects.create(
                    name=record_name,
                    principal=principal,
                    current_principal=current_principal,
                    annual_interest_rate=annual_rate,
                    tenure_months=tenure_months,
                    notes=initial['notes'],
                )
                messages.success(request, f'EMI loan record "{selected_emi.name}" saved successfully.')
                emi_loans = EmiLoan.objects.all()
        except Exception as exc:
            messages.error(request, f'Unable to calculate EMI: {exc}')

    context = {
        'title': 'EMI Loans',
        'page_title': 'EMI Loans',
        'emi_loans': emi_loans,
        'selected_emi': selected_emi,
        'initial': initial,
        'result': result,
    }
    return render(request, 'finance/loan_emi_separate.html', context)


@login_required
def loan_emi_delete(request, pk):
    """Delete a saved EMI loan record."""
    emi_loan = get_object_or_404(EmiLoan, pk=pk)
    if request.method == 'POST':
        loan_name = emi_loan.name
        emi_loan.delete()
        messages.success(request, f'EMI loan record "{loan_name}" deleted successfully.')
    return redirect('finance:loan_emi_calculator')
