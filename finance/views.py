from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl import Workbook

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from .forms import ExpenseForm, EmployeeForm, EmployeeSalaryForm, SundryDebtorForm, DebtorTransactionForm, SundryCreditorForm, CreditorTransactionForm
from .models import Expense, Employee, EmployeeSalary, SundryDebtor, DebtorTransaction, SundryCreditor, CreditorTransaction


# FINANCE DASHBOARD
@login_required
def finance_dashboard(request):
    """Finance dashboard with key metrics and summaries"""
    
    # Expenses metrics
    total_expenses = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    expenses_by_category = Expense.objects.values('category').annotate(total=Sum('amount')).order_by('-total')
    recent_expenses = Expense.objects.all().order_by('-expense_date')[:5]
    
    # Employee metrics
    total_employees = Employee.objects.filter(is_active=True).count()
    total_salaries = EmployeeSalary.objects.filter(status='paid').aggregate(Sum('total_salary'))['total_salary__sum'] or 0
    pending_salaries = EmployeeSalary.objects.filter(status='pending').aggregate(Sum('total_salary'))['total_salary__sum'] or 0
    
    # Debtor metrics - calculate from transactions
    active_debtors = SundryDebtor.objects.filter(is_active=True, is_paid=False)
    total_debtor_balance = sum(debtor.get_calculated_balance() for debtor in active_debtors)
    total_debtors = active_debtors.count()
    
    # Creditor metrics
    total_creditor_balance = SundryCreditor.objects.filter(is_active=True).aggregate(Sum('current_balance'))['current_balance__sum'] or 0
    total_creditors = SundryCreditor.objects.filter(is_active=True).count()
    
    context = {
        'total_expenses': total_expenses,
        'expenses_by_category': expenses_by_category,
        'recent_expenses': recent_expenses,
        'total_employees': total_employees,
        'total_salaries': total_salaries,
        'pending_salaries': pending_salaries,
        'total_debtor_balance': total_debtor_balance,
        'total_debtors': total_debtors,
        'total_creditor_balance': total_creditor_balance,
        'total_creditors': total_creditors,
    }
    return render(request, 'finance/dashboard.html', context)


# EXPENSE VIEWS
@login_required
def expense_list(request):
    expenses = Expense.objects.all()
    category = request.GET.get('category')
    if category:
        expenses = expenses.filter(category=category)
    
    total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    
    context = {
        'expenses': expenses,
        'categories': Expense.CATEGORY_CHOICES,
        'total_expenses': total_expenses,
    }
    return render(request, 'finance/expense_list.html', context)


@login_required
def expense_create(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('finance:expense_list')
    else:
        form = ExpenseForm()
    return render(request, 'finance/expense_form.html', {'form': form, 'title': 'Add Expense'})


@login_required
def expense_update(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            return redirect('finance:expense_list')
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'finance/expense_form.html', {'form': form, 'title': 'Edit Expense', 'object': expense})


@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense.delete()
        return redirect('finance:expense_list')
    return render(request, 'finance/expense_confirm_delete.html', {'object': expense})


# EMPLOYEE VIEWS
@login_required
def employee_list(request):
    employees = Employee.objects.all()
    active_only = request.GET.get('active_only')
    if active_only:
        employees = employees.filter(is_active=True)
    
    context = {
        'employees': employees,
    }
    return render(request, 'finance/employee_list.html', context)


@login_required
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('finance:employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'finance/employee_form.html', {'form': form, 'title': 'Add Employee'})


@login_required
def employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('finance:employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'finance/employee_form.html', {'form': form, 'title': 'Edit Employee', 'object': employee})


@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('finance:employee_list')
    return render(request, 'finance/employee_confirm_delete.html', {'object': employee})


# EMPLOYEE SALARY VIEWS
@login_required
def salary_list(request):
    salaries = EmployeeSalary.objects.select_related('employee').all()
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')
    
    if employee_id:
        salaries = salaries.filter(employee_id=employee_id)
    if status:
        salaries = salaries.filter(status=status)
    
    total_salary = salaries.aggregate(Sum('total_salary'))['total_salary__sum'] or 0
    total_paid = salaries.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    
    context = {
        'salaries': salaries,
        'employees': Employee.objects.filter(is_active=True),
        'statuses': EmployeeSalary.STATUS_CHOICES,
        'total_salary': total_salary,
        'total_paid': total_paid,
    }
    return render(request, 'finance/salary_list.html', context)


@login_required
def salary_create(request):
    if request.method == 'POST':
        form = EmployeeSalaryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('finance:salary_list')
    else:
        form = EmployeeSalaryForm()
    return render(request, 'finance/salary_form.html', {'form': form, 'title': 'Add Salary'})


@login_required
def salary_update(request, pk):
    salary = get_object_or_404(EmployeeSalary, pk=pk)
    if request.method == 'POST':
        form = EmployeeSalaryForm(request.POST, instance=salary)
        if form.is_valid():
            form.save()
            return redirect('finance:salary_list')
    else:
        form = EmployeeSalaryForm(instance=salary)
    return render(request, 'finance/salary_form.html', {'form': form, 'title': 'Edit Salary', 'object': salary})


@login_required
def salary_delete(request, pk):
    salary = get_object_or_404(EmployeeSalary, pk=pk)
    if request.method == 'POST':
        salary.delete()
        return redirect('finance:salary_list')
    return render(request, 'finance/salary_confirm_delete.html', {'object': salary})


# SUNDRY DEBTOR VIEWS
@login_required
def debtor_list(request):
    debtors = SundryDebtor.objects.all()
    active_only = request.GET.get('active_only')
    sort_by = request.GET.get('sort_by', 'latest')  # Default sort by latest BS date
    
    if active_only:
        debtors = debtors.filter(is_active=True)
    
    # Recalculate all debtor balances from transactions
    for debtor in debtors:
        debtor.update_balance_from_transactions()
    
    # Apply sorting
    if sort_by == 'created_first':
        debtors = debtors.order_by('created_at')
    elif sort_by == 'created_recent':
        debtors = debtors.order_by('-created_at')
    elif sort_by == 'balance':
        debtors = debtors.order_by('-current_balance')
    elif sort_by == 'latest':
        debtors = debtors.order_by('-bs_date', '-created_at')
    else:  # Default: sort by name
        debtors = debtors.order_by('name')
    
    total_balance = sum(debtor.get_calculated_balance() for debtor in debtors)
    
    context = {
        'debtors': debtors,
        'total_balance': total_balance,
        'sort_by': sort_by,
    }
    return render(request, 'finance/debtor_list.html', context)


@login_required
def debtor_create(request):
    if request.method == 'POST':
        form = SundryDebtorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('finance:debtor_list')
    else:
        form = SundryDebtorForm()
    return render(request, 'finance/debtor_form.html', {'form': form, 'title': 'Add Debtor'})


@login_required
def debtor_update(request, pk):
    debtor = get_object_or_404(SundryDebtor, pk=pk)
    if request.method == 'POST':
        form = SundryDebtorForm(request.POST, instance=debtor)
        if form.is_valid():
            form.save()
            return redirect('finance:debtor_list')
    else:
        form = SundryDebtorForm(instance=debtor)
    return render(request, 'finance/debtor_form.html', {'form': form, 'title': 'Edit Debtor', 'object': debtor})


@login_required
def debtor_delete(request, pk):
    debtor = get_object_or_404(SundryDebtor, pk=pk)
    if request.method == 'POST':
        debtor.delete()
        return redirect('finance:debtor_list')
    return render(request, 'finance/debtor_confirm_delete.html', {'object': debtor})


@login_required
def debtor_mark_paid(request, pk):
    debtor = get_object_or_404(SundryDebtor, pk=pk)
    if request.method == 'POST':
        debtor.is_paid = True
        debtor.save(update_fields=['is_paid'])
    return redirect('finance:debtor_list')


@login_required
def debtor_mark_unpaid(request, pk):
    debtor = get_object_or_404(SundryDebtor, pk=pk)
    if request.method == 'POST':
        debtor.is_paid = False
        debtor.save(update_fields=['is_paid'])
    return redirect('finance:debtor_list')


@login_required
def debtor_detail(request, pk):
    debtor = get_object_or_404(SundryDebtor, pk=pk)
    transactions = debtor.transactions.all()
    
    # Update balance from transactions
    debtor.update_balance_from_transactions()
    calculated_balance = debtor.get_calculated_balance()
    
    context = {
        'debtor': debtor,
        'transactions': transactions,
        'calculated_balance': calculated_balance,
    }
    return render(request, 'finance/debtor_detail.html', context)


@login_required
def debtor_transaction_create(request, debtor_id):
    debtor = get_object_or_404(SundryDebtor, pk=debtor_id)
    if request.method == 'POST':
        form = DebtorTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.debtor = debtor
            transaction.save()
            return redirect('finance:debtor_detail', pk=debtor.pk)
    else:
        form = DebtorTransactionForm(initial={'debtor': debtor})
    return render(request, 'finance/debtor_transaction_form.html', {'form': form, 'debtor': debtor, 'title': 'Add Transaction'})


@login_required
def debtor_transaction_update(request, pk):
    transaction = get_object_or_404(DebtorTransaction, pk=pk)
    debtor = transaction.debtor
    if request.method == 'POST':
        form = DebtorTransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            return redirect('finance:debtor_detail', pk=debtor.pk)
    else:
        form = DebtorTransactionForm(instance=transaction)
    return render(request, 'finance/debtor_transaction_form.html', {'form': form, 'debtor': debtor, 'title': 'Edit Transaction', 'object': transaction})


@login_required
def debtor_transaction_delete(request, pk):
    transaction = get_object_or_404(DebtorTransaction, pk=pk)
    debtor = transaction.debtor
    if request.method == 'POST':
        transaction.delete()
        debtor.update_balance_from_transactions()
        return redirect('finance:debtor_detail', pk=debtor.pk)
    return render(request, 'finance/debtor_transaction_confirm_delete.html', {'object': transaction, 'debtor': debtor})


# SUNDRY CREDITOR VIEWS
@login_required
def creditor_list(request):
    creditors = SundryCreditor.objects.all()
    active_only = request.GET.get('active_only')
    if active_only:
        creditors = creditors.filter(is_active=True)

    creditors = creditors.order_by('-bs_date', '-created_at')
    
    total_balance = creditors.aggregate(Sum('current_balance'))['current_balance__sum'] or 0
    
    context = {
        'creditors': creditors,
        'total_balance': total_balance,
    }
    return render(request, 'finance/creditor_list.html', context)


@login_required
def creditor_create(request):
    if request.method == 'POST':
        form = SundryCreditorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('finance:creditor_list')
    else:
        form = SundryCreditorForm()
    return render(request, 'finance/creditor_form.html', {'form': form, 'title': 'Add Creditor'})


@login_required
def creditor_update(request, pk):
    creditor = get_object_or_404(SundryCreditor, pk=pk)
    if request.method == 'POST':
        form = SundryCreditorForm(request.POST, instance=creditor)
        if form.is_valid():
            form.save()
            return redirect('finance:creditor_list')
    else:
        form = SundryCreditorForm(instance=creditor)
    return render(request, 'finance/creditor_form.html', {'form': form, 'title': 'Edit Creditor', 'object': creditor})


@login_required
def creditor_delete(request, pk):
    creditor = get_object_or_404(SundryCreditor, pk=pk)
    if request.method == 'POST':
        creditor.delete()
        return redirect('finance:creditor_list')
    return render(request, 'finance/creditor_confirm_delete.html', {'object': creditor})


@login_required
def creditor_mark_paid(request, pk):
    creditor = get_object_or_404(SundryCreditor, pk=pk)
    if request.method == 'POST':
        creditor.is_paid = True
        creditor.save(update_fields=['is_paid'])
    return redirect('finance:creditor_list')


@login_required
def creditor_mark_unpaid(request, pk):
    creditor = get_object_or_404(SundryCreditor, pk=pk)
    if request.method == 'POST':
        creditor.is_paid = False
        creditor.save(update_fields=['is_paid'])
    return redirect('finance:creditor_list')


@login_required
def creditor_detail(request, pk):
    creditor = get_object_or_404(SundryCreditor, pk=pk)
    transactions = creditor.transactions.all()
    
    creditor.update_balance_from_transactions()
    calculated_balance = creditor.get_calculated_balance()

    context = {
        'creditor': creditor,
        'transactions': transactions,
        'calculated_balance': calculated_balance,
    }
    return render(request, 'finance/creditor_detail.html', context)


@login_required
def creditor_transaction_create(request, creditor_id):
    creditor = get_object_or_404(SundryCreditor, pk=creditor_id)
    if request.method == 'POST':
        form = CreditorTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.creditor = creditor
            transaction.save()
            return redirect('finance:creditor_detail', pk=creditor.pk)
    else:
        form = CreditorTransactionForm(initial={'creditor': creditor})
    return render(request, 'finance/creditor_transaction_form.html', {'form': form, 'creditor': creditor, 'title': 'Add Transaction'})


@login_required
def creditor_transaction_update(request, pk):
    transaction = get_object_or_404(CreditorTransaction, pk=pk)
    creditor = transaction.creditor
    if request.method == 'POST':
        form = CreditorTransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            return redirect('finance:creditor_detail', pk=creditor.pk)
    else:
        form = CreditorTransactionForm(instance=transaction)
    return render(request, 'finance/creditor_transaction_form.html', {'form': form, 'creditor': creditor, 'title': 'Edit Transaction', 'object': transaction})


@login_required
def creditor_transaction_delete(request, pk):
    transaction = get_object_or_404(CreditorTransaction, pk=pk)
    creditor = transaction.creditor
    if request.method == 'POST':
        transaction.delete()
        creditor.update_balance_from_transactions()
        return redirect('finance:creditor_detail', pk=creditor.pk)
    return render(request, 'finance/creditor_transaction_confirm_delete.html', {'object': transaction, 'creditor': creditor})


# IMPORT/EXPORT HELPERS
def _safe_decimal(val):
    try:
        return Decimal(str(val)) if val not in (None, '') else Decimal('0')
    except Exception:
        return Decimal('0')


def _extract_date(val):
    """Extract date in YYYY-MM-DD format from various input types."""
    if not val:
        return None
    try:
        # Convert to string and extract only date part (YYYY-MM-DD)
        date_str = str(val)
        # If it has time component, split it
        if ' ' in date_str:
            date_str = date_str.split(' ')[0]
        return date_str
    except Exception:
        return None


# EXPENSE IMPORT/EXPORT
@login_required
def expense_export(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    ws.append(["category", "description", "amount", "expense_date", "notes"])
    for e in Expense.objects.all().order_by('-expense_date'):
        ws.append([
            e.category,
            e.description,
            float(e.amount),
            str(e.expense_date) if e.expense_date else '',
            e.notes or "",
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'
    return response


@login_required
def expense_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                category, description, amount, expense_date, notes = row[:5]
                expense_date = _extract_date(expense_date)
                if not expense_date:
                    continue
                
                Expense.objects.create(
                    category=category or 'other',
                    description=description or '',
                    amount=_safe_decimal(amount),
                    expense_date=expense_date,
                    notes=notes or '',
                )
                created += 1
            messages.success(request, f"Imported {created} expenses")
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
    return redirect('finance:expense_list')


# EMPLOYEE IMPORT/EXPORT
@login_required
def employee_export(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    ws.append(["first_name", "last_name", "email", "phone", "position", "base_salary", "hire_date", "is_active"])
    for emp in Employee.objects.all().order_by('first_name', 'last_name'):
        ws.append([
            emp.first_name,
            emp.last_name,
            emp.email or '',
            emp.phone or '',
            emp.position,
            float(emp.base_salary),
            str(emp.hire_date) if emp.hire_date else '',
            'true' if emp.is_active else 'false',
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employees.xlsx"'
    return response


@login_required
def employee_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                first_name, last_name, email, phone, position, base_salary, hire_date, is_active = row[:8]
                
                is_active_bool = str(is_active).lower() in ['1', 'true', 'yes', 'y']
                emp, _created = Employee.objects.update_or_create(
                    email=email or None,
                    defaults={
                        'first_name': first_name or '',
                        'last_name': last_name or '',
                        'phone': phone or '',
                        'position': position or '',
                        'base_salary': _safe_decimal(base_salary),
                        'hire_date': _extract_date(hire_date),
                        'is_active': is_active_bool,
                    }
                )
                if _created:
                    created += 1
            messages.success(request, f"Imported {created} employees (existing emails updated)")
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
    return redirect('finance:employee_list')


# SALARY IMPORT/EXPORT
@login_required
def salary_export(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salaries"
    
    ws.append([
        "employee_email", "employee_name", "month", "base_salary", "bonus", "deductions",
        "total_salary", "amount_paid", "status", "paid_date", "notes"
    ])
    for s in EmployeeSalary.objects.select_related('employee').all().order_by('-month'):
        ws.append([
            s.employee.email or '',
            f"{s.employee.first_name} {s.employee.last_name}",
            str(s.month) if s.month else '',
            float(s.base_salary),
            float(s.bonus),
            float(s.deductions),
            float(s.total_salary),
            float(s.amount_paid),
            s.status,
            str(s.paid_date) if s.paid_date else '',
            s.notes or '',
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="salaries.xlsx"'
    return response


@login_required
def salary_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            skipped = 0
            errors = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                try:
                    employee_email, employee_name, month, base_salary, bonus, deductions, total_salary, amount_paid, status, paid_date, notes = row[:11]
                    
                    # Try to find employee by email first
                    employee = None
                    email_str = str(employee_email).strip() if employee_email else ''
                    name_str = str(employee_name).strip() if employee_name else ''
                    
                    if email_str:
                        try:
                            employee = Employee.objects.get(email__iexact=email_str)
                        except Employee.DoesNotExist:
                            pass
                    
                    # If not found by email, try by name
                    if not employee and name_str:
                        name_parts = name_str.split()
                        if len(name_parts) >= 2:
                            first_name = name_parts[0]
                            last_name = ' '.join(name_parts[1:])
                            try:
                                employee = Employee.objects.get(first_name__iexact=first_name, last_name__iexact=last_name)
                            except Employee.DoesNotExist:
                                pass
                    
                    # If still not found, skip this row with error message
                    if not employee:
                        error_msg = f"Row {row_idx}: Employee not found (email: {email_str}, name: {name_str})"
                        errors.append(error_msg)
                        skipped += 1
                        continue

                    month_str = _extract_date(month)
                    if not month_str:
                        error_msg = f"Row {row_idx}: Invalid month date format"
                        errors.append(error_msg)
                        skipped += 1
                        continue

                    # Calculate total_salary if not provided
                    base_sal = _safe_decimal(base_salary)
                    bonus_sal = _safe_decimal(bonus)
                    deductions_sal = _safe_decimal(deductions)
                    calc_total = base_sal + bonus_sal - deductions_sal

                    _, created_flag = EmployeeSalary.objects.update_or_create(
                        employee=employee,
                        month=month_str,
                        defaults={
                            'base_salary': base_sal,
                            'bonus': bonus_sal,
                            'deductions': deductions_sal,
                            'total_salary': calc_total,
                            'amount_paid': _safe_decimal(amount_paid),
                            'status': status or 'pending',
                            'paid_date': _extract_date(paid_date),
                            'notes': notes or '',
                        }
                    )
                    if created_flag:
                        created += 1
                except Exception as e:
                    error_msg = f"Row {row_idx} error: {str(e)}"
                    errors.append(error_msg)
                    skipped += 1
                    continue
            
            # Build success message
            message = f"Imported {created} salary rows"
            if skipped > 0:
                message += f" ({skipped} rows skipped)"
            
            # Add up to 5 error details
            for error in errors[:5]:
                messages.warning(request, error)
            
            if skipped > 5:
                messages.warning(request, f"... and {skipped - 5} more rows skipped")
            
            messages.success(request, message)
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
    return redirect('finance:salary_list')


# DEBTOR IMPORT/EXPORT
@login_required
def debtor_export(request):
    wb = Workbook()
    
    # Debtors Sheet
    ws_debtors = wb.active
    ws_debtors.title = "Debtors"
    ws_debtors.append(["name", "contact_person", "phone", "address", "bs_date", "opening_balance", "current_balance", "credit_limit", "is_active", "notes"])
    for d in SundryDebtor.objects.all().order_by('name'):
        ws_debtors.append([
            d.name,
            d.contact_person or '',
            d.phone or '',
            d.address or '',
            str(d.bs_date) if d.bs_date else '',
            float(d.opening_balance),
            float(d.current_balance),
            float(d.credit_limit),
            'true' if d.is_active else 'false',
            d.notes or '',
        ])
    
    # Debtor Transactions Sheet
    ws_transactions = wb.create_sheet("Debtor Transactions")
    ws_transactions.append(["debtor_name", "transaction_type", "reference_no", "amount", "transaction_date", "due_date", "description"])
    for transaction in DebtorTransaction.objects.select_related('debtor').all().order_by('debtor__name', '-transaction_date'):
        ws_transactions.append([
            transaction.debtor.name,
            transaction.transaction_type,
            transaction.reference_no or '',
            float(transaction.amount),
            str(transaction.transaction_date) if transaction.transaction_date else '',
            str(transaction.due_date) if transaction.due_date else '',
            transaction.description or '',
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="debtors_with_transactions.xlsx"'
    return response


@login_required
def debtor_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            created_debtors = 0
            created_transactions = 0
            
            # Import Debtors from first sheet
            if 'Debtors' in wb.sheetnames:
                ws_debtors = wb['Debtors']
                for row in ws_debtors.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    name, contact_person, phone, address, bs_date, opening_balance, current_balance, credit_limit, is_active, notes = row[:10]
                    
                    is_active_bool = str(is_active).lower() in ['1', 'true', 'yes', 'y']
                    debtor, created_flag = SundryDebtor.objects.update_or_create(
                        name=name or '',
                        defaults={
                            'contact_person': contact_person or '',
                            'phone': phone or '',
                            'address': address or '',
                            'bs_date': _extract_date(bs_date),
                            'opening_balance': _safe_decimal(opening_balance),
                            'current_balance': _safe_decimal(current_balance),
                            'credit_limit': _safe_decimal(credit_limit),
                            'is_active': is_active_bool,
                            'notes': notes or '',
                        }
                    )
                    if created_flag:
                        created_debtors += 1
            
            # Import Debtor Transactions from second sheet
            if 'Debtor Transactions' in wb.sheetnames:
                ws_transactions = wb['Debtor Transactions']
                for row in ws_transactions.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    debtor_name, transaction_type, reference_no, amount, transaction_date, due_date, description = row[:7]
                    
                    # Find the debtor by name
                    try:
                        debtor = SundryDebtor.objects.get(name=debtor_name)
                        transaction_date_str = _extract_date(transaction_date)
                        if not transaction_date_str:
                            continue
                        
                        # Create transaction
                        DebtorTransaction.objects.create(
                            debtor=debtor,
                            transaction_type=transaction_type or 'invoice',
                            reference_no=reference_no or '',
                            amount=_safe_decimal(amount),
                            transaction_date=transaction_date_str,
                            due_date=_extract_date(due_date),
                            description=description or '',
                        )
                        created_transactions += 1
                    except SundryDebtor.DoesNotExist:
                        continue
            
            messages.success(request, f"Imported {created_debtors} debtors and {created_transactions} transactions")
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
    return redirect('finance:debtor_list')


# CREDITOR IMPORT/EXPORT
@login_required
def creditor_export(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Creditors"
    
    ws.append(["name", "contact_person", "phone", "address", "bs_date", "opening_balance", "current_balance", "is_active", "notes"])
    for c in SundryCreditor.objects.all().order_by('name'):
        ws.append([
            c.name,
            c.contact_person or '',
            c.phone or '',
            c.address or '',
            str(c.bs_date) if c.bs_date else '',
            float(c.opening_balance),
            float(c.current_balance),
            'true' if c.is_active else 'false',
            c.notes or '',
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="creditors.xlsx"'
    return response


@login_required
def creditor_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                name, contact_person, phone, address, bs_date, opening_balance, current_balance, is_active, notes = row[:9]
                
                is_active_bool = str(is_active).lower() in ['1', 'true', 'yes', 'y']
                _, created_flag = SundryCreditor.objects.update_or_create(
                    name=name or '',
                    defaults={
                        'contact_person': contact_person or '',
                        'phone': phone or '',
                        'address': address or '',
                        'bs_date': _extract_date(bs_date),
                        'opening_balance': _safe_decimal(opening_balance),
                        'current_balance': _safe_decimal(current_balance),
                        'is_active': is_active_bool,
                        'notes': notes or '',
                    }
                )
                if created_flag:
                    created += 1
            messages.success(request, f"Imported {created} creditors (matching by name)")
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
    return redirect('finance:creditor_list')


# BULK FINANCE EXPORT (Single XLSX with multiple sheets)
@login_required
def finance_export_all(request):
    """Export all finance data as a single Excel file with multiple sheets."""
    wb = Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Expenses Sheet
    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses.append(["category", "description", "amount", "expense_date", "notes"])
    for e in Expense.objects.all().order_by('-expense_date'):
        ws_expenses.append([
            e.category,
            e.description,
            float(e.amount),
            str(e.expense_date) if e.expense_date else '',
            e.notes or "",
        ])
    
    # Employees Sheet
    ws_employees = wb.create_sheet("Employees")
    ws_employees.append(["first_name", "last_name", "email", "phone", "position", "base_salary", "hire_date", "is_active"])
    for emp in Employee.objects.all().order_by('first_name', 'last_name'):
        ws_employees.append([
            emp.first_name,
            emp.last_name,
            emp.email or '',
            emp.phone or '',
            emp.position,
            float(emp.base_salary),
            str(emp.hire_date) if emp.hire_date else '',
            'true' if emp.is_active else 'false',
        ])
    
    # Salaries Sheet
    ws_salaries = wb.create_sheet("Salaries")
    ws_salaries.append([
        "employee_email", "employee_name", "month", "base_salary", "bonus", "deductions",
        "total_salary", "amount_paid", "status", "paid_date", "notes"
    ])
    for s in EmployeeSalary.objects.select_related('employee').all().order_by('-month'):
        ws_salaries.append([
            s.employee.email or '',
            f"{s.employee.first_name} {s.employee.last_name}",
            str(s.month) if s.month else '',
            float(s.base_salary),
            float(s.bonus),
            float(s.deductions),
            float(s.total_salary),
            float(s.amount_paid),
            s.status,
            str(s.paid_date) if s.paid_date else '',
            s.notes or '',
        ])
    
    # Debtors Sheet
    ws_debtors = wb.create_sheet("Debtors")
    ws_debtors.append(["name", "contact_person", "phone", "address", "opening_balance", "current_balance", "credit_limit", "is_active", "notes"])
    for d in SundryDebtor.objects.all().order_by('name'):
        ws_debtors.append([
            d.name,
            d.contact_person or '',
            d.phone or '',
            d.address or '',
            float(d.opening_balance),
            float(d.current_balance),
            float(d.credit_limit),
            'true' if d.is_active else 'false',
            d.notes or '',
        ])
    
    # Creditors Sheet
    ws_creditors = wb.create_sheet("Creditors")
    ws_creditors.append(["name", "contact_person", "phone", "address", "opening_balance", "current_balance", "is_active", "notes"])
    for c in SundryCreditor.objects.all().order_by('name'):
        ws_creditors.append([
            c.name,
            c.contact_person or '',
            c.phone or '',
            c.address or '',
            float(c.opening_balance),
            float(c.current_balance),
            'true' if c.is_active else 'false',
            c.notes or '',
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=\"finance_data.xlsx\"'
    return response


# BULK FINANCE IMPORT (Single XLSX with multiple sheets)
@login_required
def finance_import_all(request):
    """Import all finance data from a single Excel file with multiple sheets."""
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        stats = {'expenses': 0, 'employees': 0, 'salaries': 0, 'debtors': 0, 'creditors': 0}
        
        try:
            wb = openpyxl.load_workbook(file)
            
            # Import Expenses
            if 'Expenses' in wb.sheetnames:
                ws = wb['Expenses']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    category, description, amount, expense_date, notes = row[:5]
                    expense_date = _extract_date(expense_date)
                    if expense_date:
                        Expense.objects.create(
                            category=category or 'other',
                            description=description or '',
                            amount=_safe_decimal(amount),
                            expense_date=expense_date,
                            notes=notes or '',
                        )
                        stats['expenses'] += 1
            
            # Import Employees
            if 'Employees' in wb.sheetnames:
                ws = wb['Employees']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    first_name, last_name, email, phone, position, base_salary, hire_date, is_active = row[:8]
                    is_active_bool = str(is_active).lower() in ['1', 'true', 'yes', 'y']
                    emp, created = Employee.objects.update_or_create(
                        email=email or None,
                        defaults={
                            'first_name': first_name or '',
                            'last_name': last_name or '',
                            'phone': phone or '',
                            'position': position or '',
                            'base_salary': _safe_decimal(base_salary),
                            'hire_date': _extract_date(hire_date),
                            'is_active': is_active_bool,
                        }
                    )
                    if created:
                        stats['employees'] += 1
            
            # Import Salaries
            if 'Salaries' in wb.sheetnames:
                ws = wb['Salaries']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    employee_email, employee_name, month, base_salary, bonus, deductions, total_salary, amount_paid, status, paid_date, notes = row[:11]
                    
                    try:
                        employee = Employee.objects.get(email=employee_email)
                    except Employee.DoesNotExist:
                        continue
                    
                    month = _extract_date(month)
                    if not month:
                        continue
                    
                    _, created = EmployeeSalary.objects.update_or_create(
                        employee=employee,
                        month=month,
                        defaults={
                            'base_salary': _safe_decimal(base_salary),
                            'bonus': _safe_decimal(bonus),
                            'deductions': _safe_decimal(deductions),
                            'amount_paid': _safe_decimal(amount_paid),
                            'status': status or 'pending',
                            'paid_date': _extract_date(paid_date),
                            'notes': notes or '',
                        }
                    )
                    if created:
                        stats['salaries'] += 1
            
            # Import Debtors
            if 'Debtors' in wb.sheetnames:
                ws = wb['Debtors']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    name, contact_person, phone, email, address, opening_balance, current_balance, credit_limit, is_active, notes = row[:10]
                    is_active_bool = str(is_active).lower() in ['1', 'true', 'yes', 'y']
                    _, created = SundryDebtor.objects.update_or_create(
                        name=name or '',
                        defaults={
                            'contact_person': contact_person or '',
                            'phone': phone or '',
                            'email': email or '',
                            'address': address or '',
                            'opening_balance': _safe_decimal(opening_balance),
                            'current_balance': _safe_decimal(current_balance),
                            'credit_limit': _safe_decimal(credit_limit),
                            'is_active': is_active_bool,
                            'notes': notes or '',
                        }
                    )
                    if created:
                        stats['debtors'] += 1
            
            # Import Creditors
            if 'Creditors' in wb.sheetnames:
                ws = wb['Creditors']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    name, contact_person, phone, email, address, opening_balance, current_balance, is_active, notes = row[:9]
                    is_active_bool = str(is_active).lower() in ['1', 'true', 'yes', 'y']
                    _, created = SundryCreditor.objects.update_or_create(
                        name=name or '',
                        defaults={
                            'contact_person': contact_person or '',
                            'phone': phone or '',
                            'email': email or '',
                            'address': address or '',
                            'opening_balance': _safe_decimal(opening_balance),
                            'current_balance': _safe_decimal(current_balance),
                            'is_active': is_active_bool,
                            'notes': notes or '',
                        }
                    )
                    if created:
                        stats['creditors'] += 1
            
            msg = f"Finance import: {stats['expenses']} expenses, {stats['employees']} employees, {stats['salaries']} salaries, {stats['debtors']} debtors, {stats['creditors']} creditors"
            messages.success(request, msg)
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
    
    return redirect('gsp:data_settings')


# API ENDPOINTS
@login_required
def debtors_list_api(request):
    """API endpoint to fetch all debtors for payment selection (both active and inactive)"""
    debtors = SundryDebtor.objects.all().values('id', 'name', 'contact_person', 'phone', 'email', 'current_balance', 'is_active').order_by('name')
    return JsonResponse(list(debtors), safe=False)
