from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from .models import CashBank
from .forms import CashBankForm, OtherInvestmentForm


@login_required
def cashbank_list(request):
    """List cash and bank accounts"""
    cash_accounts = CashBank.objects.filter(account_type='cash').order_by('account_name')
    bank_accounts = CashBank.objects.filter(account_type='bank').order_by('bank_name', 'account_name')
    
    total_cash = cash_accounts.aggregate(
        total=Coalesce(Sum('balance'), Decimal('0'))
    )['total']
    
    total_bank = bank_accounts.aggregate(
        total=Coalesce(Sum('balance'), Decimal('0'))
    )['total']

    total_all = total_cash + total_bank
    
    context = {
        'cash_accounts': cash_accounts,
        'bank_accounts': bank_accounts,
        'total_cash': total_cash,
        'total_bank': total_bank,
        'total_all': total_all,
    }
    return render(request, 'finance/cashbank_list.html', context)


@login_required
def other_investment_list(request):
    """List other investment accounts"""
    accounts = CashBank.objects.filter(account_type='other_investment').order_by('account_name')
    
    total_invested = Decimal('0')
    total_current = Decimal('0')
    for acc in accounts:
        if acc.investment_amount:
            total_invested += acc.investment_amount
        if acc.current_amount:
            total_current += acc.current_amount
    
    total_profit_loss = total_current - total_invested
    
    context = {
        'accounts': accounts,
        'total_invested': total_invested,
        'total_current': total_current,
        'total_profit_loss': total_profit_loss,
    }
    return render(request, 'finance/other_investment_list.html', context)


@login_required
def other_investment_create(request):
    """Create a new other investment"""
    if request.method == 'POST':
        form = OtherInvestmentForm(request.POST)
        if form.is_valid():
            inv = form.save()
            messages.success(request, f"Investment '{inv.account_name}' added successfully!")
            return redirect('finance:other_investment_list')
    else:
        form = OtherInvestmentForm()
    return render(request, 'finance/other_investment_form.html', {'form': form, 'action': 'Add'})


@login_required
def other_investment_update(request, pk):
    """Edit an existing investment"""
    account = get_object_or_404(CashBank, pk=pk, account_type='other_investment')
    if request.method == 'POST':
        form = OtherInvestmentForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, f"Investment '{account.account_name}' updated successfully!")
            return redirect('finance:other_investment_list')
    else:
        form = OtherInvestmentForm(instance=account)
    return render(request, 'finance/other_investment_form.html', {'form': form, 'action': 'Edit', 'account': account})


@login_required
def investment_export_excel(request):
    """Export all other investment accounts to XLSX"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Other Investments"

    ws.append(['account_name', 'investment_date', 'investment_amount', 'current_amount', 'notes', 'is_active'])
    for inv in CashBank.objects.filter(account_type='other_investment').order_by('account_name'):
        ws.append([
            inv.account_name,
            str(inv.investment_date) if inv.investment_date else '',
            float(inv.investment_amount) if inv.investment_amount is not None else '',
            float(inv.current_amount) if inv.current_amount is not None else '',
            inv.notes or '',
            'Yes' if inv.is_active else 'No',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="other_investments.xlsx"'
    return response


@login_required
def investment_import_excel(request):
    """Import other investment accounts from XLSX"""
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            updated = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                account_name = row[0] if len(row) > 0 else None
                investment_date = str(row[1]) if len(row) > 1 and row[1] else None
                investment_amount = row[2] if len(row) > 2 else None
                current_amount = row[3] if len(row) > 3 else None
                notes = row[4] if len(row) > 4 else ''
                is_active_raw = row[5] if len(row) > 5 else 'Yes'

                if not account_name:
                    continue

                is_active = str(is_active_raw).strip().lower() not in ('no', 'false', '0')

                def safe_dec(val):
                    try:
                        return Decimal(str(val)) if val not in (None, '') else None
                    except Exception:
                        return None

                inv_amount = safe_dec(investment_amount)
                cur_amount = safe_dec(current_amount)
                balance = cur_amount if cur_amount is not None else (inv_amount or Decimal('0'))

                inv, created_flag = CashBank.objects.update_or_create(
                    account_name=account_name,
                    account_type='other_investment',
                    defaults={
                        'investment_date': investment_date or None,
                        'investment_amount': inv_amount,
                        'current_amount': cur_amount,
                        'balance': balance,
                        'notes': notes or '',
                        'is_active': is_active,
                    }
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1

            messages.success(request, f'Imported {created} new investments, updated {updated} existing.')
        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')

    return redirect('finance:other_investment_list')


@login_required
def cashbank_create(request):
    """Create a new cash or bank account"""
    initial_type = request.GET.get('type', '')
    next_url = request.GET.get('next', '')
    
    if request.method == 'POST':
        form = CashBankForm(request.POST)
        if form.is_valid():
            cashbank = form.save()
            messages.success(request, f"Account '{cashbank.account_name}' created successfully!")
            redirect_to = request.POST.get('next', '')
            if redirect_to == 'other_investment_list':
                return redirect('finance:other_investment_list')
            return redirect('finance:cashbank_list')
    else:
        initial = {}
        if initial_type:
            initial['account_type'] = initial_type
        form = CashBankForm(initial=initial)
    
    context = {'form': form, 'action': 'Create', 'next': next_url}
    return render(request, 'finance/cashbank_form.html', context)


@login_required
def cashbank_update(request, pk):
    """Update an existing cash or bank account"""
    cashbank = get_object_or_404(CashBank, pk=pk)
    
    if request.method == 'POST':
        form = CashBankForm(request.POST, instance=cashbank)
        if form.is_valid():
            cashbank = form.save()
            messages.success(request, f"Account '{cashbank.account_name}' updated successfully!")
            return redirect('finance:cashbank_list')
    else:
        form = CashBankForm(instance=cashbank)
    
    context = {'form': form, 'cashbank': cashbank, 'action': 'Update'}
    return render(request, 'finance/cashbank_form.html', context)


@login_required
def cashbank_delete(request, pk):
    """Delete a cash or bank account"""
    cashbank = get_object_or_404(CashBank, pk=pk)
    
    if request.method == 'POST':
        account_name = cashbank.account_name
        cashbank.delete()
        messages.success(request, f"Account '{account_name}' deleted successfully!")
        return redirect('finance:cashbank_list')
    
    context = {'cashbank': cashbank}
    return render(request, 'finance/cashbank_confirm_delete.html', context)


@login_required
def cashbank_toggle_active(request, pk):
    """Toggle active status of an account"""
    cashbank = get_object_or_404(CashBank, pk=pk)
    cashbank.is_active = not cashbank.is_active
    cashbank.save()
    
    status = "activated" if cashbank.is_active else "deactivated"
    messages.success(request, f"Account '{cashbank.account_name}' {status}!")
    return redirect('finance:cashbank_list')
