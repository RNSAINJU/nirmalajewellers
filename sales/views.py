from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, UpdateView, CreateView, DeleteView, FormView, TemplateView
from decimal import Decimal
from datetime import timedelta
from io import BytesIO
import json
import nepali_datetime as ndt
from django.http import HttpResponse
from django.contrib import messages
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Sum, Q
from django.core.files.storage import default_storage
from django.conf import settings

from order.models import Order, OrderOrnament, OrderPayment, DebtorPayment
from order.forms import OrderForm, OrnamentFormSet
from ornament.models import Ornament, Kaligar
from .models import Sale
from .forms import ExcelImportForm
from finance.models import SundryDebtor


class CreateSaleFromOrderView(View):
    """Create a Sale entry from an existing Order.

    This is triggered from the order list via a button. If a Sale already
    exists for the given order, it simply redirects without creating a
    duplicate.
    """

    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)

        # If sale already exists for this order, do nothing
        if hasattr(order, "sale") and order.sale is not None:
            return redirect("order:list")

        # Prefer deliver_date, then order_date, otherwise today's Nepali date
        sale_date = order.deliver_date or order.order_date
        if sale_date is None:
            try:
                sale_date = ndt.date.today()
            except Exception:
                sale_date = None

        # Default bill number for the sale (can be edited later)
        bill_no = str(order.sn)

        # Create the Sale record
        Sale.objects.create(
            order=order,
            sale_date=sale_date,
            bill_no=bill_no,
            pan_number=order.pan_number,
            address=order.address,
        )

        # Mark ornaments on this order as sales items
        Ornament.objects.filter(order=order).update(
            ornament_type=Ornament.OrnamentCategory.SALES
        )

        # Mark the order as delivered once it is added to sales
        order.status = "delivered"
        order.save(update_fields=["status", "updated_at"])

        return redirect("order:list")


class SalesListView(ListView):
    """List all Sales with their related Orders."""

    model = Sale
    template_name = "sales/sales_list.html"
    context_object_name = "sales"
    ordering = ["-bill_no"]

    def get_queryset(self):
        # We'll add total_weight and total_amount in get_context_data for both ornaments and raw metals
        queryset = (
            super()
            .get_queryset()
            .select_related("order")
            .prefetch_related("order__order_ornaments__ornament", "sale_metals")
        )
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(bill_no__icontains=search)
                | Q(order__sn__icontains=search)
                | Q(order__customer_name__icontains=search)
                | Q(order__phone_number__icontains=search)
            )

        return queryset.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sales_qs = context["sales"]
        context["sales_count"] = sales_qs.count() if hasattr(sales_qs, "count") else len(sales_qs)

        gold_metal = Ornament.MetalTypeCategory.GOLD
        silver_metal = Ornament.MetalTypeCategory.SILVER
        diamond_metal = Ornament.MetalTypeCategory.DIAMOND

        context["gold_sales"] = sales_qs.filter(
            Q(order__order_ornaments__ornament__metal_type__iexact=gold_metal)
            | Q(sale_metals__metal_type="gold")
        ).distinct()
        context["silver_sales"] = sales_qs.filter(
            Q(order__order_ornaments__ornament__metal_type__iexact=silver_metal)
            | Q(sale_metals__metal_type="silver")
        ).distinct()
        context["diamond_sales"] = sales_qs.filter(
            Q(order__order_ornaments__ornament__metal_type__iexact=diamond_metal)
        ).distinct()
        context["all_sales_count"] = context["sales_count"]
        context["gold_sales_count"] = context["gold_sales"].count()
        context["silver_sales_count"] = context["silver_sales"].count()
        context["diamond_sales_count"] = context["diamond_sales"].count()

        # Calculate gold/silver weights and total sales amount (including raw metals)
        purity_factors = {
            Ornament.TypeCategory.TWENTYFOURKARAT: Decimal("1"),
            Ornament.TypeCategory.TWENTHREEKARAT: Decimal("23") / Decimal("24"),
            Ornament.TypeCategory.TWENTYTWOKARAT: Decimal("22") / Decimal("24"),
            Ornament.TypeCategory.EIGHTEENKARAT: Decimal("18") / Decimal("24"),
            Ornament.TypeCategory.FOURTEENKARAT: Decimal("14") / Decimal("24"),
        }

        gold_24_weight = Decimal("0")
        silver_weight = Decimal("0")
        total_sales_amount = Decimal("0")

        # For each sale, sum ornament weights and sale_metals
        for sale in context["sales"]:
            # Ornaments
            for line in sale.order.order_ornaments.all():
                weight = line.ornament.weight or Decimal("0")
                factor = purity_factors.get(getattr(line.ornament, 'type', None), Decimal("1.00"))
                if getattr(line.ornament, 'metal_type', None) == getattr(Ornament.MetalTypeCategory, 'GOLD', 'gold'):
                    gold_24_weight += weight * factor
                elif getattr(line.ornament, 'metal_type', None) == getattr(Ornament.MetalTypeCategory, 'SILVER', 'silver'):
                    silver_weight += weight
            # Raw metals
            for metal in sale.sale_metals.all():
                if metal.metal_type == 'gold':
                    gold_24_weight += metal.quantity
                elif metal.metal_type == 'silver':
                    silver_weight += metal.quantity
                total_sales_amount += metal.line_amount or Decimal("0")
            # Add order total (ornaments + metals)
            total_sales_amount += sale.order.total or Decimal("0")

        # Patch each sale with a total_weight that includes both ornaments and raw metals

        def apply_sale_totals(queryset):
            # Patch each sale with total_weight and display_total (ornaments + raw metals)
            for sale in queryset:
                ornament_weight = sum(
                    [(line.ornament.weight or Decimal("0")) for line in sale.order.order_ornaments.all()]
                )
                metal_weight = sum(
                    [(metal.quantity or Decimal("0")) for metal in sale.sale_metals.all()]
                )
                sale.total_weight = ornament_weight + metal_weight
                sale.gold_24_weight = Decimal("0")
                sale.silver_24_weight = Decimal("0")
                for line in sale.order.order_ornaments.all():
                    metal_type = str(getattr(line.ornament, "metal_type", "")).lower()
                    if metal_type == str(gold_metal).lower():
                        weight = line.ornament.weight or Decimal("0")
                        factor = purity_factors.get(getattr(line.ornament, "type", None), Decimal("1.00"))
                        sale.gold_24_weight += weight * factor
                    elif metal_type == str(silver_metal).lower():
                        weight = line.ornament.weight or Decimal("0")
                        factor = purity_factors.get(getattr(line.ornament, "type", None), Decimal("1.00"))
                        sale.silver_24_weight += weight * factor
                metal_total = sum([(metal.line_amount or Decimal("0")) for metal in sale.sale_metals.all()])
                if sale.order.order_ornaments.count() == 0:
                    sale.display_total = metal_total
                else:
                    sale.display_total = sale.order.total or Decimal("0")

        apply_sale_totals(context["sales"])
        apply_sale_totals(context["gold_sales"])
        apply_sale_totals(context["silver_sales"])
        apply_sale_totals(context["diamond_sales"])

        def calculate_totals(queryset):
            totals = {
                "weight": Decimal("0"),
                "total": Decimal("0"),
                "paid": Decimal("0"),
                "remaining": Decimal("0"),
                "gold_24_weight": Decimal("0"),
                "silver_24_weight": Decimal("0"),
            }
            for sale in queryset:
                ornament_weight = sum(
                    [(line.ornament.weight or Decimal("0")) for line in sale.order.order_ornaments.all()]
                )
                metal_weight = sum(
                    [(metal.quantity or Decimal("0")) for metal in sale.sale_metals.all()]
                )
                total_weight = ornament_weight + metal_weight
                metal_total = sum([(metal.line_amount or Decimal("0")) for metal in sale.sale_metals.all()])
                if sale.order.order_ornaments.count() == 0:
                    display_total = metal_total
                else:
                    display_total = sale.order.total or Decimal("0")

                for line in sale.order.order_ornaments.all():
                    metal_type = str(getattr(line.ornament, "metal_type", "")).lower()
                    if metal_type == str(gold_metal).lower():
                        weight = line.ornament.weight or Decimal("0")
                        factor = purity_factors.get(getattr(line.ornament, "type", None), Decimal("1.00"))
                        totals["gold_24_weight"] += weight * factor
                    elif metal_type == str(silver_metal).lower():
                        weight = line.ornament.weight or Decimal("0")
                        factor = purity_factors.get(getattr(line.ornament, "type", None), Decimal("1.00"))
                        totals["silver_24_weight"] += weight * factor

                totals["weight"] += total_weight
                totals["total"] += display_total
                totals["paid"] += sale.order.total_paid or Decimal("0")
                totals["remaining"] += sale.order.remaining_amount or Decimal("0")

            return totals

        context.update(
            {
                "gold_24_weight": gold_24_weight,
                "silver_weight": silver_weight,
                "total_sales_amount": total_sales_amount,
                "all_totals": calculate_totals(sales_qs),
                "gold_totals": calculate_totals(context["gold_sales"]),
                "silver_totals": calculate_totals(context["silver_sales"]),
                "diamond_totals": calculate_totals(context["diamond_sales"]),
            }
        )
        return context


class SaleUpdateView(UpdateView):
    """Edit a Sale (bill number and sale date)."""

    model = Sale
    fields = ["bill_no", "sale_date", "pan_number", "address"]
    template_name = "sales/sale_form.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add the related order to the context
        order = self.object.order
        context['order'] = order
        # Ornaments for this order
        context['order_ornaments'] = order.order_ornaments.select_related('ornament').all() if order else []
        # Raw metal lines for this sale
        context['sale_metals'] = self.object.sale_metals.select_related('stock_type').all() if self.object else []
        # Pass through the return URL parameters
        context['return_month'] = self.request.GET.get('month', '')
        context['return_year'] = self.request.GET.get('year', '')
        return context

    def get_success_url(self):
        # Check if we should return to sales by month page
        month = self.request.POST.get('return_month') or self.request.GET.get('month')
        year = self.request.POST.get('return_year') or self.request.GET.get('year')
        
        if month and year:
            return f"{reverse_lazy('sales:sales_by_month')}?month={month}&year={year}"
        return reverse_lazy("sales:sales_list")


class SaleDeleteView(DeleteView):
    """Delete a Sale and return the order to the open orders list."""

    model = Sale
    template_name = "sales/sale_confirm_delete.html"
    success_url = reverse_lazy("sales:sales_list")

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        order = self.object.order
        response = super().delete(request, *args, **kwargs)

        # Reset order status so it is no longer treated as delivered
        order.status = "order"
        order.save(update_fields=["status", "updated_at"])

        return response


class DeleteSaleAndOrderView(View):
    """Delete both the Sale and its related Order.

    Also returns all ornaments on that order back to stock by clearing
    their order FK and setting ornament_type to STOCK.
    """

    def post(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk)
        order = sale.order

        if order is not None:
            # Return ornaments on this order back to stock
            Ornament.objects.filter(order=order).update(
                order=None,
                ornament_type=Ornament.OrnamentCategory.STOCK,
            )

            # Deleting the order will cascade to the Sale via FK
            order.delete()
        else:
            # Fallback: just delete the sale if order is missing
            sale.delete()

        return redirect("sales:sales_list")


class DirectSaleCreateView(CreateView):
    """Create a new Order + Sale in one step, using the same
    auto-calculation and ornament selection logic as the Order
    create page, but finalizing it directly as a Sale.
    """

    model = Order
    form_class = OrderForm
    template_name = "order/order_form.html"

    def get_initial(self):
        initial = super().get_initial()
        if "order_date" not in initial or initial["order_date"] in (None, ""):
            try:
                initial["order_date"] = ndt.date.today()
            except Exception:
                pass

        if "deliver_date" not in initial or initial["deliver_date"] in (None, ""):
            try:
                initial["deliver_date"] = ndt.date.today() + timedelta(days=7)
            except Exception:
                pass

        # Set status to "delivered" for sales create
        initial["status"] = "delivered"

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_sale_create"] = True
        context.setdefault("form_title", "Create Sale")
        context["payment_choices"] = Order.PAYMENT_CHOICES
        context["initial_payments_json"] = json.dumps([])

        # Provide defaults for sale-specific fields
        try:
            default_sale_date = ndt.date.today()
        except Exception:
            default_sale_date = ""
        context.setdefault("sale_date", default_sale_date)

        # Auto-generate the next bill_no based on the latest bill number
        from .models import Sale
        latest_sale = Sale.objects.exclude(bill_no__isnull=True).exclude(bill_no__exact='').order_by('-id').first()
        
        if latest_sale and latest_sale.bill_no:
            try:
                latest_bill_num = int(latest_sale.bill_no)
                next_bill_num = latest_bill_num + 1
                bill_no_candidate = str(next_bill_num)
            except (ValueError, TypeError):
                # If bill_no is not a number, default to 1
                bill_no_candidate = "1"
        else:
            # No sales exist, start from 1
            bill_no_candidate = "1"
        
        context.setdefault("bill_no", bill_no_candidate)

        # Add metal stock formset for adding raw metals (same as order)
        from .forms import SalesMetalStockFormSet
        if self.request.POST:
            instance = getattr(self, 'object', None)
            context['metal_stock_formset'] = SalesMetalStockFormSet(self.request.POST, instance=instance)
        else:
            context['metal_stock_formset'] = SalesMetalStockFormSet(instance=None)
        context['metal_formset_prefix'] = context['metal_stock_formset'].prefix

        return context

    def get_success_url(self):
        return reverse_lazy("sales:sales_list")

    def form_valid(self, form):
        # Save the order first
        self.object = form.save()

        # Create the Sale for this order using sales-specific fields from the form when provided.
        raw_sale_date = self.request.POST.get("sale_date") or None
        bill_no = self.request.POST.get("bill_no") or str(self.object.sn)

        if raw_sale_date:
            sale_date = raw_sale_date
        else:
            sale_date = self.object.deliver_date or self.object.order_date
            if sale_date is None:
                try:
                    sale_date = ndt.date.today()
                except Exception:
                    sale_date = None

        sale_obj = Sale.objects.create(
            order=self.object,
            sale_date=sale_date,
            bill_no=bill_no,
            pan_number=self.object.pan_number,
            address=self.object.address,
        )

        # --- Save SalesMetalStockFormSet (raw metal details) and update MetalStock ---
        from .forms import SalesMetalStockFormSet
        from goldsilverpurchase.models import MetalStock, MetalStockMovement
        metal_stock_formset = SalesMetalStockFormSet(self.request.POST, instance=sale_obj)
        if metal_stock_formset.is_valid():
            sale_metals = metal_stock_formset.save()
            # Deduct from MetalStock and create MetalStockMovement for each raw metal sold
            for metal in sale_obj.sale_metals.all():
                try:
                    metal_stock = MetalStock.objects.get(
                        metal_type=metal.metal_type,
                        purity=metal.purity,
                        stock_type=metal.stock_type
                    )
                    # Deduct quantity
                    metal_stock.quantity -= metal.quantity
                    metal_stock.save()
                    # Create movement record
                    MetalStockMovement.objects.create(
                        metal_stock=metal_stock,
                        movement_type='out',
                        quantity=metal.quantity,
                        rate=metal.rate_per_gram,
                        reference_type='Sale',
                        reference_id=f"Sale-{sale_obj.pk}",
                        notes=f"Raw metal sold in sale {sale_obj.pk} (Order {self.object.sn})"
                    )
                except MetalStock.DoesNotExist:
                    # Optionally, add a warning message
                    messages.warning(self.request, f"No matching metal stock found for {metal.get_metal_type_display()} ({metal.purity}, {metal.stock_type}). Please add to inventory first.")
        # ---

        # Create per-line OrderOrnament entries from JSON payload
        order_lines_raw = form.cleaned_data.get("order_lines_json") or "[]"
        try:
            lines = json.loads(order_lines_raw)
        except (TypeError, ValueError):
            lines = []

        for line in lines:
            ornament_id = line.get("ornament_id")
            if not ornament_id:
                continue
            try:
                ornament = Ornament.objects.get(pk=ornament_id)
            except Ornament.DoesNotExist:
                continue

            OrderOrnament.objects.create(
                order=self.object,
                ornament=ornament,
                gold_rate=Decimal(str(line.get("gold_rate", 0) or 0)),
                diamond_rate=Decimal(str(line.get("diamond_rate", 0) or 0)),
                zircon_rate=Decimal(str(line.get("zircon_rate", 0) or 0)),
                stone_rate=Decimal(str(line.get("stone_rate", 0) or 0)),
                jarti=Decimal(str(line.get("jarti", 0) or 0)),
                jyala=Decimal(str(line.get("jyala", 0) or 0)),
                line_amount=Decimal(str(line.get("line_amount", 0) or 0)),
            )

            # Also mark ornament as belonging to this order and flag as ORDER type
            ornament.order = self.object
            try:
                ornament.ornament_type = Ornament.OrnamentCategory.SALES
            except AttributeError:
                ornament.ornament_type = "sales"
            ornament.save()

        # Persist payment breakdown for the sale
        payment_lines_raw = form.cleaned_data.get("payment_lines_json") or "[]"
        try:
            payments_data = json.loads(payment_lines_raw)
        except (TypeError, ValueError):
            payments_data = []

        self.object.payments.all().delete()
        for payment in payments_data:
            amount = Decimal(str(payment.get("amount", 0) or 0))
            if amount <= 0:
                continue
            mode = payment.get("payment_mode") or payment.get("mode") or "cash"
            if mode not in dict(Order.PAYMENT_CHOICES):
                mode = "cash"
            OrderPayment.objects.create(
                order=self.object,
                payment_mode=mode,
                amount=amount,
            )

        # Recompute order totals from created lines (amount/subtotal/total/remaining)
        self.object.recompute_totals_from_lines()

        # Mark order as delivered for direct sales
        self.object.status = "delivered"
        self.object.save(update_fields=["status", "updated_at"])

        return redirect(self.get_success_url())


def sales_print_view(request):
    """Printable view of sales (with related orders)."""

    view = SalesListView()
    view.request = request
    sales = view.get_queryset()

    return render(request, "sales/print_view.html", {"sales": sales})


def sales_export_excel(request):
    """Export sales to Excel, similar to ornaments/purchases export."""

    view = SalesListView()
    view.request = request
    sales = view.get_queryset()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    headers = [
        "Bill No",
        "Sale Date (BS)",
        "Order No",
        "Order Date (BS)",
        "Customer",
        "Phone",
        "Status",
        "Amount",
        "Discount",
        "Subtotal",
        "Tax",
        "Total",
        "Payment Modes",
        "Paid Amount",
        "Remaining Amount",
    ]
    ws.append(headers)

    for s in sales:
        o = s.order
        ws.append([
            s.bill_no,
            str(s.sale_date) if s.sale_date else "",
            o.sn,
            str(o.order_date) if o.order_date else "",
            o.customer_name,
            o.phone_number,
            o.get_status_display(),
            o.amount,
            o.discount,
            o.subtotal,
            o.tax,
            o.total,
            ", ".join([p.get_payment_mode_display() for p in o.payments.all()]) if o.payments.exists() else "",
            o.total_paid,
            o.remaining_amount,
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="sales.xlsx"'

    return response


def sales_monthly_tax_report(request):
    """Export monthly sales tax report to Excel."""

    month = request.GET.get("month")
    year = request.GET.get("year")

    if not month or not year:
        messages.error(request, "Please select a month and year to generate the tax report.")
        return redirect("sales:sales_by_month")

    try:
        month_int = int(month)
        year_int = int(year)

        start_date = ndt.date(year_int, month_int, 1)
        if month_int == 12:
            end_date = ndt.date(year_int + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = ndt.date(year_int, month_int + 1, 1) - timedelta(days=1)
    except (ValueError, TypeError):
        messages.error(request, "Invalid month or year provided for the tax report.")
        return redirect("sales:sales_by_month")

    sales = (
        Sale.objects.select_related("order")
        .filter(
            sale_date__gte=start_date,
            sale_date__lte=end_date,
            order__tax__gt=0,
        )
        .order_by("bill_no")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Tax Report"

    headers = ["Bill No", "Customer Name", "PAN", "Address", "Total Amount", "Tax"]
    ws.append(headers)

    for sale in sales:
        order = sale.order
        ws.append([
            sale.bill_no,
            order.customer_name,
            order.pan_number or "",
            order.address or "",
            order.total,
            order.tax,
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="monthly_tax_report_{year_int}_{month_int:02d}.xlsx"'
    )

    return response


def sales_import_excel(request):
    """Import sales from Excel in detailed format.

    Expected columns (in order):
    SN, Order Date, Purchase Date, Customer name, PAN, Address, Taxable sales, Phone no,
    Ornament, Metal Type, Type, Weight, Jarti, Total weight, Own gold, Rate, Jyala, Stones,
    Total, Discount, Tax, All total, Payment, Status, Kaligar
    """

    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("sales:import_excel")

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            imported = 0
            skipped = 0
            errors = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                try:
                    # Parse columns based on the format: SN, Order Date, Purchase Date, Customer name, PAN, Address, Taxable sales, Phone no, Ornament, Metal Type, Type, Weight, Jarti, Total weight, Own gold, Rate, Jyala, Stones, Total, Discount, Tax, All total, Payment, Status, Kaligar
                    sn = row[0] if len(row) > 0 else None
                    order_date = row[1] if len(row) > 1 else None
                    purchase_date = row[2] if len(row) > 2 else None  # This is actually sale_date/deliver_date
                    customer_name = row[3] if len(row) > 3 else None
                    pan = row[4] if len(row) > 4 else None
                    address = row[5] if len(row) > 5 else None
                    taxable_sales = row[6] if len(row) > 6 else None
                    phone_no = row[7] if len(row) > 7 else None
                    ornament_name = row[8] if len(row) > 8 else None
                    metal_type = row[9] if len(row) > 9 else None
                    purity_type = row[10] if len(row) > 10 else None
                    weight = row[11] if len(row) > 11 else None
                    jarti = row[12] if len(row) > 12 else None
                    total_weight = row[13] if len(row) > 13 else None
                    own_gold = row[14] if len(row) > 14 else None
                    rate = row[15] if len(row) > 15 else None
                    jyala = row[16] if len(row) > 16 else None
                    stones = row[17] if len(row) > 17 else None
                    ornament_total = row[18] if len(row) > 18 else None
                    discount = row[19] if len(row) > 19 else None
                    tax = row[20] if len(row) > 20 else None
                    all_total = row[21] if len(row) > 21 else None
                    payment_mode = row[22] if len(row) > 22 else None
                    status = row[23] if len(row) > 23 else None
                    kaligar_name = row[24] if len(row) > 24 else None

                    if not customer_name or not ornament_name:
                        skipped += 1
                        continue

                    # Create or get customer
                    from goldsilverpurchase.models import Customer
                    customer, _ = Customer.objects.get_or_create(
                        name=str(customer_name).strip(),
                        defaults={
                            'phone_number': str(phone_no or '').strip(),
                            'address': str(address or '').strip(),
                            'pan_no': str(pan or '').strip(),
                        }
                    )

                    # Create or get ornament
                    from ornament.models import Ornament, Kaligar
                    
                    # Get or create Kaligar
                    kaligar = None
                    if kaligar_name:
                        kaligar, _ = Kaligar.objects.get_or_create(
                            name=str(kaligar_name).strip()
                        )

                    # Parse metal type and purity
                    metal_type_str = str(metal_type or 'Gold').strip()
                    purity_str = str(purity_type or '24K').strip()
                    
                    ornament, _ = Ornament.objects.get_or_create(
                        name=str(ornament_name).strip(),
                        metal_type=metal_type_str,
                        defaults={
                            'purity_type': purity_str,
                            'weight': Decimal(str(weight or 0)),
                            'kaligar': kaligar,
                        }
                    )

                    # Create Order
                    from order.models import Order, OrderOrnament
                    
                    order = Order.objects.create(
                        customer=customer,
                        order_date=str(order_date or '').strip() if order_date else None,
                        deliver_date=str(purchase_date or '').strip() if purchase_date else None,
                        total=Decimal(str(all_total or 0)),
                        discount=Decimal(str(discount or 0)),
                        taxable_amount=Decimal(str(taxable_sales or 0)),
                        tax=Decimal(str(tax or 0)),
                        status=str(status or 'pending').strip().lower(),
                    )

                    # Create OrderOrnament
                    OrderOrnament.objects.create(
                        order=order,
                        ornament=ornament,
                        quantity=1,
                        customer_jarti=Decimal(str(jarti or 0)),
                        ornament_jarti=Decimal(str(weight or 0)),
                        jyala=Decimal(str(jyala or 0)),
                        stone_amount=Decimal(str(stones or 0)),
                        rate=Decimal(str(rate or 0)),
                        own_gold=Decimal(str(own_gold or 0)),
                        own_silver=Decimal('0'),
                    )

                    # Create Payment if payment mode is specified
                    if payment_mode and str(payment_mode).strip().lower() != 'none':
                        from order.models import OrderPayment
                        payment_method = str(payment_mode).strip().lower()
                        if payment_method in ['cash', 'fonepay', 'bank', 'gold', 'silver', 'sundry_debtor']:
                            OrderPayment.objects.create(
                                order=order,
                                amount=Decimal(str(all_total or 0)),
                                payment_method=payment_method,
                                payment_date=str(purchase_date or '').strip() if purchase_date else None,
                            )

                    # Create Sale
                    Sale.objects.create(
                        order=order,
                        bill_no=str(sn or '').strip(),
                        sale_date=str(purchase_date or '').strip() if purchase_date else None,
                    )

                    imported += 1

                except Exception as exc:
                    errors.append(f"Row {row_num}: {str(exc)}")
                    skipped += 1
                    continue

            if errors and len(errors) <= 10:
                for error in errors:
                    messages.warning(request, error)

            messages.success(
                request,
                f"Imported {imported} sales, skipped {skipped} rows. {len(errors)} errors.",
            )
            return redirect("sales:sales_list")

        except Exception as exc:  # noqa: BLE001
            messages.error(request, f"Failed to import Excel: {exc}")
            return redirect("sales:import_excel")

    return render(request, "sales/import_excel.html")


def download_import_template(request):
    """Download Excel template for bulk import."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Import"
    
    # Define columns matching the user's format
    columns = [
        'SN',
        'Order Date',
        'Purchase Date',
        'Customer name',
        'PAN',
        'ठेगाना',
        'Taxable sales',
        'Phone no',
        'Ornament',
        'Metal Type',
        'Type',
        'Weight',
        'Jarti',
        'Total weight',
        'Own gold (tola)',
        'Rate (tola)',
        'Jyala',
        'Stones',
        'Total',
        'Discount',
        'Tax',
        'All total',
        'Payment',
        'Status',
        'Kaligar'
    ]
    
    # Add headers
    ws.append(columns)
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data rows
    sample_data = [
        ['207', '2082-09-01', '2082-09-01', 'Sandhya Nagarkoti', 'Gundu', 'Kathmandu', '4710', '9841234567', 'Diamond Nosepin', 'Diamond', '14KARAT', '0.08', '0', '0.08', '0', '15230', '702', '2960', '4710', '0', '94.2', '4807.2', 'Fonepay', 'Complete', 'Ritu Gems'],
        ['208', '2082-09-02', '2082-09-02', 'Ram Prasad', '', 'Pokhara', '8500', '9845678901', 'Gold Ring', 'Gold', '22K', '5.5', '500', '6', '0', '60000', '1500', '5000', '8500', '1000', '170', '8670', 'cash', 'delivered', 'Ramesh'],
        ['209', '2082-09-03', '2082-09-03', 'Sita Devi', 'PAN12345', 'Lalitpur', '12000', '9801234567', 'Silver Necklace', 'Silver', '92.5', '25', '0', '25', '0', '900', '0', '0', '12000', '500', '240', '12240', 'bank', 'pending', 'Ashok'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 20)
    
    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename="Sales_Import_Template.xlsx"'
    
    return response


class ImportWizardStepOneView(TemplateView):
    """Step 1: Welcome and download template."""
    
    template_name = 'sales/import_wizard_step1.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = 1
        return context


class ImportWizardStepTwoView(FormView):
    """Step 2: Upload and preview Excel file."""
    
    form_class = ExcelImportForm
    template_name = 'sales/import_wizard_step2.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = 2
        return context
    
    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        
        try:
            # Read and parse the file
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Get headers
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
            
            # Parse data rows (limit to first 10 for preview)
            preview_data = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row) or row_num > 11:  # First 10 data rows
                    continue
                preview_data.append({
                    'row_num': row_num,
                    'data': row[:len(headers)]
                })
            
            # Store file temporarily in session
            import tempfile
            import os
            
            # Save file to temp directory
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, f'import_{self.request.user.id or "anonymous"}_{excel_file.name}')
            
            with open(temp_file, 'wb') as f:
                for chunk in excel_file.chunks():
                    f.write(chunk)
            
            # Store path in session
            self.request.session['import_file_path'] = temp_file
            self.request.session['import_headers'] = headers
            
            # Store in context
            context = self.get_context_data()
            context['headers'] = headers
            context['preview_data'] = preview_data
            context['row_count'] = ws.max_row - 1  # Exclude header
            context['step'] = 2
            
            return render(self.request, self.template_name, context)
        
        except Exception as e:
            messages.error(self.request, f'Failed to read Excel file: {str(e)}')
            return redirect('sales:import_wizard_step2')


class ImportWizardStepThreeView(View):
    """Step 3: Confirm and import."""
    
    def get(self, request):
        """Show confirmation page."""
        file_path = request.session.get('import_file_path')
        headers = request.session.get('import_headers')
        
        if not file_path or not headers:
            messages.error(request, 'Session expired. Please start import again.')
            return redirect('sales:import_wizard_step1')
        
        context = {
            'step': 3,
            'headers': headers,
            'file_path': file_path,
        }
        
        return render(request, 'sales/import_wizard_step3.html', context)
    
    def post(self, request):
        """Execute the import."""
        file_path = request.session.get('import_file_path')
        
        if not file_path:
            messages.error(request, 'Session expired. Please start import again.')
            return redirect('sales:import_wizard_step1')
        
        try:
            # Process the file
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Parse headers
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
            
            # Map columns
            col_map = self._map_columns(headers)
            
            # Group orders by Bill No
            orders_data = {}
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    row_data = {}
                    for header, col_idx in col_map.items():
                        if col_idx is not None and col_idx < len(row):
                            cell_value = row[col_idx].value
                            row_data[header] = cell_value
                    
                    if not any(row_data.values()):
                        continue
                    
                    # Get bill_no, use auto-generated if empty
                    bill_no = row_data.get('bill_no') or f'AUTO_{row_num}'
                    if bill_no not in orders_data:
                        orders_data[bill_no] = {'order_info': {}, 'ornaments': []}
                    
                    if not orders_data[bill_no]['order_info']:
                        orders_data[bill_no]['order_info'] = {
                            'customer_name': row_data.get('customer_name', ''),
                            'phone_number': row_data.get('phone_number', ''),
                            'order_date': row_data.get('order_date'),
                            'deliver_date': row_data.get('deliver_date'),
                            'discount': self._parse_decimal(row_data.get('discount', 0)),
                            'tax': self._parse_decimal(row_data.get('tax', 0)),
                            'payment_mode': row_data.get('payment_mode', 'cash'),
                            'status': row_data.get('status', 'delivered'),
                        }
                    
                    ornament_data = {
                        'ornament_name': row_data.get('ornament_name', ''),
                        'ornament_type': row_data.get('ornament_type', ''),
                        'metal_type': row_data.get('metal_type', ''),
                        'purity_type': row_data.get('purity_type', 'twentyfourkarat'),
                        'weight': self._parse_decimal(row_data.get('weight', 0)),
                        'gold_rate': self._parse_decimal(row_data.get('rate', 0)),
                        'jarti': self._parse_decimal(row_data.get('jarti', 0)),
                        'jyala': self._parse_decimal(row_data.get('jyala', 0)),
                        'stone_rate': self._parse_decimal(row_data.get('stones', 0)),
                        'line_amount': self._parse_decimal(row_data.get('total', 0)),
                        'kaligar': row_data.get('kaligar', ''),
                    }
                    orders_data[bill_no]['ornaments'].append(ornament_data)
                
                except Exception as e:
                    messages.warning(request, f'Row {row_num}: Skipped - {str(e)}')
            
            # Create orders
            imported_count = 0
            errors = []
            
            importer = ExcelImportProcessor(request)
            for bill_no, order_data in orders_data.items():
                try:
                    result = importer.create_complete_order(
                        bill_no,
                        order_data['order_info'],
                        order_data['ornaments']
                    )
                    if result['success']:
                        imported_count += 1
                    else:
                        if result.get('error'):
                            errors.append(f"Bill {bill_no}: {result['error']}")
                except Exception as e:
                    errors.append(f"Bill {bill_no}: {str(e)}")
            
            # Clean up
            import os
            if os.path.exists(file_path):
                os.remove(file_path)
            
            del request.session['import_file_path']
            del request.session['import_headers']
            
            # Show results
            if imported_count > 0:
                messages.success(
                    request,
                    f'✅ Successfully imported {imported_count} complete orders with all ornaments!'
                )
            
            if errors:
                for error in errors[:10]:
                    messages.warning(request, error)
            
            return redirect('sales:sales_list')
        
        except Exception as e:
            messages.error(request, f'Failed to import: {str(e)}')
            return redirect('sales:import_wizard_step1')
    
    def _map_columns(self, headers):
        """Map columns from headers."""
        processor = ExcelImportProcessor(self.request)
        return processor.map_columns(headers)
    
    def _parse_decimal(self, value):
        """Parse decimal value."""
        processor = ExcelImportProcessor(self.request)
        return processor.parse_decimal(value)


class ExcelImportProcessor:
    """Helper class for Excel import processing."""
    
    def __init__(self, request):
        self.request = request
    
    def map_columns(self, headers):
        """Map Excel headers to model fields."""
        col_map = {
            'bill_no': None,
            'order_date': None,
            'deliver_date': None,
            'customer_name': None,
            'phone_number': None,
            'ornament_name': None,
            'ornament_type': None,
            'metal_type': None,
            'purity_type': None,
            'weight': None,
            'total': None,
            'discount': None,
            'tax': None,
            'payment_mode': None,
            'status': None,
            'rate': None,
            'jarti': None,
            'jyala': None,
            'stones': None,
            'kaligar': None,
        }
        
        for idx, header in enumerate(headers):
            header_lower = str(header or '').lower().strip()
            
            if 'bill' in header_lower:
                col_map['bill_no'] = idx
            elif 'order' in header_lower and 'date' in header_lower and 'deliver' not in header_lower:
                col_map['order_date'] = idx
            elif 'deliver' in header_lower and 'date' in header_lower:
                col_map['deliver_date'] = idx
            elif 'customer' in header_lower or (header_lower == 'name' and col_map['customer_name'] is None):
                col_map['customer_name'] = idx
            elif 'phone' in header_lower or 'contact' in header_lower or 'tel' in header_lower:
                col_map['phone_number'] = idx
            elif 'ornament' in header_lower and 'name' in header_lower:
                col_map['ornament_name'] = idx
            elif 'ornament' in header_lower and 'type' in header_lower:
                col_map['ornament_type'] = idx
            elif 'metal' in header_lower and 'type' in header_lower:
                col_map['metal_type'] = idx
            elif 'purity' in header_lower or 'karat' in header_lower or 'carat' in header_lower:
                col_map['purity_type'] = idx
            elif any(w in header_lower for w in ('weight', 'total weight', 'wt', 'gm')):
                col_map['weight'] = idx
            elif any(w in header_lower for w in ('total', 'all total', 'final total', 'sum', 'grand total')):
                col_map['total'] = idx
            elif 'discount' in header_lower or 'deduction' in header_lower:
                col_map['discount'] = idx
            elif 'tax' in header_lower or 'vat' in header_lower:
                col_map['tax'] = idx
            elif 'payment' in header_lower and 'mode' in header_lower:
                col_map['payment_mode'] = idx
            elif 'status' in header_lower or 'state' in header_lower:
                col_map['status'] = idx
            elif 'rate' in header_lower and col_map['rate'] is None:
                col_map['rate'] = idx
            elif 'jarti' in header_lower or 'jatti' in header_lower:
                col_map['jarti'] = idx
            elif 'jyala' in header_lower or 'jala' in header_lower:
                col_map['jyala'] = idx
            elif 'stone' in header_lower or 'zircon' in header_lower or 'diamond' in header_lower:
                col_map['stones'] = idx
            elif 'kaligar' in header_lower or 'kalegar' in header_lower:
                col_map['kaligar'] = idx
        
        return col_map
    
    def parse_decimal(self, value):
        """Parse decimal from various formats."""
        if value is None or value == '':
            return Decimal('0')
        
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            
            if isinstance(value, str):
                cleaned = value.strip().replace('Rs', '').replace('$', '').replace(',', '').strip()
                return Decimal(cleaned) if cleaned else Decimal('0')
            
            return Decimal(value)
        except:
            return Decimal('0')
    
    def create_complete_order(self, bill_no, order_info, ornaments_list):
        """Create complete order with all ornaments."""
        try:
            customer_name = str(order_info.get('customer_name') or '').strip()
            phone_number = str(order_info.get('phone_number') or '').strip()
            
            if not customer_name:
                return {'success': False, 'error': 'Customer name is required'}
            
            # Parse dates
            order_date = self._parse_nepali_date(order_info.get('order_date'))
            deliver_date = self._parse_nepali_date(order_info.get('deliver_date'))
            
            if not order_date:
                try:
                    order_date = ndt.date.today()
                except:
                    pass
            
            if not deliver_date and order_date:
                try:
                    deliver_date = order_date + timedelta(days=7)
                except:
                    deliver_date = order_date
            
            # Create order
            order = Order.objects.create(
                customer_name=customer_name,
                phone_number=phone_number if phone_number and len(phone_number) >= 7 else '1234567',
                order_date=order_date,
                deliver_date=deliver_date,
                status=order_info.get('status', 'delivered'),
                discount=order_info.get('discount', Decimal('0')),
                tax=order_info.get('tax', Decimal('0')),
            )
            
            # Create ornaments
            total_line_amount = Decimal('0')
            
            for ornament_data in ornaments_list:
                if not (ornament_data.get('weight') or ornament_data.get('line_amount')):
                    continue
                
                ornament_name = ornament_data.get('ornament_name') or (
                    f"{ornament_data.get('ornament_type', 'Item')} - "
                    f"{ornament_data.get('metal_type', 'Gold')}"
                )
                
                purity_type = self._get_purity_type(ornament_data.get('purity_type', 'twentyfourkarat'))
                metal_type = self._get_metal_type(ornament_data.get('metal_type'))
                
                # Get or create kaligar
                kaligar_name = ornament_data.get('kaligar', 'Default')
                kaligar, _ = Kaligar.objects.get_or_create(
                    name=kaligar_name,
                    defaults={'panno': '000000000'}  # Default PAN as placeholder
                )
                
                ornament, _ = Ornament.objects.get_or_create(
                    ornament_name=ornament_name,
                    metal_type=metal_type,
                    type=purity_type,
                    defaults={
                        'weight': ornament_data.get('weight', Decimal('0')),
                        'ornament_type': 'sales',
                        'kaligar': kaligar,
                    }
                )
                
                line_amount = ornament_data.get('line_amount', Decimal('0'))
                total_line_amount += line_amount
                
                OrderOrnament.objects.create(
                    order=order,
                    ornament=ornament,
                    gold_rate=ornament_data.get('gold_rate', Decimal('0')),
                    jarti=ornament_data.get('jarti', Decimal('0')),
                    jyala=ornament_data.get('jyala', Decimal('0')),
                    stone_rate=ornament_data.get('stone_rate', Decimal('0')),
                    line_amount=line_amount,
                )
                
                ornament.order = order
                ornament.ornament_type = 'sales'
                ornament.save(update_fields=['order', 'ornament_type', 'updated_at'])
            
            # Calculate order totals first
            payment_mode = order_info.get('payment_mode', 'cash')
            discount = order_info.get('discount', Decimal('0'))
            tax = order_info.get('tax', Decimal('0'))
            
            total_amount = total_line_amount if total_line_amount > 0 else Decimal('0')
            
            order.amount = total_line_amount or total_amount
            order.subtotal = max(Decimal('0'), order.amount - discount)
            order.total = order.subtotal + tax
            paid_amount = min(total_amount, order.total)
            order.remaining_amount = max(Decimal('0'), order.total - paid_amount)
            order.save(update_fields=[
                'amount', 'subtotal', 'total', 'remaining_amount'
            ])
            
            # Create payment
            if paid_amount > 0:
                OrderPayment.objects.create(
                    order=order,
                    payment_mode=payment_mode,
                    amount=paid_amount,
                )
                
                if payment_mode == 'sundry_debtor':
                    debtor, created = SundryDebtor.objects.get_or_create(
                        name=customer_name,
                        defaults={
                            'credit_limit': paid_amount,
                            'current_balance': paid_amount,
                        }
                    )
                    
                    if not created:
                        debtor.current_balance = debtor.current_balance + paid_amount
                        debtor.save(update_fields=['current_balance', 'updated_at'])
                    
                    payment = order.payments.first()
                    if payment:
                        DebtorPayment.objects.create(
                            order_payment=payment,
                            debtor=debtor,
                        )
            
            # Create Sale
            sale_date = deliver_date or order_date
            
            Sale.objects.create(
                order=order,
                sale_date=sale_date,
                bill_no=bill_no,
            )
            
            return {'success': True, 'order_id': order.sn}
        
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def _parse_nepali_date(self, date_value):
        """Parse Nepali date."""
        if not date_value:
            return None
        
        try:
            if isinstance(date_value, str):
                date_str = str(date_value).strip()
                
                if len(date_str) == 10 and date_str[4] == '-' and date_str[7] == '-':
                    parts = date_str.split('-')
                    if all(p.isdigit() for p in parts):
                        year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                        try:
                            return ndt.date(year, month, day)
                        except:
                            pass
            
            if hasattr(date_value, 'year'):
                return date_value
        
        except Exception:
            pass
        
        return None
    
    def _get_metal_type(self, metal_type_str):
        """Convert metal type string."""
        if not metal_type_str:
            return 'gold'
        
        metal_lower = str(metal_type_str).lower().strip()
        if 'silver' in metal_lower:
            return 'silver'
        elif 'platinum' in metal_lower:
            return 'platinum'
        else:
            return 'gold'
    
    def _get_purity_type(self, purity_str):
        """Convert purity string."""
        if not purity_str:
            return 'twentyfourkarat'
        
        purity_lower = str(purity_str).lower().strip().replace(' ', '').replace('k', '')
        
        if '24' in purity_lower or purity_lower == 'twentyfourkarat':
            return 'twentyfourkarat'
        elif '23' in purity_lower or purity_lower == 'twenthreekarat':
            return 'twenthreekarat'
        elif '22' in purity_lower or purity_lower == 'twentytwokarat':
            return 'twentytwokarat'
        elif '18' in purity_lower or purity_lower == 'eighteenkarat':
            return 'eighteenkarat'
        elif '14' in purity_lower or purity_lower == 'fourteenkarat':
            return 'fourteenkarat'
        else:
            return 'twentyfourkarat'


class SalesByMonthView(ListView):
    """List sales filtered by a selected month and year."""
    
    model = Sale
    template_name = "sales/sales_by_month.html"
    context_object_name = "sales"
    ordering = ["bill_no"]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related("order").prefetch_related(
            "order__order_ornaments__ornament", "sale_metals"
        )
        
        # Get month and year from query params, default to current month/year
        month = self.request.GET.get('month')
        year = self.request.GET.get('year')
        
        # If no params provided, use current Nepali date
        if not month or not year:
            try:
                current_nepali_date = ndt.date.today()
                month = month or str(current_nepali_date.month)
                year = year or str(current_nepali_date.year)
            except Exception:
                month = month or '1'
                year = year or '2081'
        
        try:
            month_int = int(month)
            year_int = int(year)
            
            # Create start and end dates for the month
            start_date = ndt.date(year_int, month_int, 1)
            
            # Calculate last day of the month
            if month_int == 12:
                end_date = ndt.date(year_int + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = ndt.date(year_int, month_int + 1, 1) - timedelta(days=1)
            
            # Filter sales by date range
            queryset = queryset.filter(
                sale_date__gte=start_date,
                sale_date__lte=end_date
            )
        except (ValueError, TypeError):
            pass
        
        return queryset.order_by("bill_no")
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get current Nepali date for default selection
        try:
            current_nepali_date = ndt.date.today()
            current_month = current_nepali_date.month
            current_year = current_nepali_date.year
        except Exception:
            current_month = 1
            current_year = 2081
        
        context['current_month'] = current_month
        context['current_year'] = current_year
        
        # Get current filter values, default to current month/year
        selected_month = self.request.GET.get('month', str(current_month))
        selected_year = self.request.GET.get('year', str(current_year))
        
        context['selected_month'] = selected_month
        context['selected_year'] = selected_year
        
        # Nepali month names
        context['nepali_months'] = [
            {'value': 1, 'name': 'Baishakh'},
            {'value': 2, 'name': 'Jestha'},
            {'value': 3, 'name': 'Ashadh'},
            {'value': 4, 'name': 'Shrawan'},
            {'value': 5, 'name': 'Bhadra'},
            {'value': 6, 'name': 'Ashwin'},
            {'value': 7, 'name': 'Kartik'},
            {'value': 8, 'name': 'Mangsir'},
            {'value': 9, 'name': 'Poush'},
            {'value': 10, 'name': 'Magh'},
            {'value': 11, 'name': 'Falgun'},
            {'value': 12, 'name': 'Chaitra'},
        ]
        
        # Calculate statistics for filtered sales
        if context['sales']:
            purity_factors = {
                Ornament.TypeCategory.TWENTYFOURKARAT: Decimal("1.00"),
                Ornament.TypeCategory.TWENTHREEKARAT: Decimal("0.99"),
                Ornament.TypeCategory.TWENTYTWOKARAT: Decimal("0.98"),
                Ornament.TypeCategory.EIGHTEENKARAT: Decimal("0.75"),
                Ornament.TypeCategory.FOURTEENKARAT: Decimal("0.58"),
            }
            
            gold_24_weight = Decimal("0")
            silver_weight = Decimal("0")
            total_sales_amount = Decimal("0")
            total_remaining_amount = Decimal("0")
            total_profit = Decimal("0")
            total_tax = Decimal("0")
            
            # Payment method breakdown
            payment_methods = {
                'cash': Decimal("0"),
                'fonepay': Decimal("0"),
                'bank': Decimal("0"),
                'gold': Decimal("0"),
                'silver': Decimal("0"),
                'sundry_debtor': Decimal("0"),
            }
            
            for sale in context["sales"]:
                # Ornaments
                for line in sale.order.order_ornaments.all():
                    weight = line.ornament.weight or Decimal("0")
                    factor = purity_factors.get(getattr(line.ornament, 'type', None), Decimal("1.00"))
                    if getattr(line.ornament, 'metal_type', None) == getattr(Ornament.MetalTypeCategory, 'GOLD', 'gold'):
                        gold_24_weight += weight * factor
                    elif getattr(line.ornament, 'metal_type', None) == getattr(Ornament.MetalTypeCategory, 'SILVER', 'silver'):
                        silver_weight += weight
                    
                    # Calculate profit: (customer_jarti - ornament_jarti) / 11.664 * rate + jyala
                    customer_jarti = line.jarti or Decimal("0")
                    ornament_jarti = line.ornament.jarti or Decimal("0")
                    rate = line.gold_rate or Decimal("0")
                    jyala = line.jyala or Decimal("0")
                    
                    jarti_difference = customer_jarti - ornament_jarti
                    profit = (jarti_difference / Decimal("11.664") * rate) + jyala
                    total_profit += profit
                
                # Raw metals
                for metal in sale.sale_metals.all():
                    if metal.metal_type == 'gold':
                        gold_24_weight += metal.quantity
                    elif metal.metal_type == 'silver':
                        silver_weight += metal.quantity
                
                # Add order total, remaining amount, and tax
                total_sales_amount += sale.order.total or Decimal("0")
                total_remaining_amount += sale.order.remaining_amount or Decimal("0")
                total_tax += sale.order.tax or Decimal("0")
                
                # Collect payment method totals
                for payment in sale.order.payments.all():
                    payment_mode = payment.payment_mode
                    if payment_mode in payment_methods:
                        payment_methods[payment_mode] += payment.amount or Decimal("0")
            
            # Patch each sale with total_weight, display_total, and profit
            for sale in context["sales"]:
                ornament_weight = sum([(line.ornament.weight or Decimal("0")) for line in sale.order.order_ornaments.all()])
                metal_weight = sum([(metal.quantity or Decimal("0")) for metal in sale.sale_metals.all()])
                sale.total_weight = ornament_weight + metal_weight
                
                metal_total = sum([(metal.line_amount or Decimal("0")) for metal in sale.sale_metals.all()])
                if sale.order.order_ornaments.count() == 0:
                    sale.display_total = metal_total
                else:
                    sale.display_total = sale.order.total or Decimal("0")
                
                # Calculate profit for this sale
                sale_profit = Decimal("0")
                for line in sale.order.order_ornaments.all():
                    customer_jarti = line.jarti or Decimal("0")
                    ornament_jarti = line.ornament.jarti or Decimal("0")
                    rate = line.gold_rate or Decimal("0")
                    jyala = line.jyala or Decimal("0")
                    
                    jarti_difference = customer_jarti - ornament_jarti
                    profit = (jarti_difference / Decimal("11.664") * rate) + jyala
                    sale_profit += profit
                
                sale.profit = sale_profit
            
            context['gold_24_weight'] = gold_24_weight
            context['silver_weight'] = silver_weight
            context['total_sales_amount'] = total_sales_amount
            context['total_remaining_amount'] = total_remaining_amount
            context['total_profit'] = total_profit
            context['total_tax'] = total_tax
            context['payment_methods'] = payment_methods
            context['sales_count'] = len(context['sales'])
        else:
            context['gold_24_weight'] = Decimal("0")
            context['silver_weight'] = Decimal("0")
            context['total_sales_amount'] = Decimal("0")
            context['total_remaining_amount'] = Decimal("0")
            context['total_profit'] = Decimal("0")
            context['total_tax'] = Decimal("0")
            context['payment_methods'] = {
                'cash': Decimal("0"),
                'fonepay': Decimal("0"),
                'bank': Decimal("0"),
                'gold': Decimal("0"),
                'silver': Decimal("0"),
                'sundry_debtor': Decimal("0"),
            }
            context['sales_count'] = 0
        
        return context
